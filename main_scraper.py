import logging
import traceback
import os
import json
import re
from datetime import date
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from ai import OllamaAssistant
from jobs_scraper import scrape_all_jobs

# --------------------------- Config ---------------------------

required_columns = [
    "AI_recommendation",
    "company",
    "title",
    "link",
    "years_required",
    "role_level",
    "skills_match_pct",
    "matched_skills",
    "missing_skills",
    "reasoning",
    "description",
    "posted_date",
]

pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
logging.basicConfig(level=logging.INFO)

excel_file = "jobs.xlsx"

PERSONAL_JOB_FINDER_PROMPT = """
You are a ruthless but fair IT job screener evaluating whether a candidate should apply to a job posting.

-----------------------------------------------------
CANDIDATE PROFILE
-----------------------------------------------------
- Name: Sneh Patel
- Experience: ~3 years (internships + full-time roles combined)
- Education: Master of Applied Computer Science (Dalhousie University, 2025)
- Location: Halifax, NS, Canada — open to Remote, Hybrid, or Relocation across Canada
- Core Stack: Python, Java, TypeScript, JavaScript, React, Node.js, AWS, Docker, Kubernetes, PostgreSQL, FastAPI, LangChain, Terraform, Apache Kafka, Spark
- Domains: Full-Stack Development, Cloud/DevOps, Data Engineering, AI/ML Integration, IT Support/Systems Analysis
- NOT a fit for: Pure QA, manual testing only, non-technical roles, embedded/FPGA, hardware engineering

-----------------------------------------------------
INPUTS
-----------------------------------------------------
JOB TITLE: {title}
JOB DESCRIPTION: {description}
RESUME: {resume_text}

-----------------------------------------------------
EVALUATION RULES (apply in order)
-----------------------------------------------------

STEP 1 — DOMAIN FILTER
Accept ONLY if the role clearly belongs to: Software Development, Data Engineering, Cloud, DevOps, AI/ML, IT Support, Systems Analysis, Cybersecurity, or adjacent technical domains.
Reject immediately if: Marketing, HR, Finance, Sales, Manual QA, Hardware Engineering, or non-technical.
If rejected here — verdict = "no", stop evaluation.

STEP 2 — EXPERIENCE GAP CHECK
Extract the required years of experience from the JD (look for phrases like "X+ years", "minimum X years", "X to Y years").
Compare against candidate's 3 years:
- 0 to 4 years required — PASS (good fit range)
- 5 years required — BORDERLINE (check if skills compensate)
- 6+ years required — FAIL (too senior, do not apply)
- Not mentioned — NEUTRAL (continue to skill check)

STEP 3 — SKILLS MATCH SCORING
List every distinct technical skill, tool, framework, or platform mentioned in the JD.
For each one, check if it appears in the resume (directly or through a clearly equivalent technology).
Compute a match percentage: matched_skills / total_jd_skills * 100

Scoring bands:
- 70% or above — Strong match
- 45% to 69% — Partial match
- Below 45% — Weak match

STEP 4 — ROLE LEVEL CALIBRATION
Detect the seniority signal from the JD title and language:
- Junior / Entry-level / Associate / New Grad — BONUS (bump verdict up one level if on the fence)
- Intermediate / Mid-level / no seniority label — NEUTRAL
- Senior / Lead / Staff / Principal / Manager — PENALTY (bump verdict down one level)

STEP 5 — FINAL VERDICT MATRIX

Experience Check | Skills Match | Role Level   | Verdict
PASS             | Strong       | Any          | yes
PASS             | Partial      | Junior/Mid   | yes
PASS             | Partial      | Senior       | maybe
PASS             | Weak         | Junior       | maybe
PASS             | Weak         | Mid/Senior   | no
BORDERLINE       | Strong       | Junior/Mid   | maybe
BORDERLINE       | Partial/Weak | Any          | no
FAIL             | Any          | Any          | no

-----------------------------------------------------
OUTPUT FORMAT — return ONLY valid JSON, no backticks, no extra text
-----------------------------------------------------

{{
  "verdict": "<yes | maybe | no>",
  "years_required": "<exact number, range like 3-5, or unspecified>",
  "role_level": "<junior | mid | senior | unspecified>",
  "skills_match_pct": <integer 0-100>,
  "matched_skills": ["skill1", "skill2"],
  "missing_skills": ["skill1", "skill2"],
  "reasoning": "<2 sentences max — why this verdict, what the main gap or strength is>"
}}
"""


# --------------------------- Helpers ---------------------------

def load_env_file(file_path):
    with open(file_path, encoding="utf-8") as f:
        for line in f:
            if line.strip() and not line.startswith("#") and not line.startswith(";"):
                key, value = line.strip().split("=", 1)
                os.environ[key] = value


def load_resume_text(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        logging.warning("Could not load resume at %s.", path)
        return ""


def format_prompt(template: str, **kwargs) -> str:
    escaped = template.replace("{", "{{").replace("}", "}}")
    for k, v in kwargs.items():
        escaped = escaped.replace(f"{{{{{k}}}}}", str(v))
    return escaped


def load_df():
    if os.path.exists(excel_file):
        df = pd.read_excel(excel_file, engine="openpyxl")
        for col in required_columns:
            if col not in df.columns:
                df[col] = ""
        return df
    return pd.DataFrame(columns=required_columns)


def parse_json_response(response_text: str) -> dict:
    txt = response_text.strip()
    if txt.startswith("```"):
        m = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", txt, re.IGNORECASE)
        if m:
            txt = m.group(1).strip()
    txt = re.sub(r",\s*([}\]])", r"\1", txt)
    try:
        data = json.loads(txt)
    except Exception as e:
        logging.error("Unparsable JSON from model:\n%s", response_text)
        raise e

    verdict = str(data.get("verdict", "")).lower().strip()
    if verdict not in {"yes", "no", "maybe"}:
        raise ValueError(f"Invalid verdict: {verdict!r}")

    matched = data.get("matched_skills", [])
    missing = data.get("missing_skills", [])

    return {
        "verdict":          verdict,
        "years_required":   str(data.get("years_required", "unspecified")).strip(),
        "role_level":       str(data.get("role_level", "unspecified")).lower().strip(),
        "skills_match_pct": int(data.get("skills_match_pct", 0)),
        "matched_skills":   ", ".join(matched) if isinstance(matched, list) else str(matched),
        "missing_skills":   ", ".join(missing) if isinstance(missing, list) else str(missing),
        "reasoning":        str(data.get("reasoning", "")).strip(),
    }


def beautify_excel(path: str = None):
    target = path or excel_file
    wb = load_workbook(target)
    ws = wb.active

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.auto_filter.ref = ws.dimensions

    red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).lower()] = cell.column_letter

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col = cell.column_letter

            if col == header_map.get("ai_recommendation"):
                val = str(cell.value).lower() if cell.value else ""
                if val == "no":
                    cell.fill = red_fill
                elif val == "yes":
                    cell.fill = green_fill
                elif val == "maybe":
                    cell.fill = yellow_fill

            if col == header_map.get("skills_match_pct"):
                try:
                    pct = int(cell.value)
                    cell.fill = green_fill if pct >= 70 else (yellow_fill if pct >= 45 else orange_fill)
                except (TypeError, ValueError):
                    pass

            if col == header_map.get("link") and cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=10
        )
        ws.column_dimensions[col_letter].width = max(10, min(60, max_len + 2))

    wb.save(target)


def write_excel_safely(df: pd.DataFrame, path: str) -> str:
    try:
        df.to_excel(path, index=False, engine="openpyxl")
        beautify_excel(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        fallback = f"{base}_{date.today().isoformat()}{ext}"
        logging.warning("File locked. Saving to fallback: %s", fallback)
        df.to_excel(fallback, index=False, engine="openpyxl")
        beautify_excel(fallback)
        return fallback


def send_with_retries(assistant, msg: str, tries: int = 3, backoff_sec: float = 1.5):
    last_err = None
    for attempt in range(1, tries + 1):
        try:
            return assistant.submit_message(msg)
        except Exception as e:
            last_err = e
            logging.warning("Model call failed (attempt %d/%d): %s", attempt, tries, e)
            time.sleep(backoff_sec * attempt)
    raise last_err


# --------------------------- Core flow ---------------------------

def scrape_and_filter_ai(unique_urls, assistant, resume_text):
    offset   = 0
    new_data = [0]
    df       = pd.DataFrame(columns=required_columns)

    while len(new_data) > 0:
        try:
            logging.info("Scraping at offset %d", offset)
            new_data = scrape_all_jobs(
                site_name=os.getenv("sites", "jobright,indeed,glassdoor"),
                search_term=os.getenv("search_term", "software engineer"),
                location=os.getenv("location", "Halifax, Nova Scotia"),
                hours_old=os.getenv("hours_old", "72"),
                results_wanted=os.getenv("results_wanted", "20"),
                offset=offset,
                country_indeed=os.getenv("country_indeed", "canada"),
                is_remote=os.getenv("is_remote", "false").lower() == "true",
            )
            logging.info("%d jobs scraped", len(new_data))
        except Exception as e:
            logging.error("Scraping error: %s\n%s", e, traceback.format_exc())
            new_data = pd.DataFrame(columns=required_columns)

        for _, row in tqdm(new_data.iterrows(), total=len(new_data), desc="Analyzing Jobs"):
            try:
                job_url = row.get("job_url", "")
                if job_url in unique_urls:
                    continue

                msg = format_prompt(
                    PERSONAL_JOB_FINDER_PROMPT,
                    title=row.get("title", ""),
                    description=row.get("description", ""),
                    resume_text=resume_text,
                )

                ai_response = send_with_retries(assistant, msg, tries=3, backoff_sec=1.5)
                result      = parse_json_response(ai_response)

                new_row = {
                    "AI_recommendation": result["verdict"],
                    "company":           row.get("company", ""),
                    "title":             row.get("title", ""),
                    "link":              job_url,
                    "years_required":    result["years_required"],
                    "role_level":        result["role_level"],
                    "skills_match_pct":  result["skills_match_pct"],
                    "matched_skills":    result["matched_skills"],
                    "missing_skills":    result["missing_skills"],
                    "reasoning":         result["reasoning"],
                    "description":       row.get("description", ""),
                    "posted_date":       row.get("date_posted", ""),
                }
                df.loc[len(df)] = new_row
                unique_urls.add(job_url)

                logging.info(
                    "[%s] %s @ %s | match=%s%% | level=%s",
                    result["verdict"].upper(),
                    row.get("title", ""),
                    row.get("company", ""),
                    result["skills_match_pct"],
                    result["role_level"],
                )

            except Exception as e:
                logging.error("AI error: %s\n%s", e, traceback.format_exc())

        offset += len(new_data)
        break

    return df


def main():
    load_env_file(".env")
    data        = load_df()
    unique_urls = set(data["link"])
    resume_text = load_resume_text(os.getenv("RESUME_PATH", "resume.txt"))

    assistant = OllamaAssistant(model=os.getenv("model", "gemma3:e4b"))
    logging.info("Ollama Assistant ready. Model: %s", assistant.model)

    new_df = scrape_and_filter_ai(unique_urls, assistant, resume_text)
    df     = pd.concat([data, new_df], ignore_index=True)
    path   = write_excel_safely(df, excel_file)
    logging.info("Excel written to: %s", path)


if __name__ == "__main__":
    main()