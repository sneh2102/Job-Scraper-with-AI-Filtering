"""
db_manager.py
=============
SQLite databases for scraped jobs and applied jobs.

  jobs.db    — mirrors jobs.xlsx  (every scraped job)
  applied.db — mirrors Job-Tracker.xlsx (every pipeline-generated resume)
               BUT also stores tex_content + cover_letter_content (not in Excel)

Both databases stay in sync with their Excel counterparts: any write to SQLite
immediately regenerates the corresponding Excel file.
"""
import sqlite3
import os
import logging
import pandas as pd
from datetime import datetime

log = logging.getLogger("db_manager")


# ── Column lists ──────────────────────────────────────────────────────────────

JOBS_COLS = [
    "ai_recommendation", "company", "title", "link", "location",
    "site", "years_required", "role_level", "skills_match_pct",
    "matched_skills", "missing_skills", "reasoning", "description",
    "posted_date", "application_status",
]

APPLIED_EXCEL_COLS = [      # columns that go into the Excel (no tex/cover)
    "company", "title", "job_url", "location", "applied_date",
    "ats_score", "status", "resume_path", "cover_path",
]


# ══════════════════════════════════════════════════════════════════════════════
# JOBS DATABASE
# ══════════════════════════════════════════════════════════════════════════════

class JobsDB:
    """
    SQLite store for scraped jobs. Mirrors jobs.xlsx.

    Keyed on `link` (UNIQUE). All writes are followed by sync_to_excel().
    """

    def __init__(self, db_path: str = "jobs.db"):
        self.db_path = db_path
        self._init()

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init(self):
        with self._conn() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS jobs (
                    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                    ai_recommendation  TEXT,
                    company            TEXT,
                    title              TEXT,
                    link               TEXT UNIQUE,
                    location           TEXT,
                    site               TEXT,
                    years_required     TEXT,
                    role_level         TEXT,
                    skills_match_pct   TEXT,
                    matched_skills     TEXT,
                    missing_skills     TEXT,
                    reasoning          TEXT,
                    description        TEXT,
                    posted_date        TEXT,
                    application_status TEXT DEFAULT 'Not Applied',
                    created_at         TEXT DEFAULT (datetime('now'))
                )
            """)

    # ── Writes ────────────────────────────────────────────────────────────────

    def upsert(self, row: dict):
        """Insert or update a job keyed by link. Keys are normalised to lowercase."""
        norm = {k.lower(): v for k, v in row.items()}
        link = str(norm.get("link", "")).strip()
        if not link or link in ("nan", "none", ""):
            return
        cols = [c for c in JOBS_COLS if c in norm]
        if "link" not in cols:
            cols.insert(0, "link")
        sql = (
            f"INSERT INTO jobs ({', '.join(cols)}) "
            f"VALUES ({', '.join(':' + c for c in cols)}) "
            f"ON CONFLICT(link) DO UPDATE SET "
            f"{', '.join(f'{c}=excluded.{c}' for c in cols if c != 'link')}"
        )
        with self._conn() as c:
            c.execute(sql, {col: str(norm.get(col, "")) for col in cols})

    def delete_by_verdict(self, verdict: str):
        with self._conn() as c:
            c.execute("DELETE FROM jobs WHERE LOWER(ai_recommendation)=LOWER(?)", (verdict,))

    def delete_not_applied(self):
        with self._conn() as c:
            c.execute("DELETE FROM jobs WHERE LOWER(COALESCE(application_status,''))!='applied'")

    def delete_all(self):
        with self._conn() as c:
            c.execute("DELETE FROM jobs")

    def delete_by_link(self, link: str):
        with self._conn() as c:
            c.execute("DELETE FROM jobs WHERE link=?", (link,))

    def set_application_status(self, link: str, status: str):
        with self._conn() as c:
            c.execute("UPDATE jobs SET application_status=? WHERE link=?", (status, link))

    def set_verdict(self, link: str, verdict: str):
        with self._conn() as c:
            c.execute("UPDATE jobs SET ai_recommendation=? WHERE link=?", (verdict, link))

    # ── Reads ─────────────────────────────────────────────────────────────────

    def get_all(self) -> pd.DataFrame:
        with self._conn() as c:
            return pd.read_sql_query(
                "SELECT * FROM jobs ORDER BY id DESC", c
            )

    def count(self) -> int:
        with self._conn() as c:
            return c.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]

    # ── Sync helpers ──────────────────────────────────────────────────────────

    def import_from_excel(self, excel_path: str) -> int:
        """Bulk-import Excel into SQLite using a single transaction (fast)."""
        if not os.path.exists(excel_path):
            return 0
        try:
            df = pd.read_excel(excel_path, engine="openpyxl")
            df.columns = [c.lower() for c in df.columns]

            valid_cols = [c for c in JOBS_COLS if c in df.columns]
            if "link" not in df.columns:
                return 0

            records = []
            for _, row in df.iterrows():
                link = str(row.get("link", "")).strip()
                if not link or link.lower() in ("nan", "none", ""):
                    continue
                records.append({c: str(row.get(c, "")) for c in valid_cols})

            if not records:
                return 0

            col_str = ", ".join(valid_cols)
            ph_str  = ", ".join(f":{c}" for c in valid_cols)
            upd_str = ", ".join(f"{c}=excluded.{c}" for c in valid_cols if c != "link")
            sql = (
                f"INSERT INTO jobs ({col_str}) VALUES ({ph_str}) "
                f"ON CONFLICT(link) DO UPDATE SET {upd_str}"
            )
            with sqlite3.connect(self.db_path) as conn:
                conn.executemany(sql, records)

            log.info("Imported %d rows from %s into jobs.db", len(records), excel_path)
            return len(records)
        except Exception as e:
            log.error("jobs import_from_excel: %s", e)
            return 0

    def sync_to_excel(self, excel_path: str):
        """Overwrite Excel with current SQLite data (verdict-coloured)."""
        df = self.get_all()
        df = df.drop(columns=["id", "created_at"], errors="ignore")
        # Capitalise column so the existing tracker UI doesn't break
        if "ai_recommendation" in df.columns:
            df = df.rename(columns={"ai_recommendation": "AI_recommendation"})
        _write_excel(df, excel_path, verdict_col="AI_recommendation")


# ══════════════════════════════════════════════════════════════════════════════
# APPLIED DATABASE
# ══════════════════════════════════════════════════════════════════════════════

class AppliedDB:
    """
    SQLite store for pipeline-generated resumes.
    Stores tex_content + cover_letter_content (never written to Excel).
    """

    def __init__(self, db_path: str = "applied.db"):
        self.db_path = db_path
        self._init()

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init(self):
        with self._conn() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS applied_jobs (
                    id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                    company              TEXT,
                    title                TEXT,
                    job_url              TEXT,
                    location             TEXT,
                    applied_date         TEXT,
                    ats_score            INTEGER DEFAULT 0,
                    status               TEXT DEFAULT 'applied',
                    tex_content          TEXT,
                    cover_letter_content TEXT,
                    resume_path          TEXT,
                    cover_path           TEXT,
                    created_at           TEXT DEFAULT (datetime('now'))
                )
            """)

    # ── Writes ────────────────────────────────────────────────────────────────

    def add(self, data: dict) -> int:
        """Insert and return the new row id."""
        with self._conn() as c:
            cur = c.execute("""
                INSERT INTO applied_jobs
                    (company, title, job_url, location, applied_date,
                     ats_score, status, tex_content, cover_letter_content,
                     resume_path, cover_path)
                VALUES
                    (:company, :title, :job_url, :location, :applied_date,
                     :ats_score, :status, :tex_content, :cover_letter_content,
                     :resume_path, :cover_path)
            """, {
                "company":              data.get("company", ""),
                "title":                data.get("title", ""),
                "job_url":              data.get("job_url", ""),
                "location":             data.get("location", ""),
                "applied_date":         data.get("applied_date", datetime.now().strftime("%Y-%m-%d")),
                "ats_score":            int(data.get("ats_score", 0)),
                "status":               data.get("status", "applied"),
                "tex_content":          data.get("tex_content", ""),
                "cover_letter_content": data.get("cover_letter_content", ""),
                "resume_path":          data.get("resume_path", ""),
                "cover_path":           data.get("cover_path", ""),
            })
            return cur.lastrowid

    def delete_by_id(self, job_id: int):
        with self._conn() as c:
            c.execute("DELETE FROM applied_jobs WHERE id=?", (job_id,))

    def delete_all(self):
        with self._conn() as c:
            c.execute("DELETE FROM applied_jobs")

    # ── Reads ─────────────────────────────────────────────────────────────────

    def get_all(self) -> pd.DataFrame:
        """Returns all rows WITHOUT tex/cover content (for UI listing)."""
        with self._conn() as c:
            return pd.read_sql_query(
                "SELECT id, company, title, job_url, location, applied_date, "
                "ats_score, status, resume_path, cover_path, created_at "
                "FROM applied_jobs ORDER BY id DESC",
                c,
            )

    def get_tex(self, job_id: int) -> str:
        with self._conn() as c:
            row = c.execute(
                "SELECT tex_content FROM applied_jobs WHERE id=?", (job_id,)
            ).fetchone()
            return row["tex_content"] if row and row["tex_content"] else ""

    def get_cover(self, job_id: int) -> str:
        with self._conn() as c:
            row = c.execute(
                "SELECT cover_letter_content FROM applied_jobs WHERE id=?", (job_id,)
            ).fetchone()
            return row["cover_letter_content"] if row and row["cover_letter_content"] else ""

    def get_by_id(self, job_id: int) -> dict:
        with self._conn() as c:
            row = c.execute(
                "SELECT * FROM applied_jobs WHERE id=?", (job_id,)
            ).fetchone()
            return dict(row) if row else {}

    def count(self) -> int:
        with self._conn() as c:
            return c.execute("SELECT COUNT(*) FROM applied_jobs").fetchone()[0]

    # ── Sync ─────────────────────────────────────────────────────────────────

    def sync_to_excel(self, excel_path: str):
        """Export to Excel WITHOUT tex/cover letter columns."""
        df = self.get_all()
        df = df.drop(columns=["id", "created_at"], errors="ignore")
        _write_excel(df, excel_path, verdict_col=None)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED EXCEL WRITER
# ══════════════════════════════════════════════════════════════════════════════

def _write_excel(df: pd.DataFrame, path: str, verdict_col: str = None):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    try:
        df.to_excel(path, index=False, engine="openpyxl")
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f6feb", end_color="1f6feb", fill_type="solid")
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        if verdict_col:
            hmap = {str(c.value).lower(): c.column_letter for c in ws[1] if c.value}
            vcol = hmap.get(verdict_col.lower(), "")
            if vcol:
                green  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                red    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.column_letter == vcol:
                            v = str(cell.value).lower() if cell.value else ""
                            cell.fill = (green if v == "yes" else red if v == "no"
                                         else yellow if v == "maybe" else cell.fill)

        for col_idx in range(1, ws.max_column + 1):
            col_l = get_column_letter(col_idx)
            max_len = max((len(str(c.value)) for c in ws[col_l] if c.value), default=10)
            ws.column_dimensions[col_l].width = max(10, min(60, max_len + 2))

        wb.save(path)
    except PermissionError:
        log.warning("Excel file is open — could not write %s", path)
    except Exception as e:
        log.warning("Excel write failed (%s): %s", path, e)
