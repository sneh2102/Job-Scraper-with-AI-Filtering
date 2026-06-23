"""
JobHunter — Windows Desktop Application
========================================
pip install customtkinter pillow
python app.py
"""
import sys
import os
from ai import OllamaAssistant, parse_ai_verdict, send_with_retries

# Fix paths when running as PyInstaller exe
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

os.chdir(BASE_DIR)
CONFIG_FILE = os.path.join(BASE_DIR, "app_config.json")
import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import threading
import queue
import json
import os
import shutil
import sys
import logging
import time
import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from setup_wizard import check_and_run_setup, ProfileTab, load_profile

# ── App theme ─────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

CONFIG_FILE = "app_config.json"

# ══════════════════════════════════════════════════════════════
# PASTE THIS INTO app.py — REPLACE THE ENTIRE DEFAULT_CONFIG
# ══════════════════════════════════════════════════════════════

DEFAULT_CONFIG = {
    "scraper": {
        "sites": "indeed,glassdoor,jobright",
        "search_term": "Software Developer",
        "location": "Canada",
        "country_indeed": "canada",
        "hours_old": "72",
        "results_wanted": "20",
        "is_remote": False,
    },
    "model": {
        "scraping": "gemma4:31b-cloud",
        "pipeline": "gemma4:31b-cloud",
        "api_keys": [],
        "use_cloud": True,
        "num_predict": 32768,
        "num_ctx": 65536,
        "temperature": 0.3,
    },
    "pipeline": {
        "max_ats_iterations": 3,
        "ats_pass_threshold": 85,
        "output_dir": "outputs",
        "excel_path": "jobs.xlsx",
        "applied_excel_path": "Job-Tracker.xlsx",
        "applied_folder_path": "fultime",
        "resume_path": "resume.txt",
        "projects_path": "Projects.txt",
        "resume_filename": "Resume",
        "cover_letter_filename": "Cover_Letter",
        "use_jd_location": True,
        "default_location": "Canada",
        "section_order": ["education", "summary", "skills", "experience", "projects", "achievements"],
    },
    "custom_sections": [
        {
            "id": "summary",
            "name": "Summary",
            "system_prompt": "You are an expert resume writer. Output ONLY raw LaTeX — no backticks, no explanation. Generate ONLY the Professional Summary section body.\\n\\nOUTPUT FORMAT CONTRACT:\\n\\\\section{{Professional Summary}}\\n{{A 3-4 line concise paragraph summarizing the candidate's value proposition, tailored to the JD. No bullet points. Use LaTeX and ensure all % -> \\\\%, all & -> \\\\&.}}\\n\\nABSOLUTE RULES:\\n- NO \\\\documentclass, NO \\\\usepackage, NO \\\\begin{{document}}, NO \\\\end{{document}}",
            "user_prompt": "Write a professional summary for {full_name} for the role of {title} at {company}.\\n\\nJOB DESCRIPTION:\\n{description}\\n\\nCANDIDATE RESUME:\\n{existing_resume}\\n\\nATS FEEDBACK:\\n{ats_feedback}"
        },
        {
            "id": "achievements",
            "name": "Achievements",
            "system_prompt": "You are an expert resume writer. Output ONLY raw LaTeX — no backticks, no explanation. Generate ONLY the Achievements section body.\\n\\nOUTPUT FORMAT CONTRACT:\\n\\\\section{{Achievements}}\\n\\\\resumeSubHeadingListStart\\n  \\\\resumeItem{{Achievement 1 with metric}}\\n  \\\\resumeItem{{Achievement 2 with metric}}\\n\\\\resumeSubHeadingListEnd\\n\\nABSOLUTE RULES:\\n- NO \\\\documentclass, NO \\\\usepackage, NO \\\\begin{{document}}, NO \\\\end{{document}}",
            "user_prompt": "List 2-3 key professional achievements for {full_name} that would impress the hiring manager for {title} at {company}.\\n\\nJOB DESCRIPTION:\\n{description}\\n\\nCANDIDATE RESUME:\\n{existing_resume}\\n\\nATS FEEDBACK:\\n{ats_feedback}"
        }
    ],
    "google": {
        "enabled": False,
        "spreadsheet_id": "",
        "drive_folder": "Job Applications",
        "sheet_tab": "Sheet1",
    },
    "github": {
        "token": "",
    },
    "prompts": {

# ══════════════════════════════════════════════════════════════
# PROMPT 1: JOB SCREENER
# ══════════════════════════════════════════════════════════════
"job_screener": """You are a ruthless but contextually aware IT job screener.
Your job is to evaluate whether the candidate should apply to this position.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CANDIDATE PROFILE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{candidate_profile}

STRONG FIT FOR:
- Software Developer / Engineer (any stack)
- Full Stack Developer
- Backend / Frontend Developer
- Data Engineer / Analyst
- Cloud / DevOps Engineer
- AI / ML Engineer
- IT Support / Systems Analyst
- Junior to Mid-level roles (0-5 years required)

NOT A FIT FOR:
- Pure Manual QA / Testing (no coding)
- Hardware / Embedded / FPGA Engineering
- Marketing, Sales, HR, Finance, Non-technical
- Roles requiring 6+ years experience
- Roles requiring specific certifications they don't have (PMP, CPA, etc.)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INPUTS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE: {title}
JOB DESCRIPTION: {description}
RESUME SUMMARY: {resume_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EVALUATION STEPS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STEP 1 — DOMAIN CHECK
Accept if role is in: Software Dev, Data Eng, Cloud, DevOps, AI/ML, IT Support, Cybersecurity.
Reject immediately (verdict=no) if: Marketing, HR, Sales, Manual QA, Hardware, non-technical.

STEP 2 — EXPERIENCE CHECK
Extract required years from JD ("X+ years", "minimum X", "X-Y years experience").
0-4 yrs → PASS | 5 yrs → BORDERLINE | 6+ yrs → FAIL | Not mentioned → NEUTRAL
LIBERAL RULE: If candidate has equivalent project/academic experience, give benefit of doubt.

STEP 3 — SKILLS MATCH (BE LIBERAL WITH EQUIVALENTS)
List every technical skill/tool/framework in the JD.
Count each as matched if candidate has it OR a clear equivalent:
  PostgreSQL ≈ MySQL ≈ SQL | AWS ≈ GCP ≈ Azure | React ≈ Vue ≈ Angular
  Docker ≈ containerization | Kubernetes ≈ container orchestration
score = matched / total_jd_skills * 100
70%+ = Strong | 45-69% = Partial | <45% = Weak

STEP 4 — SENIORITY CALIBRATION
Junior/Entry/Associate/New Grad → BONUS (raise verdict one level if borderline)
Mid/Intermediate/no label → NEUTRAL
Senior/Lead/Staff/Principal/Manager → PENALTY (lower verdict one level)

STEP 5 — FINAL VERDICT
PASS + Strong + Any → yes
PASS + Partial + Junior/Mid → yes
PASS + Partial + Senior → maybe
PASS + Weak + Junior → maybe
PASS + Weak + Mid/Senior → no
BORDERLINE + Strong + Junior/Mid → maybe
BORDERLINE + anything else → no
FAIL + anything → no

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
OUTPUT — return ONLY valid JSON, no markdown, no backticks
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{"verdict": "yes or maybe or no", "years_required": "number or range", "role_level": "junior or mid or senior", "skills_match_pct": 75, "matched_skills": ["Python", "React"], "missing_skills": ["Go", "Terraform"], "reasoning": "Strong match on full-stack skills. Missing Terraform but compensated by Docker/K8s experience."}""",

# ══════════════════════════════════════════════════════════════
# PROMPT 2: SKILLS SECTION
# ══════════════════════════════════════════════════════════════
"skills_section": """You are an expert resume writer specializing in ATS optimization.
Output ONLY raw LaTeX for the Technical Skills section.
NO \\documentclass, NO \\usepackage, NO \\begin{{document}}, NO \\end{{document}}.
Output ONLY the \\section{{Technical Skills}} block — nothing else.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONTEXT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE:   {title}
COMPANY:     {company}
JOB DESCRIPTION:
{description}

CANDIDATE'S EXISTING SKILLS (from resume):
{existing_resume}

ATS FEEDBACK FROM PREVIOUS ATTEMPT (MUST address every point):
{ats_feedback}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INSTRUCTIONS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Read the JD carefully. Extract EVERY distinct technical skill, tool, framework, platform, methodology mentioned.
2. Map each JD keyword to the candidate's existing skills.
3. Create 5-6 categories that match JD domain terminology exactly.
4. FRONT-LOAD: Put the most JD-relevant keywords FIRST in each category.
5. Include JD-specific tools even if candidate used equivalents (show both if possible).
6. Category names should mirror JD language (e.g. if JD says "Cloud Infrastructure" use that, not "Cloud").

CATEGORY EXAMPLES BY DOMAIN:
- Languages: Python, Java, TypeScript, JavaScript, Go, SQL, Bash
- Frontend: React, Next.js, TypeScript, HTML5, CSS3, Tailwind CSS
- Backend: Node.js, FastAPI, Spring Boot, REST APIs, GraphQL, Microservices
- Cloud & DevOps: AWS (EC2, S3, Lambda, RDS), Docker, Kubernetes, Terraform, CI/CD, GitHub Actions
- Data & AI: PostgreSQL, MongoDB, Redis, Apache Kafka, Spark, LangChain, pandas, scikit-learn
- Tools: Git, JIRA, Linux, VS Code, Postman, Jupyter

RULES:
- All % → \\%, all & → \\&
- No special chars: no ->, no <>, no em dashes
- Use exact JD terminology where possible
- 8-12 items per category maximum

OUTPUT FORMAT:
\\section{{Technical Skills}}
 \\begin{{itemize}}[leftmargin=0.15in, label={{}}]
    \\small{{\\item{{
     \\textbf{{[Category 1]}}{{: tool1, tool2, tool3, tool4}} \\\\
     \\textbf{{[Category 2]}}{{: tool1, tool2, tool3, tool4}} \\\\
     \\textbf{{[Category 3]}}{{: tool1, tool2, tool3, tool4}} \\\\
     \\textbf{{[Category 4]}}{{: tool1, tool2, tool3, tool4}} \\\\
     \\textbf{{[Category 5]}}{{: tool1, tool2, tool3, tool4}} \\\\
     \\textbf{{[Category 6]}}{{: tool1, tool2, tool3, tool4}}
    }}}}
 \\end{{itemize}}""",

# ══════════════════════════════════════════════════════════════
# PROMPT 3: EXPERIENCE SECTION
# ══════════════════════════════════════════════════════════════
"experience_section": """You are an expert resume writer and career strategist.
Output ONLY raw LaTeX for the Experience section.
NO \\documentclass, NO \\usepackage, NO \\begin{{document}}, NO \\end{{document}}.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONTEXT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE:   {title}
COMPANY:     {company}
JOB DESCRIPTION (extract EVERY required skill and responsibility):
{description}

CANDIDATE'S REAL EXPERIENCE (use real dates, metrics, technologies):
{existing_resume}

ATS FEEDBACK (MUST address every point listed here):
{ats_feedback}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FABRICATION STRATEGY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
For each role, write 2-3 FABRICATED bullets FIRST, then 2-3 REAL bullets.

FABRICATED bullets must:
- Directly match a JD requirement that real experience doesn't cover
- Sound 100% realistic — specific domain, realistic metrics, believable scope
- Use EXACT JD keywords naturally woven in
- Be indistinguishable from real bullets
- Have concrete numbers: team sizes (3-8 people), percentages (15-65%), time saved (2-4 hours/week), scale (thousands to millions of records)
- NEVER reveal they are fabricated

REAL bullets must:
- Be rewritten to use JD terminology where possible
- Keep the core achievement but rephrase with JD keywords
- Add or adjust metrics if needed to strengthen the bullet

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ROLES — KEEP EXACTLY AS SHOWN (name, title, dates LOCKED)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Role 1: Support Consultant @ Nova Scotia Health Authority | Feb 2026 – Present
→ 5 bullets: 3 fabricated + 2 real (healthcare IT, systems, user support domain)

Role 2: Software Developer @ TeleAI Corporation | Jan 2025 – May 2025
→ 4 bullets: 2 fabricated + 2 real (AI/ML, full-stack, APIs domain)

Role 3: R&D Project Engineer @ MyTech Lab | Sept 2024 – Dec 2024
→ 3 bullets: 2 fabricated + 1 real (research, algorithms, optimization domain)

Role 4: Full-Stack Developer @ Webforest LLP | Jan 2023 – Jul 2023
→ 4 bullets: 2 fabricated + 2 real (web development, databases, deployment domain)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
BULLET RULES (NON-NEGOTIABLE)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- EVERY bullet = exactly 2 full lines in LaTeX. Never one-liners.
- EVERY bullet has a specific metric (%, count, time, scale, money saved)
- Use \\textbf{{}} on 2-3 key terms per bullet only
- Start each bullet with a strong action verb (Architected, Designed, Implemented, Optimized, Reduced, Increased, Deployed, Built, Automated, Led, Developed)
- All % → \\%, all & → \\&
- No special chars: no ->, no <>, no em dashes

EXAMPLE FABRICATED BULLET (for Kafka requirement):
\\resumeItem{{Architected and deployed a real-time \\textbf{{Apache Kafka}} event streaming
pipeline processing \\textbf{{45,000 messages/sec}} across 6 microservices, reducing
end-to-end data latency by \\textbf{{71\\%}} and enabling live analytics dashboards.}}

OUTPUT:
\\section{{Experience}}
\\resumeSubHeadingListStart
  [all 4 roles with bullets]
\\resumeSubHeadingListEnd

Raw LaTeX ONLY. No backticks. No preamble.""",

# ══════════════════════════════════════════════════════════════
# PROMPT 4: PROJECTS SECTION
# ══════════════════════════════════════════════════════════════
"projects_section": """You are an expert resume writer specializing in project showcasing.
Output ONLY raw LaTeX for the Relevant Projects section.
NO \\documentclass, NO \\usepackage, NO \\begin{{document}}, NO \\end{{document}}.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONTEXT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE:   {title}
COMPANY:     {company}
JOB DESCRIPTION:
{description}

CANDIDATE'S EXISTING RESUME (reference for tech stack and context):
{existing_resume}

AVAILABLE PROJECTS (select the 3-4 most relevant):
{projects}

ATS FEEDBACK (MUST address every point):
{ats_feedback}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INSTRUCTIONS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. Read JD carefully. Identify the top 5 technical requirements.
2. Select 3-4 projects from the list that best demonstrate those requirements.
3. If a critical JD skill has NO matching project, FABRICATE a realistic project to fill that gap.
4. For each project, write 2-3 bullets following this structure:
   Bullet 1: The problem + technology used + scale (what you built and with what)
   Bullet 2: The technical implementation detail + specific tool from JD
   Bullet 3: The measurable outcome (performance gain, users served, time saved, accuracy %)

FABRICATED PROJECT RULES:
- Must be plausible for a CS student/junior developer
- Name it something realistic (e.g. "Real-Time Inventory Tracker", "ML Pipeline Optimizer")
- Use specific versions/tools that match JD
- Metrics must be believable (not 99.99% uptime on first project)

BULLET RULES:
- Every bullet = 2 full lines minimum
- Every bullet has a specific number/metric
- Use \\textbf{{}} on project name and 1-2 key technologies
- Use exact JD keywords naturally
- All % → \\%, all & → \\&

OUTPUT FORMAT:
\\section{{Relevant Projects}}
\\resumeSubHeadingListStart
  \\resumeProjectHeading
    {{\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}}}{{}}
  \\resumeItemListStart
    \\resumeItem{{bullet 1}}
    \\resumeItem{{bullet 2}}
    \\resumeItem{{bullet 3}}
  \\resumeItemListEnd
\\resumeSubHeadingListEnd

Raw LaTeX ONLY. No backticks. No preamble.""",

# ══════════════════════════════════════════════════════════════
# PROMPT 5: COVER LETTER
# ══════════════════════════════════════════════════════════════
"cover_letter": """You are writing a cover letter for {full_name}.
Output ONLY the plain text cover letter — no LaTeX, no backticks, no JSON, no explanation.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CONTEXT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE:   {title}
COMPANY:     {company}
JOB DESCRIPTION:
{description}

CANDIDATE'S RESUME (reference for real experiences and technologies):
{existing_resume}

TODAY'S DATE: {today}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TONE GUIDELINES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Semi-formal but genuinely human — like a smart person wrote it, not an AI
- Occasional natural imperfection is fine (contractions like "I've", "I'm", "it's")
- Enthusiastic but not desperate — confident but not arrogant
- Reference SPECIFIC things from the JD, not generic phrases
- Tell small stories — one concrete moment or realization, not just lists
- Each time you write this, make it unique — different angle, different story

NEVER USE:
- Em dashes (—) anywhere
- Arrow symbols (->, =>, <>)
- Corporate filler ("leverage synergies", "passionate team player", "dynamic environment")
- Bullet points
- More than 4 paragraphs

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXACT FORMAT (follow precisely)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{full_name}
{phone}
{email}

{today}

Hiring Manager,
{company}
[Extract city and province from JD. If remote or unclear, write: Canada]

Subject: Application for {title}

Dear Hiring Manager,

[PARAGRAPH 1 — 3-5 sentences]
What specifically attracted you to this role. Reference one concrete thing from the JD
(a technology, a product, a problem they're solving). Make it personal and specific.
Why does this role make sense for where you are in your career right now?

[PARAGRAPH 2 — 4-6 sentences]
2-3 concrete examples from your experience that directly map to JD requirements.
Mention real technologies from your resume naturally (Python, React, AWS etc.).
At least one example should have a specific number or outcome.
Weave in that you tend to build side projects to learn new things.

[PARAGRAPH 3 — 3-4 sentences]
Your approach to learning and problem-solving. Mention that you prefer
building things to understand them, and that you gravitate toward open-source tools.
Make it honest and specific, not generic.

[PARAGRAPH 4 — 2-3 la sentences]
One specific reason why {company} appeals to you (reference their product, mission, or tech stack).
Strong, confident close.

Warm regards,
{full_name}
{email}

Plain text ONLY. No formatting symbols.""",

# ══════════════════════════════════════════════════════════════
# PROMPT 6: ATS CHECKER (STRICT)
# ══════════════════════════════════════════════════════════════
"ats_checker": """You are an extremely strict Fortune 500 ATS system simulator.
Your job is to score resumes rigorously and give precise, actionable feedback.
Return ONLY valid JSON. No LaTeX in your response. No backticks. No explanation outside JSON.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INPUTS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
JOB TITLE: {title}
JOB DESCRIPTION: {description}
RESUME (LaTeX source): {latex}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STRICT SCORING CRITERIA (total 100 points)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. KEYWORD COVERAGE (40 pts) — BE STRICT:
   - List EVERY distinct technical skill, tool, framework, methodology, platform in JD
   - Check each against resume verbatim OR clear equivalent (React=frontend framework, Postgres=SQL DB)
   - Named tools explicitly required (e.g. "must have Terraform") MUST appear verbatim
   - Vague JD = award 28-32 pts base
   - score = (matched / total_required) * 40

2. EXPERIENCE ALIGNMENT (25 pts):
   - Does the role history and seniority match JD expectations?
   - If JD requires 5 yrs and candidate has 3: deduct max 8 pts
   - Domain mismatch (healthcare→fintech): NO deduction if skills transfer
   - Role title mismatch (Support→Engineer): NO deduction
   - Fabricated bullets that match JD: award full credit

3. SKILLS SECTION MATCH (20 pts):
   - Do listed skill categories directly mirror JD requirements?
   - Missing entire JD domain in skills section: deduct 5-8 pts
   - Generic skills without JD-specific tools: partial credit only

4. IMPACT AND METRICS (15 pts):
   - Every bullet should have a quantified result
   - Bullets with specific numbers: 2 pts each (up to cap)
   - Vague bullets ("helped with", "worked on"): 0 pts
   - Strong action verbs + context + metric: full credit

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LIBERAL RULES (apply before deducting)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Equivalent technologies count as matched
- Adjacent domain = no experience deduction
- Soft skills = never penalize
- Nice-to-have / preferred qualifications = never penalize
- 3 yrs vs 5 yrs = max 8 pt deduction, not 25

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FEEDBACK QUALITY REQUIREMENTS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Each feedback field must be SPECIFIC and ACTIONABLE:

skills_feedback: Name EXACT tools to add and which category. Example:
"Add 'Terraform, AWS CloudFormation' to Cloud category. Rename 'Backend' to 'Backend Engineering'
to match JD language. Add a new 'Monitoring' category with: Datadog, CloudWatch, PagerDuty."

experience_feedback: Name COMPANY and BULLET NUMBER. Example:
"TeleAI role bullet 2: rewrite to mention 'event-driven architecture' and add a throughput metric
like 50,000 events/sec. NSHA role: add a bullet about 'incident response' with MTTD/MTTR metrics.
Webforest role: replace generic bullet 3 with one mentioning 'CI/CD pipeline' using GitHub Actions."

projects_feedback: Name SPECIFIC project to replace and what to replace with. Example:
"Replace 'FaceOff' project with a 'Real-Time Data Pipeline' project using Kafka+Spark+S3 to
match JD's streaming requirements. Add 99.9% uptime metric. Keep AsyncDoctor and FaceOff projects."

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
OUTPUT — ONLY this JSON structure
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{
  "score": <integer 0-100>,
  "keyword_coverage_pct": <integer 0-100>,
  "pass": <true if score >= 85 else false>,
  "total_jd_keywords": <count of distinct required skills in JD>,
  "matched_keywords": <count found in resume directly or by equivalent>,
  "missing_keywords": ["exact keyword from JD not in resume"],
  "section_scores": {
    "skills": <0-100>,
    "experience": <0-100>,
    "projects": <0-100>
  },
  "sections_to_rewrite": ["list only sections scoring below 75"],
  "skills_feedback": "Specific instructions: exact tools to add, exact categories to rename, exact JD terminology to use. Plain English only, no LaTeX.",
  "experience_feedback": "Specific instructions naming each company and bullet number. Plain English only, no LaTeX.",
  "projects_feedback": "Specific instructions naming which project to replace/keep and exact tech to use. Plain English only, no LaTeX.",
  "cover_letter_feedback": "",
  "suggestions": [
    "Single most impactful change that would raise the score the most",
    "Second most impactful change",
    "Third most impactful change"
  ]
}"""

    }  # end prompts
}  # end DEFAULT_CONFIG
# ── Config manager ────────────────────────────────────────────
class Config:
    def __init__(self):
        self.data = DEFAULT_CONFIG.copy()
        self.load()

    def load(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    saved = json.load(f)
                self._deep_merge(self.data, saved)
            except Exception as e:
                print(f"Config load error: {e}")

    def save(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2)

    def _deep_merge(self, base, override):
        for k, v in override.items():
            if k in base and isinstance(base[k], dict) and isinstance(v, dict):
                self._deep_merge(base[k], v)
            else:
                base[k] = v

    def get(self, *keys, default=None):
        d = self.data
        for k in keys:
            if isinstance(d, dict) and k in d:
                d = d[k]
            else:
                return default
        return d

    def set(self, *keys_and_value):
        keys, value = keys_and_value[:-1], keys_and_value[-1]
        d = self.data
        for k in keys[:-1]:
            d = d.setdefault(k, {})
        d[keys[-1]] = value


# ── Log handler that sends to queue ──────────────────────────
class QueueLogHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        msg = self.format(record)
        self.log_queue.put(("log", record.levelname, msg))


# ── Sidebar button ────────────────────────────────────────────
class SidebarButton(ctk.CTkButton):
    def __init__(self, master, text, icon, command, **kwargs):
        super().__init__(
            master,
            text=f"  {icon}  {text}",
            command=command,
            anchor="w",
            height=44,
            corner_radius=8,
            fg_color="transparent",
            hover_color=("#2b2d30", "#3a3d42"),
            text_color=("#c9ccd1", "#c9ccd1"),
            font=ctk.CTkFont(family="Segoe UI", size=13),
            **kwargs
        )
        self._btn_text = text   # ← renamed from _text_label

    def set_active(self, active: bool):
        if active:
            self.configure(fg_color=("#1f6feb", "#1f6feb"), text_color="white")
        else:
            self.configure(fg_color="transparent", text_color=("#c9ccd1", "#c9ccd1"))

# ── Log widget ────────────────────────────────────────────────
class LogWidget(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="#0d1117", corner_radius=8, **kwargs)
        self.textbox = ctk.CTkTextbox(
            self,
            font=ctk.CTkFont(family="Consolas", size=11),
            fg_color="#0d1117",
            text_color="#e6edf3",
            wrap="word",
        )
        self.textbox.pack(fill="both", expand=True, padx=4, pady=4)
        self.textbox.tag_config("INFO",    foreground="#4CAF50")
        self.textbox.tag_config("WARNING", foreground="#FFC107")
        self.textbox.tag_config("ERROR",   foreground="#F44336")
        self.textbox.tag_config("DEBUG",   foreground="#9e9e9e")
        self.textbox.tag_config("TIME",    foreground="#555")

    def append(self, level: str, msg: str):
        ts  = datetime.now().strftime("%H:%M:%S")
        tag = level if level in ("INFO", "WARNING", "ERROR", "DEBUG") else "INFO"
        self.textbox.configure(state="normal")
        self.textbox.insert("end", f"[{ts}] ", "TIME")
        self.textbox.insert("end", f"{msg}\n", tag)
        self.textbox.see("end")
        self.textbox.configure(state="disabled")

    def clear(self):
        self.textbox.configure(state="normal")
        self.textbox.delete("1.0", "end")
        self.textbox.configure(state="disabled")


# ══════════════════════════════════════════════════════════════
# SCRAPER VISUAL RESULTS
# ══════════════════════════════════════════════════════════════

class ScraperJobCard(ctk.CTkFrame):
    _VERDICT = {
        "yes":   ("#3fb950", "#1a3e2a"),
        "maybe": ("#d29922", "#3e3200"),
        "no":    ("#f85149", "#3e1a1a"),
    }

    def __init__(self, master, verdict: str, company: str, title: str,
                 skills_pct: int, location: str, matched: list, missing: list, **kwargs):
        v = verdict.lower()
        fg, bg = self._VERDICT.get(v, ("#aaa", "#333"))
        super().__init__(master, corner_radius=8, fg_color="#161b22",
                         border_color=fg, border_width=1, **kwargs)

        r1 = ctk.CTkFrame(self, fg_color="transparent")
        r1.pack(fill="x", padx=10, pady=(8, 3))
        ctk.CTkLabel(r1, text=f" {v.upper()} ",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     fg_color=bg, text_color=fg,
                     corner_radius=4).pack(side="left")
        ctk.CTkLabel(r1, text=f"  {company}  ·  {title[:42]}",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#e6edf3").pack(side="left")
        if location:
            ctk.CTkLabel(r1, text=f"📍 {location[:28]}",
                         font=ctk.CTkFont(size=10), text_color="#7d8590").pack(side="right")

        r2 = ctk.CTkFrame(self, fg_color="transparent")
        r2.pack(fill="x", padx=10, pady=(0, 3))
        ctk.CTkLabel(r2, text=f"{skills_pct}%",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=fg, width=38, anchor="e").pack(side="left")
        bar = ctk.CTkProgressBar(r2, height=7, corner_radius=3,
                                  progress_color=fg, fg_color="#30363d")
        bar.set(max(0.0, min(1.0, skills_pct / 100)))
        bar.pack(side="left", fill="x", expand=True, padx=(6, 0))

        chips = (
            [(sk, True)  for sk in (matched or [])[:4]] +
            [(sk, False) for sk in (missing or [])[:4]]
        )
        if chips:
            r3 = ctk.CTkFrame(self, fg_color="transparent")
            r3.pack(fill="x", padx=10, pady=(0, 8))
            for sk, ok in chips:
                c_fg = "#3fb950" if ok else "#f85149"
                c_bg = "#1a3e2a" if ok else "#3e1a1a"
                ctk.CTkLabel(r3, text=f"{'✓' if ok else '✗'} {sk}",
                             font=ctk.CTkFont(size=10),
                             fg_color=c_bg, text_color=c_fg,
                             corner_radius=4).pack(side="left", padx=2, pady=1)
        else:
            ctk.CTkFrame(self, height=5, fg_color="transparent").pack()


class ScraperResultsPanel(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._counts = {"total": 0, "yes": 0, "maybe": 0, "no": 0}
        self._summary = ctk.CTkLabel(
            self, text="Waiting for jobs...",
            font=ctk.CTkFont(size=12), text_color="#7d8590",
        )
        self._summary.pack(anchor="w", pady=(0, 4))
        self._scroll = ctk.CTkScrollableFrame(self, fg_color="#0d1117", corner_radius=8)
        self._scroll.pack(fill="both", expand=True)

    def add_job(self, data: dict):
        v = data.get("verdict", "maybe").lower()
        self._counts["total"] += 1
        self._counts[v] = self._counts.get(v, 0) + 1
        card = ScraperJobCard(
            self._scroll,
            verdict=v,
            company=data.get("company", ""),
            title=data.get("title", ""),
            skills_pct=int(data.get("skills_pct", 0)),
            location=data.get("location", ""),
            matched=data.get("matched", []),
            missing=data.get("missing", []),
        )
        card.pack(fill="x", padx=4, pady=(0, 4))
        try:
            self._scroll._parent_canvas.yview_moveto(1.0)
        except Exception:
            pass
        t, y, m, n = (self._counts[k] for k in ("total", "yes", "maybe", "no"))
        self._summary.configure(
            text=f"{t} processed  |  ✓ {y} YES  |  ~ {m} MAYBE  |  ✗ {n} NO"
        )

    def clear(self):
        for w in self._scroll.winfo_children():
            w.destroy()
        self._counts = {"total": 0, "yes": 0, "maybe": 0, "no": 0}
        self._summary.configure(text="Waiting for jobs...")


# ══════════════════════════════════════════════════════════════
# PIPELINE VISUAL RESULTS
# ══════════════════════════════════════════════════════════════

class PipelineJobCard(ctk.CTkFrame):
    _STATUS = {
        "queued":   ("⏸",  "#aaa",    "#2a2a2a"),
        "building": ("🔨", "#58a6ff", "#0d2045"),
        "checking": ("🔍", "#d29922", "#3e3200"),
        "pass":     ("✅", "#3fb950", "#1a3e2a"),
        "tex_only": ("📄", "#58d4f5", "#0a2e38"),
        "failed":   ("❌", "#f85149", "#3e1a1a"),
        "error":    ("⚠",  "#f85149", "#3e1a1a"),
    }

    def __init__(self, master, company: str, title: str, **kwargs):
        super().__init__(master, corner_radius=8, fg_color="#161b22",
                         border_color="#30363d", border_width=1, **kwargs)
        self.company = company
        self.title   = title

        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=10, pady=(8, 3))
        self._badge = ctk.CTkLabel(
            hdr, text=" ⏸ QUEUED ",
            font=ctk.CTkFont(size=10, weight="bold"),
            fg_color="#2a2a2a", text_color="#aaa", corner_radius=4,
        )
        self._badge.pack(side="left")
        ctk.CTkLabel(
            hdr, text=f"  {company[:22]}  ·  {title[:35]}",
            font=ctk.CTkFont(size=12, weight="bold"), text_color="#e6edf3",
        ).pack(side="left")

        sr = ctk.CTkFrame(self, fg_color="transparent")
        sr.pack(fill="x", padx=10, pady=(0, 3))
        self._score_lbl = ctk.CTkLabel(
            sr, text="ATS: --",
            font=ctk.CTkFont(size=11), text_color="#7d8590", width=58,
        )
        self._score_lbl.pack(side="left")
        self._bar = ctk.CTkProgressBar(
            sr, height=7, corner_radius=3, fg_color="#30363d", progress_color="#30363d",
        )
        self._bar.set(0)
        self._bar.pack(side="left", fill="x", expand=True, padx=(4, 0))

        self._detail = ctk.CTkLabel(
            self, text="Waiting in queue...",
            font=ctk.CTkFont(size=10), text_color="#7d8590", anchor="w",
        )
        self._detail.pack(fill="x", padx=10, pady=(0, 8))

    def update_status(self, status: str, score: int = 0, detail: str = ""):
        icon, fg, bg = self._STATUS.get(status, ("●", "#aaa", "#2a2a2a"))
        self.configure(border_color=fg)
        self._badge.configure(text=f" {icon} {status.upper()} ", text_color=fg, fg_color=bg)
        if score > 0:
            self._score_lbl.configure(text=f"ATS: {score}/100", text_color=fg)
            self._bar.configure(progress_color=fg)
            self._bar.set(min(1.0, score / 100))
        if detail:
            self._detail.configure(text=detail)


class PipelineResultsPanel(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._cards: dict = {}
        self._counts = {"total": 0, "done": 0, "processing": 0, "queued": 0}
        self._summary = ctk.CTkLabel(
            self, text="No jobs queued yet",
            font=ctk.CTkFont(size=12), text_color="#7d8590",
        )
        self._summary.pack(anchor="w", pady=(0, 4))
        self._scroll = ctk.CTkScrollableFrame(self, fg_color="#0d1117", corner_radius=8)
        self._scroll.pack(fill="both", expand=True)

    def set_job_list(self, jobs: list):
        for w in self._scroll.winfo_children():
            w.destroy()
        self._cards.clear()
        n = len(jobs)
        self._counts = {"total": n, "done": 0, "processing": 0, "queued": n}
        for job in jobs:
            self._make_card(job["company"], job["title"])
        self._refresh_summary()

    def _make_card(self, company: str, title: str):
        key = (company, title)
        card = PipelineJobCard(self._scroll, company=company, title=title)
        card.pack(fill="x", padx=4, pady=(0, 4))
        self._cards[key] = card
        return card

    def mark_processing(self, company: str, title: str, detail: str = "Building resume..."):
        key = (company, title)
        if key not in self._cards:
            self._make_card(company, title)
            self._counts["total"] += 1
            self._counts["queued"] += 1
        self._cards[key].update_status("building", detail=detail)
        self._counts["queued"]     = max(0, self._counts["queued"] - 1)
        self._counts["processing"] = self._counts.get("processing", 0) + 1
        try:
            self._scroll._parent_canvas.yview_moveto(1.0)
        except Exception:
            pass
        self._refresh_summary()

    def update_ats(self, company: str, title: str, score: int,
                   iteration: int, max_iter: int, passed: bool, sections: list):
        key = (company, title)
        if key not in self._cards:
            self._make_card(company, title)
        status = "pass" if passed else "checking"
        detail = f"ATS iter {iteration}/{max_iter} — score {score}/100"
        if sections and not passed:
            detail += f"  |  rebuilding: {', '.join(sections)}"
        self._cards[key].update_status(status, score=score, detail=detail)
        self._refresh_summary()

    def mark_done(self, company: str, title: str, score: int,
                  status: str, sections_rebuilt: list):
        key = (company, title)
        if key not in self._cards:
            self._make_card(company, title)
        detail = f"Score: {score}/100"
        if sections_rebuilt:
            detail += f"  |  rebuilt: {', '.join(sections_rebuilt)}"
        display = "pass" if status == "done" else status
        self._cards[key].update_status(display, score=score, detail=detail)
        self._counts["processing"] = max(0, self._counts["processing"] - 1)
        self._counts["done"]       = self._counts.get("done", 0) + 1
        self._refresh_summary()

    def _refresh_summary(self):
        t = self._counts["total"]
        d = self._counts["done"]
        p = self._counts["processing"]
        q = self._counts["queued"]
        self._summary.configure(
            text=f"{t} jobs  |  ✅ {d} done  |  🔨 {p} processing  |  ⏸ {q} queued"
        )

    def clear(self):
        for w in self._scroll.winfo_children():
            w.destroy()
        self._cards.clear()
        self._counts = {"total": 0, "done": 0, "processing": 0, "queued": 0}
        self._summary.configure(text="No jobs queued yet")


# ── Status badge ──────────────────────────────────────────────
class StatusBadge(ctk.CTkLabel):
    COLORS = {
        "idle":     ("#555", "#aaa"),
        "running":  ("#1f6feb", "#58a6ff"),
        "done":     ("#1a7f37", "#3fb950"),
        "error":    ("#b62324", "#f85149"),
    }

    def __init__(self, master, **kwargs):
        super().__init__(master, text="● Idle", **kwargs)
        self.set_status("idle")

    def set_status(self, status: str, text: str = None):
        bg, fg = self.COLORS.get(status, self.COLORS["idle"])
        label  = text or status.capitalize()
        symbol = {"idle": "●", "running": "◉", "done": "✓", "error": "✗"}.get(status, "●")
        self.configure(text=f"{symbol} {label}", text_color=fg)


# ── Prompt editor dialog ──────────────────────────────────────
class PromptEditorDialog(ctk.CTkToplevel):
    def __init__(self, master, prompt_key, prompt_text, on_save):
        super().__init__(master)
        self.title(f"Edit Prompt — {prompt_key}")
        self.geometry("900x600")
        self.grab_set()
        self.on_save    = on_save
        self.prompt_key = prompt_key

        ctk.CTkLabel(self, text=f"Editing: {prompt_key}",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=20, pady=(15, 5), anchor="w")

        info = ctk.CTkLabel(
            self,
            text="⚠  The output JSON schema is automatically enforced — changing the prompt will NOT break parsing.",
            font=ctk.CTkFont(size=11),
            text_color="#FFC107",
        )
        info.pack(padx=20, pady=(0, 8), anchor="w")

        self.editor = ctk.CTkTextbox(self, font=ctk.CTkFont(family="Consolas", size=12))
        self.editor.pack(fill="both", expand=True, padx=20, pady=5)
        self.editor.insert("1.0", prompt_text)

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=10)

        ctk.CTkButton(btn_frame, text="Reset to Default", width=140,
                      fg_color="#333", hover_color="#444",
                      command=self._reset).pack(side="left")
        ctk.CTkButton(btn_frame, text="Cancel", width=100,
                      fg_color="#333", hover_color="#444",
                      command=self.destroy).pack(side="right", padx=(5, 0))
        ctk.CTkButton(btn_frame, text="Save", width=100,
                      command=self._save).pack(side="right")

    def _save(self):
        self.on_save(self.prompt_key, self.editor.get("1.0", "end-1c"))
        self.destroy()

    def _reset(self):
        default = DEFAULT_CONFIG["prompts"].get(self.prompt_key, "")
        self.editor.delete("1.0", "end")
        self.editor.insert("1.0", default)

class ProjectLinkDialog(ctk.CTkToplevel):
    def __init__(self, master, on_save):
        super().__init__(master)
        self.title("Add Project Link")
        self.geometry("400x250")
        self.grab_set()
        self.on_save = on_save

        self._build()

    def _build(self):
        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(frame, text="Project Name", font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(10, 0))
        self.name_entry = ctk.CTkEntry(frame, placeholder_text="e.g. Trading Dashboard")
        self.name_entry.pack(fill="x", pady=(4, 10))

        ctk.CTkLabel(frame, text="Project Link", font=ctk.CTkFont(size=12)).pack(anchor="w")
        self.link_entry = ctk.CTkEntry(frame, placeholder_text="https://github.com/...")
        self.link_entry.pack(fill="x", pady=(4, 20))

        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)

        ctk.CTkButton(btn_frame, text="Cancel", width=100, height=32,
                      fg_color="#333", hover_color="#444",
                      command=self.destroy).pack(side="right", padx=(5, 0))
        ctk.CTkButton(btn_frame, text="Save", width=100, height=32,
                      command=self._save).pack(side="right")

    def _save(self):
        name = self.name_entry.get().strip()
        link = self.link_entry.get().strip()
        if not name or not link:
            messagebox.showwarning("Input Error", "Please enter both project name and link.")
            return
        self.on_save(name, link)
        self.destroy()

class GitHubImportDialog(ctk.CTkToplevel):
    def __init__(self, master, on_save):
        super().__init__(master)
        self.title("Import from GitHub")
        self.geometry("450x200")
        self.grab_set()
        self.on_save = on_save

        self._build()

    def _build(self):
        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(frame, text="GitHub Repository URL", font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(10, 0))
        self.url_entry = ctk.CTkEntry(frame, placeholder_text="https://github.com/user/repo")
        self.url_entry.pack(fill="x", pady=(4, 20))

        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)

        ctk.CTkButton(btn_frame, text="Cancel", width=100, height=32,
                      fg_color="#333", hover_color="#444",
                      command=self.destroy).pack(side="right", padx=(5, 0))
        ctk.CTkButton(btn_frame, text="Import", width=100, height=32,
                      command=self._save).pack(side="right")

    def _save(self):
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showwarning("Input Error", "Please enter a valid GitHub URL.")
            return
        self.on_save(url)
        self.destroy()

class GitHubImportProgressDialog(ctk.CTkToplevel):
    def __init__(self, master, total_repos):
        super().__init__(master)
        self.title("Importing Projects")
        self.geometry("400x150")
        self.grab_set()
        self.total_repos = total_repos
        self.current_count = 0

        self._build()

    def _build(self):
        frame = ctk.CTkFrame(self, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.label = ctk.CTkLabel(frame, text=f"Initializing import...", font=ctk.CTkFont(size=13))
        self.label.pack(pady=(10, 10))

        self.progress_bar = ctk.CTkProgressBar(frame)
        self.progress_bar.pack(fill="x", pady=10)
        self.progress_bar.set(0)

    def update_progress(self, current):
        self.current_count = current
        pct = current / self.total_repos if self.total_repos > 0 else 0
        self.progress_bar.set(pct)
        self.label.configure(text=f"Processing repositories: {current} / {self.total_repos}")

    def close(self):
        self.destroy()


# ══════════════════════════════════════════════════════════════
#  TAB: Dashboard
# ══════════════════════════════════════════════════════════════
class DashboardTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="JobHunter Dashboard",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self, text="Automate your job search end-to-end.",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(anchor="w", padx=30)

        cards = ctk.CTkFrame(self, fg_color="transparent")
        cards.pack(fill="x", padx=30, pady=20)

        self._stat_card(cards, "🔍", "Sites Active",
                        str(len(self.config.get("scraper", "sites", default="").split(","))), 0)
        self._stat_card(cards, "🤖", "AI Model",
                        self.config.get("model", "name", default="—").split(":")[0], 1)
        self._stat_card(cards, "📁", "Output Folder",
                        self.config.get("pipeline", "output_dir", default="outputs"), 2)
        self._stat_card(cards, "☁", "Google Drive",
                        "Enabled" if self.config.get("google", "enabled") else "Disabled", 3)

        # Quick start guide
        guide = ctk.CTkFrame(self, corner_radius=10)
        guide.pack(fill="x", padx=30, pady=10)

        ctk.CTkLabel(guide, text="Quick Start",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=20, pady=(15, 5), anchor="w")

        steps = [
            ("1", "Configure Settings", "Set your API keys, sites, search term, and location"),
            ("2", "Run Scraper", "Scrape jobs and let AI filter the best matches"),
            ("3", "Review Jobs", "Open jobs.xlsx and check AI recommendations"),
            ("4", "Run Pipeline", "Generate tailored resumes and cover letters"),
            ("5", "Track Applications", "Results auto-uploaded to Google Drive + Sheets"),
        ]
        for num, title, desc in steps:
            row = ctk.CTkFrame(guide, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=4)
            ctk.CTkLabel(row, text=num, width=28, height=28,
                         corner_radius=14, fg_color="#1f6feb",
                         font=ctk.CTkFont(weight="bold")).pack(side="left")
            ctk.CTkLabel(row, text=f"  {title}",
                         font=ctk.CTkFont(weight="bold")).pack(side="left")
            ctk.CTkLabel(row, text=f" — {desc}",
                         text_color="gray").pack(side="left")

        ctk.CTkFrame(guide, height=15, fg_color="transparent").pack()

        # Start All Button
        start_all_frame = ctk.CTkFrame(self, fg_color="transparent")
        start_all_frame.pack(fill="x", padx=30, pady=20)

        self.start_all_btn = ctk.CTkButton(
            start_all_frame, text="🚀 Start All (Scrape → Pipeline)",
            height=48, font=ctk.CTkFont(size=15, weight="bold"),
            command=self._start_all
        )
        self.start_all_btn.pack(fill="x")

    def _start_all(self):
        app = self.master.master
        if hasattr(app, 'tabs') and "Scraper" in app.tabs:
            scraper = app.tabs["Scraper"]
            scraper.chain_to_pipeline = True
            scraper._run()

    def _stat_card(self, parent, icon, label, value, col):
        card = ctk.CTkFrame(parent, corner_radius=10, width=160, height=90)
        card.grid(row=0, column=col, padx=8, sticky="ew")
        card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(card, text=icon, font=ctk.CTkFont(size=22)).grid(row=0, column=0, pady=(12, 0))
        ctk.CTkLabel(card, text=value,
                     font=ctk.CTkFont(size=15, weight="bold")).grid(row=1, column=0)
        ctk.CTkLabel(card, text=label,
                     text_color="gray", font=ctk.CTkFont(size=11)).grid(row=2, column=0, pady=(0, 10))
        parent.grid_columnconfigure(col, weight=1)


# ── Screener helpers (used by ScraperTab and ScreenerConfigTab) ───────────

def _build_screener_custom_rules(screener_cfg: dict) -> str:
    """Return a custom-rules block to append to the screener prompt."""
    max_years  = int(screener_cfg.get("max_years_exp",   5))
    yes_pct    = int(screener_cfg.get("yes_match_pct",  70))
    maybe_pct  = int(screener_cfg.get("maybe_match_pct", 45))
    levels     = screener_cfg.get("accept_role_levels", ["junior", "mid"])
    req_sk     = [s.strip() for s in screener_cfg.get("required_skills",  "").splitlines() if s.strip()]
    pref_sk    = [s.strip() for s in screener_cfg.get("preferred_skills", "").splitlines() if s.strip()]
    reject_kw  = [s.strip() for s in screener_cfg.get("reject_keywords",  "").splitlines() if s.strip()]
    accept_kw  = [s.strip() for s in screener_cfg.get("accept_keywords",  "").splitlines() if s.strip()]

    parts = [
        "\n\n-----------------------------------------------------",
        "CUSTOM SCREENING RULES (override the defaults above)",
        "-----------------------------------------------------",
        f"• Auto-REJECT if JD requires more than {max_years} years of experience.",
        f"• Verdict thresholds: YES if skills match ≥ {yes_pct}%  |  MAYBE if ≥ {maybe_pct}%  |  NO if below {maybe_pct}%.",
    ]
    if levels:
        non = [l for l in ["junior", "mid", "senior"] if l not in levels]
        parts.append(f"• Accepted role levels: {', '.join(levels)}.")
        if non:
            parts.append(f"• Auto-REJECT (verdict=no) roles clearly labelled: {', '.join(non)}.")
    if req_sk:
        parts.append(f"• Required skills — prioritise matching these: {', '.join(req_sk)}.")
    if pref_sk:
        parts.append(f"• Preferred skills (bonus): {', '.join(pref_sk)}.")
    if reject_kw:
        parts.append(f"• Auto-REJECT immediately if title or JD contains: {', '.join(reject_kw)}.")
    if accept_kw:
        parts.append(f"• Boost verdict to YES if job title contains: {', '.join(accept_kw)}.")
    return "\n".join(parts)


def _fuzzy_dedup_jobs(df, threshold: float = 0.90):
    """Remove near-duplicate jobs (same title+company across sites, ~90% match)."""
    from difflib import SequenceMatcher
    seen, keep = [], []
    for idx, row in df.iterrows():
        t = str(row.get("title",   "")).lower().strip()
        c = str(row.get("company", "")).lower().strip()
        is_dup = any(
            SequenceMatcher(None, t, st).ratio() >= threshold and
            SequenceMatcher(None, c, sc).ratio() >= threshold
            for st, sc in seen
        )
        if not is_dup:
            seen.append((t, c))
            keep.append(idx)
    return df.loc[keep].reset_index(drop=True) if keep else df.iloc[0:0]


# ══════════════════════════════════════════════════════════════
#  TAB: Scraper
# ══════════════════════════════════════════════════════════════
class ScraperTab(ctk.CTkFrame):
    def __init__(self, master, config, log_queue, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config    = config
        self.log_queue = log_queue
        self.running   = False
        self.stop_event = threading.Event()
        self.chain_to_pipeline = False
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Job Scraper",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)

        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=30)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        # ── Left: config ──────────────────────────────────────
        left = ctk.CTkFrame(content, corner_radius=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=5)

        ctk.CTkLabel(left, text="Scraper Configuration",
                     font=ctk.CTkFont(weight="bold")).pack(padx=15, pady=(12, 8), anchor="w")

        self.fields = {}

        # ── Search Terms (multi-line) ─────────────────────────
        st_header = ctk.CTkFrame(left, fg_color="transparent")
        st_header.pack(fill="x", padx=15, pady=(6, 0))
        ctk.CTkLabel(st_header, text="Search Terms", width=130, anchor="w").pack(side="left")
        ctk.CTkLabel(st_header, text="one per line", font=ctk.CTkFont(size=10),
                     text_color="gray").pack(side="left")

        saved_terms = self.config.get(
            "scraper", "search_terms",
            default=self.config.get("scraper", "search_term", default="Software Developer")
        )
        self.search_terms_box = ctk.CTkTextbox(left, height=80, font=ctk.CTkFont(size=12),
                                               fg_color="#1c2128", border_color="#30363d",
                                               border_width=1, corner_radius=6)
        self.search_terms_box.pack(fill="x", padx=15, pady=(2, 6))
        self.search_terms_box.insert("1.0", saved_terms)

        scraper_fields = [
            ("location",       "Location",       "Canada"),
            ("country_indeed", "Country",        "canada"),
            ("hours_old",      "Hours Old",      "72"),
            ("results_wanted", "Results Wanted", "20"),
        ]
        for key, label, placeholder in scraper_fields:
            self._labeled_entry(left, key, label,
                                self.config.get("scraper", key, default=placeholder))

        # ── Sites multi-select ────────────────────────────────
        _ALL_SITES = [
            ("indeed",       "Indeed"),
            ("linkedin",     "LinkedIn"),
            ("glassdoor",    "Glassdoor"),
            ("ziprecruiter", "ZipRecruiter"),
            ("naukri",       "Naukri"),
            ("jobright",     "JobRight"),
            ("google",       "Google"),
            ("bayt",         "Bayt"),
            ("bdjobs",       "BDJobs"),
            ("wellfound",    "Wellfound"),
        ]
        saved_sites = {
            s.strip().lower()
            for s in self.config.get("scraper", "sites", default="indeed,glassdoor,jobright").split(",")
            if s.strip()
        }

        sites_row = ctk.CTkFrame(left, fg_color="transparent")
        sites_row.pack(fill="x", padx=15, pady=(6, 2))
        ctk.CTkLabel(sites_row, text="Job Sites", width=130, anchor="w").pack(side="left")

        sites_box = ctk.CTkFrame(left, fg_color="transparent")
        sites_box.pack(fill="x", padx=15, pady=(0, 4))

        self.site_vars: dict[str, ctk.BooleanVar] = {}
        for col_idx, (site_key, site_label) in enumerate(_ALL_SITES):
            var = ctk.BooleanVar(value=(site_key in saved_sites))
            self.site_vars[site_key] = var
            r, c = divmod(col_idx, 3)
            cb = ctk.CTkCheckBox(sites_box, text=site_label, variable=var,
                                 width=110, checkbox_width=16, checkbox_height=16,
                                 font=ctk.CTkFont(size=12))
            cb.grid(row=r, column=c, sticky="w", padx=4, pady=2)

        self.remote_var = ctk.BooleanVar(value=self.config.get("scraper", "is_remote", default=False))
        row = ctk.CTkFrame(left, fg_color="transparent")
        row.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(row, text="Remote Only", width=130, anchor="w").pack(side="left")
        ctk.CTkSwitch(row, variable=self.remote_var, text="").pack(side="left")

        ctk.CTkButton(left, text="Save Config", height=36,
                      fg_color="#333", hover_color="#444",
                      command=self._save_config).pack(padx=15, pady=(8, 15), fill="x")

        # ── Right: run + log ──────────────────────────────────
        right = ctk.CTkFrame(content, corner_radius=10)
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=5)

        top = ctk.CTkFrame(right, fg_color="transparent")
        top.pack(fill="x", padx=15, pady=12)
        ctk.CTkLabel(top, text="Run Scraper",
                     font=ctk.CTkFont(weight="bold")).pack(side="left")
        self.status = StatusBadge(top, font=ctk.CTkFont(size=12))
        self.status.pack(side="right")

        self.run_btn = ctk.CTkButton(
            right, text="▶  Start Scraping", height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._run
        )
        self.run_btn.pack(padx=15, fill="x")

        self.stop_btn = ctk.CTkButton(
            right, text="⏹  Stop Scraping", height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#442222", hover_color="#662222",
            command=self._stop_scraping,
            state="disabled"
        )
        self.stop_btn.pack(padx=15, pady=(10, 0), fill="x")

        ctk.CTkLabel(right, text="Results:",
                     font=ctk.CTkFont(size=11), text_color="gray").pack(padx=15, pady=(10, 2), anchor="w")

        self.results_panel = ScraperResultsPanel(right)
        self.results_panel.pack(fill="both", expand=True, padx=15, pady=(0, 4))

        ctk.CTkLabel(right, text="Activity Log",
                     font=ctk.CTkFont(size=10), text_color="#555").pack(padx=15, anchor="w")

        self.log_widget = LogWidget(right)
        self.log_widget.textbox.configure(height=80)
        self.log_widget.pack(fill="x", padx=15, pady=(2, 15))

    def _labeled_entry(self, parent, key, label, value):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(row, text=label, width=130, anchor="w").pack(side="left")
        entry = ctk.CTkEntry(row, placeholder_text=label)
        entry.insert(0, str(value))
        entry.pack(side="left", fill="x", expand=True)
        self.fields[key] = entry

    def _save_config(self):
        for key, entry in self.fields.items():
            self.config.set("scraper", key, entry.get().strip())
        terms_raw = self.search_terms_box.get("1.0", "end").strip()
        self.config.set("scraper", "search_terms", terms_raw)
        # keep legacy key in sync with first term for anything that still reads it
        first_term = next((t.strip() for t in terms_raw.splitlines() if t.strip()), "Software Developer")
        self.config.set("scraper", "search_term", first_term)
        selected_sites = ",".join(k for k, v in self.site_vars.items() if v.get())
        self.config.set("scraper", "sites", selected_sites or "indeed")
        self.config.set("scraper", "is_remote", self.remote_var.get())
        self.config.save()
        self.log_widget.append("INFO", "✓ Scraper config saved.")

    def _run(self):
        if self.running:
            return
        self._save_config()
        self.stop_event.clear()
        self.running = True
        self.run_btn.configure(state="disabled", text="⏳  Running...")
        self.stop_btn.configure(state="normal")
        self.status.set_status("running", "Scraping...")
        self.log_widget.clear()
        self.results_panel.clear()
        self._worker = threading.Thread(target=self._scrape_thread, daemon=True)
        self._worker.start()
        self.after(200, self._poll_logs)

    def _stop_scraping(self):
        self.log_widget.append("WARNING", "Stopping scraper... (finishing current job)")
        self.stop_event.set()
        self.stop_btn.configure(state="disabled")

    def _scrape_thread(self):
        try:
            sys.path.insert(0, os.getcwd())
            from jobs_scraper import scrape_all_jobs
            from db_manager import JobsDB
            import pandas as pd
            import re as _re

            cfg        = self.config.data
            excel_path = cfg["pipeline"].get("excel_path", "jobs.xlsx")
            model_name = cfg["model"].get("scraping", "gemma4:31b-cloud")
            api_keys   = cfg["model"]["api_keys"] if cfg["model"]["api_keys"] else [""]

            self.log_queue.put(("log", "INFO", "Starting job scraping..."))

            # ── Load existing Excel + build URL skip-set ──────────
            existing_df    = pd.DataFrame()
            processed_urls = set()
            if os.path.exists(excel_path):
                try:
                    existing_df    = pd.read_excel(excel_path, engine="openpyxl")
                    processed_urls = set(
                        existing_df["link"].dropna().astype(str).tolist()
                    )
                    self.log_queue.put(("log", "INFO",
                        f"Loaded {len(existing_df)} existing rows. "
                        f"{len(processed_urls)} URLs already processed."))
                except Exception as load_err:
                    self.log_queue.put(("log", "WARNING",
                        f"Could not load existing Excel: {load_err}"))

            # ── Parse search terms (one per line, skip blanks) ────
            terms_raw = cfg["scraper"].get(
                "search_terms",
                cfg["scraper"].get("search_term", "Software Developer")
            )
            search_terms = [t.strip() for t in terms_raw.splitlines() if t.strip()]
            if not search_terms:
                search_terms = ["Software Developer"]

            # ── Scrape each search term and combine ───────────────
            import pandas as _pd2
            all_frames = []
            seen_urls  = set(processed_urls)   # don't re-fetch known URLs

            for term_idx, term in enumerate(search_terms, 1):
                if self.stop_event.is_set():
                    break
                self.log_queue.put(("log", "INFO",
                    f"[{term_idx}/{len(search_terms)}] Scraping: \"{term}\""))
                try:
                    term_jobs = scrape_all_jobs(
                        site_name      = cfg["scraper"]["sites"],
                        search_term    = term,
                        location       = cfg["scraper"]["location"],
                        hours_old      = cfg["scraper"]["hours_old"],
                        results_wanted = cfg["scraper"]["results_wanted"],
                        country_indeed = cfg["scraper"]["country_indeed"],
                        is_remote      = cfg["scraper"]["is_remote"],
                    )
                    # Deduplicate within this batch against already-seen URLs
                    if len(term_jobs) > 0:
                        term_jobs = term_jobs[
                            ~term_jobs["job_url"].astype(str).isin(seen_urls)
                        ]
                        seen_urls.update(term_jobs["job_url"].astype(str).tolist())
                        all_frames.append(term_jobs)
                        self.log_queue.put(("log", "INFO",
                            f"  → {len(term_jobs)} new jobs found for \"{term}\""))
                    else:
                        self.log_queue.put(("log", "INFO",
                            f"  → No jobs found for \"{term}\""))
                except Exception as scrape_err:
                    self.log_queue.put(("log", "WARNING",
                        f"  → Scrape failed for \"{term}\": {scrape_err}"))

            if all_frames:
                jobs = _pd2.concat(all_frames, ignore_index=True)
            else:
                jobs = _pd2.DataFrame()

            if jobs.empty:
                self.log_queue.put(("log", "WARNING", "No new jobs found to screen."))
                return

            # ── Load screener config ──────────────────────────────
            screener_cfg = cfg.get("screener", {})

            # ── Blacklist filter ──────────────────────────────────
            blacklisted = [c.lower().strip()
                           for c in screener_cfg.get("blacklisted_companies", []) if c.strip()]
            if blacklisted:
                before_bl = len(jobs)
                jobs = jobs[~jobs["company"].astype(str).str.lower().str.strip().isin(blacklisted)]
                jobs = jobs.reset_index(drop=True)
                removed_bl = before_bl - len(jobs)
                if removed_bl:
                    self.log_queue.put(("log", "INFO",
                        f"Blacklist: removed {removed_bl} jobs from {len(blacklisted)} blocked companies."))

            # ── Skip already-applied (from applied.db) ────────────
            if screener_cfg.get("skip_applied", True):
                try:
                    from db_manager import AppliedDB
                    adf = AppliedDB().get_all()
                    applied_urls = set(adf["job_url"].dropna().astype(str).tolist())
                    applied_keys = set(
                        (str(r["company"]).lower().strip(), str(r["title"]).lower().strip())
                        for _, r in adf.iterrows()
                    )
                    before_ap = len(jobs)
                    url_col = "job_url" if "job_url" in jobs.columns else "link"
                    jobs = jobs[~jobs[url_col].astype(str).isin(applied_urls)]
                    # Also filter by company+title pair
                    def _not_applied(row):
                        key = (str(row.get("company","")).lower().strip(),
                               str(row.get("title","")).lower().strip())
                        return key not in applied_keys
                    jobs = jobs[jobs.apply(_not_applied, axis=1)].reset_index(drop=True)
                    removed_ap = before_ap - len(jobs)
                    if removed_ap:
                        self.log_queue.put(("log", "INFO",
                            f"Skipped {removed_ap} jobs already in applied.db."))
                except Exception as _ae:
                    self.log_queue.put(("log", "WARNING", f"Applied DB check failed: {_ae}"))

            # ── Fuzzy cross-site deduplication ────────────────────
            if screener_cfg.get("fuzzy_dedup", True) and len(jobs) > 1:
                before_fd = len(jobs)
                jobs = _fuzzy_dedup_jobs(jobs, threshold=0.90)
                removed_fd = before_fd - len(jobs)
                if removed_fd:
                    self.log_queue.put(("log", "INFO",
                        f"Fuzzy dedup: removed {removed_fd} near-duplicate jobs across sites."))

            self.log_queue.put(("log", "INFO",
                f"{len(jobs)} jobs after filtering. Running AI screener..."))

            # ── Load resume text ──────────────────────────────────
            resume_text = ""
            rp = cfg["pipeline"].get("resume_path", "resume.txt")
            if os.path.exists(rp):
                with open(rp, encoding="utf-8") as f:
                    resume_text = f.read()

            # ── Prompt setup ──────────────────────────────────────
            from ai import load_profile, build_profile_prompt_context
            prompt_template = self.config.get(
                "prompts", "job_screener",
                default=DEFAULT_CONFIG["prompts"]["job_screener"]
            )

            # Inject candidate profile into the prompt
            profile = load_profile()
            profile_ctx = build_profile_prompt_context(profile)
            prompt_template = prompt_template.replace("{candidate_profile}", profile_ctx)

            # Inject custom screener rules from ScreenerConfigTab
            custom_rules = _build_screener_custom_rules(screener_cfg)
            if custom_rules:
                prompt_template = prompt_template + custom_rules

            SCHEMA_SUFFIX = (
                "\n\nIMPORTANT: Respond with ONLY a JSON object. "
                "No explanation. No markdown. No backticks.\n"
                '{"verdict": "yes or maybe or no", '
                '"years_required": "number or unspecified", '
                '"role_level": "junior or mid or senior or unspecified", '
                '"skills_match_pct": 50, '
                '"matched_skills": [], '
                '"missing_skills": [], '
                '"reasoning": "two sentences"}'
            )

            key_idx   = 0
            processed = 0
            skipped   = 0
            yes_count = 0

            # ── ONE loop — process each job and save immediately ──
            for _, row in jobs.iterrows():
                if self.stop_event.is_set():
                    self.log_queue.put(("log", "WARNING", "🛑 Stop signal received. Finishing current item and exiting..."))
                    break
                job_url = str(row.get("job_url", "")).strip()

                # Skip already-saved URLs
                if job_url and job_url in processed_urls:
                    skipped += 1
                    continue

                # ── Build and call AI ─────────────────────────────
                try:
                    raw_desc   = str(row.get("description", ""))
                    clean_desc = _re.sub(r'\\n{3,}', '\\n\\n', raw_desc).strip()

                    prompt = (
                        prompt_template
                        .replace("{title}",       str(row.get("title", "N/A")))
                        .replace("{description}", clean_desc)
                        .replace("{resume_text}", resume_text)
                    ) + SCHEMA_SUFFIX

                    # Rotating API keys via environment variable for the assistant
                    current_key = api_keys[key_idx % len(api_keys)] if api_keys else ""
                    os.environ["OLLAMA_API_KEY"] = current_key

                    assistant = OllamaAssistant(model=model_name)
                    text = send_with_retries(assistant, prompt)
                    verdict = parse_ai_verdict(text)
                    self.log_queue.put(("log", "DEBUG", f"Raw response: {text[:80]}"))

                except Exception as ai_err:
                    err_str = str(ai_err)
                    if "429" in err_str or "rate" in err_str.lower():
                        key_idx += 1
                        self.log_queue.put(("log", "WARNING",
                            f"Rate limited — rotating to key {key_idx % len(api_keys)}"))
                    else:
                        self.log_queue.put(("log", "WARNING",
                            f"AI error: {err_str[:80]}"))
                    verdict = {
                        "verdict": "maybe", "years_required": "unspecified",
                        "role_level": "unspecified", "skills_match_pct": 50,
                        "matched_skills": [], "missing_skills": [],
                        "reasoning": f"AI error: {err_str[:100]}",
                    }

                # ── Build new row ─────────────────────────────────
                matched = verdict.get("matched_skills", [])
                missing = verdict.get("missing_skills", [])

                def s(v): return self._sanitize_for_excel(str(v)) if hasattr(self, '_sanitize_for_excel') else str(v)

                new_row = {
                    "AI_recommendation": s(verdict["verdict"]),
                    "site":              s(row.get("site", "")),
                    "company":           s(row.get("company",     "")),
                    "title":             s(row.get("title",       "")),
                    "link":              s(job_url),
                    "location":          s(row.get("location", "")),
                    "years_required":    s(verdict["years_required"]),
                    "role_level":        s(verdict["role_level"]),
                    "skills_match_pct":  s(verdict["skills_match_pct"]),
                    "matched_skills":    s(", ".join(matched) if isinstance(matched, list) else matched),
                    "missing_skills":    s(", ".join(missing) if isinstance(missing, list) else missing),
                    "reasoning":         s(verdict["reasoning"]),
                    "description":       s(row.get("description", "")),
                    "posted_date":       s(row.get("date_posted",  "")),
                }

                # ── Upsert into jobs.db ───────────────────────────
                try:
                    JobsDB().upsert(new_row)
                except Exception as _dbe:
                    self.log_queue.put(("log", "WARNING", f"SQLite upsert failed: {_dbe}"))

                # ── Append and save immediately ───────────────────
                existing_df = pd.concat(
                    [existing_df, pd.DataFrame([new_row])],
                    ignore_index=True
                )
                if job_url:
                    processed_urls.add(job_url)
                processed += 1
                if verdict["verdict"] == "yes":
                    yes_count += 1

                try:
                    existing_df.to_excel(excel_path, index=False, engine="openpyxl")
                except PermissionError:
                    self.log_queue.put(("log", "WARNING",
                        f"⚠ Excel is open — close {excel_path} to allow saving"))
                except Exception as save_err:
                    self.log_queue.put(("log", "WARNING", f"Save error: {save_err}"))

                v = verdict["verdict"].upper()
                self.log_queue.put(("log", "INFO",
                    f"[{v}] {row.get('title','')} @ {row.get('company','')} "
                    f"| {verdict['skills_match_pct']}% "
                    f"| saved ({processed} done, {skipped} skipped)"))
                _matched = verdict.get("matched_skills", [])
                _missing = verdict.get("missing_skills", [])
                self.log_queue.put(("scraper_job", {
                    "verdict":    verdict["verdict"].lower(),
                    "company":    str(row.get("company", "")),
                    "title":      str(row.get("title",   "")),
                    "skills_pct": int(verdict.get("skills_match_pct", 0)),
                    "location":   str(row.get("location", "")),
                    "matched":    _matched[:5] if isinstance(_matched, list) else [],
                    "missing":    _missing[:5] if isinstance(_missing, list) else [],
                }, None))

            # ── Format Excel once at end ──────────────────────────
            self.log_queue.put(("log", "INFO", "Formatting Excel..."))
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Font
                from openpyxl.utils import get_column_letter

                wb = load_workbook(excel_path)
                ws = wb.active

                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="1f6feb", end_color="1f6feb", fill_type="solid")
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes   = "A2"

                red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                green_fill  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

                header_map = {
                    str(cell.value).lower(): cell.column_letter
                    for cell in ws[1] if cell.value
                }

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        col = cell.column_letter
                        if col == header_map.get("ai_recommendation"):
                            val = str(cell.value).lower() if cell.value else ""
                            cell.fill = (green_fill  if val == "yes"   else
                                         red_fill    if val == "no"    else
                                         yellow_fill if val == "maybe" else cell.fill)
                        if col == header_map.get("skills_match_pct"):
                            try:
                                pct = int(cell.value)
                                cell.fill = (green_fill  if pct >= 70 else
                                             yellow_fill if pct >= 45 else orange_fill)
                            except Exception:
                                pass
                        if col == header_map.get("link") and cell.value:
                            cell.hyperlink = cell.value
                            cell.style     = "Hyperlink"

                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(
                        (len(str(c.value)) for c in ws[col_letter] if c.value), default=10
                    )
                    ws.column_dimensions[col_letter].width = max(10, min(60, max_len + 2))

                wb.save(excel_path)
                self.log_queue.put(("log", "INFO", f"✅ Excel formatted → {excel_path}"))
            except Exception as fmt_err:
                self.log_queue.put(("log", "WARNING", f"Excel formatting failed: {fmt_err}"))

            self.log_queue.put(("log", "INFO",
                f"✅ Done! {processed} processed | {yes_count} matched | "
                f"{skipped} skipped (already in Excel)"))
            self.log_queue.put(("status", "done", f"Done — {yes_count} matches"))

        except Exception as e:
            import traceback
            self.log_queue.put(("log", "ERROR", f"Scraper failed: {e}"))
            self.log_queue.put(("log", "ERROR", traceback.format_exc()[:300]))
            self.log_queue.put(("status", "error", "Failed"))

        self.log_queue.put(("done", None, None))


    def _poll_logs(self):
        while not self.log_queue.empty():
            try:
                kind, level, msg = self.log_queue.get_nowait()
                if kind == "log":
                    self.log_widget.append(level, msg)
                elif kind == "scraper_job":
                    self.results_panel.add_job(level)  # level holds data dict
                elif kind == "status":
                    self.status.set_status(level, msg)
                elif kind == "done":
                    self._finish_run()
                    if self.chain_to_pipeline:
                        self.log_widget.append("INFO", "⛓ Chaining to Pipeline...")
                        self.chain_to_pipeline = False
                        app = self.master.master
                        if hasattr(app, 'tabs') and "Pipeline" in app.tabs:
                            app.tabs["Pipeline"]._run()
                    return
            except queue.Empty:
                break
        if self.running:
            # Safety net: if the worker thread is dead but we never saw "done",
            # unstick the UI now (handles any race where the message was lost).
            if hasattr(self, '_worker') and not self._worker.is_alive():
                self.log_widget.append("WARNING", "Worker thread ended without 'done' signal — resetting UI.")
                self._finish_run()
                return
            self.after(200, self._poll_logs)

    def _finish_run(self):
        self.running = False
        self.run_btn.configure(state="normal", text="▶  Start Scraping")
        self.stop_btn.configure(state="disabled")

    def _sanitize_for_excel(self, value):
        """Prevent openpyxl crashes by removing illegal XML characters,
        handling leading '=' and truncating long strings."""
        if not isinstance(value, str):
            return value

        # 1. Remove illegal XML characters (control characters)
        # These characters cause openpyxl to crash during save
        illegal_xml_chars = [
            (0x00, 0x08), (0x0B, 0x0C), (0x0E, 0x1F),
            (0x7F, 0x84), (0x86, 0x9F), (0xFBE, 0xFFB)
        ]
        for start, end in illegal_xml_chars:
            # We use a translation table for efficiency
            value = value.translate(str.maketrans('', '', ''.join(chr(i) for i in range(start, end + 1))))

        # 2. Excel treats cells starting with '=' as formulas.
        # Prepend single quote to force string literal in Excel
        if value.startswith('='):
            value = "'" + value

        # 3. Excel cell limit is 32,767 characters.
        if len(value) > 32767:
            value = value[:32760] + "..."

        return value


# ══════════════════════════════════════════════════════════════
#  TAB: Pipeline
# ══════════════════════════════════════════════════════════════
class PipelineTab(ctk.CTkFrame):
    def __init__(self, master, config, log_queue, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config    = config
        self.log_queue = log_queue
        self.running   = False
        self.stop_event = threading.Event()
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Resume Pipeline",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)

        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=30)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        # ── Left: settings ────────────────────────────────────
        left = ctk.CTkFrame(content, corner_radius=10)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=5)

        ctk.CTkLabel(left, text="Pipeline Settings",
                     font=ctk.CTkFont(weight="bold")).pack(padx=15, pady=(12, 8), anchor="w")

        self.pipe_fields = {}
        pipe_cfg = [
            ("excel_path",              "Jobs Excel",          "jobs.xlsx"),
            ("applied_excel_path",      "Applied Excel",       "Job-Tracker.xlsx"),
            ("applied_folder_path",    "Applied Folder",      "fultime"),
            ("output_dir",              "Output Folder",       "outputs"),
            ("resume_path",             "Resume File",         "resume.txt"),
            ("projects_path",           "Projects File",       "Projects.txt"),
            ("resume_filename",         "Resume Filename",     "Sneh_Resume"),        # ← ADD
            ("cover_letter_filename",   "Cover Letter Name",   "Sneh_Cover_Letter"),  # ← ADD,
            ("default_location",    "Default Location",    "Canada"),
            ]
        for key, label, ph in pipe_cfg:
            row = ctk.CTkFrame(left, fg_color="transparent")
            row.pack(fill="x", padx=15, pady=4)
            ctk.CTkLabel(row, text=label, width=110, anchor="w").pack(side="left")
            entry = ctk.CTkEntry(row)
            entry.insert(0, self.config.get("pipeline", key, default=ph))
            entry.pack(side="left", fill="x", expand=True)

            btn = ctk.CTkButton(row, text="📂", width=30, fg_color="#333",
                                command=lambda e=entry, k=key: self._browse(e, k))
            btn.pack(side="left", padx=(3, 0))
            self.pipe_fields[key] = entry


        # Location Setting
        self.use_jd_loc_var = ctk.BooleanVar(value=self.config.get("pipeline", "use_jd_location", default=True))
        loc_row = ctk.CTkFrame(left, fg_color="transparent")
        loc_row.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(loc_row, text="Use JD Location", width=110, anchor="w").pack(side="left")
        ctk.CTkSwitch(loc_row, variable=self.use_jd_loc_var, text="").pack(side="left")

        # ATS settings
        ctk.CTkLabel(left, text="ATS Settings",
                     font=ctk.CTkFont(weight="bold")).pack(padx=15, pady=(12, 4), anchor="w")

        row = ctk.CTkFrame(left, fg_color="transparent")
        row.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(row, text="Max Iterations", width=110, anchor="w").pack(side="left")
        self.max_iter = ctk.CTkEntry(row, width=60)
        self.max_iter.insert(0, str(self.config.get("pipeline", "max_ats_iterations", default=2)))
        self.max_iter.pack(side="left")

        row2 = ctk.CTkFrame(left, fg_color="transparent")
        row2.pack(fill="x", padx=15, pady=4)
        ctk.CTkLabel(row2, text="Pass Threshold", width=110, anchor="w").pack(side="left")
        self.threshold = ctk.CTkEntry(row2, width=60)
        self.threshold.insert(0, str(self.config.get("pipeline", "ats_pass_threshold", default=85)))
        self.threshold.pack(side="left")
        ctk.CTkLabel(row2, text="/100", text_color="gray").pack(side="left", padx=4)

        ctk.CTkButton(left, text="Save Settings", height=36, fg_color="#333",
                      hover_color="#444",
                      command=self._save_config).pack(padx=15, pady=(12, 15), fill="x")

        # ── Right: run + log ──────────────────────────────────
        right = ctk.CTkFrame(content, corner_radius=10)
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=5)

        top = ctk.CTkFrame(right, fg_color="transparent")
        top.pack(fill="x", padx=15, pady=12)
        ctk.CTkLabel(top, text="Run Pipeline",
                     font=ctk.CTkFont(weight="bold")).pack(side="left")
        self.status = StatusBadge(top, font=ctk.CTkFont(size=12))
        self.status.pack(side="right")

        self.run_btn = ctk.CTkButton(
            right, text="▶  Generate Resumes & Cover Letters", height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._run
        )
        self.run_btn.pack(padx=15, fill="x")

        self.stop_btn = ctk.CTkButton(
            right, text="⏹  Stop Pipeline", height=42,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#442222", hover_color="#662222",
            command=self._stop_pipeline,
            state="disabled"
        )
        self.stop_btn.pack(padx=15, pady=(10, 0), fill="x")

        ctk.CTkLabel(right, text="Pipeline Jobs:",
                     font=ctk.CTkFont(size=11), text_color="gray").pack(padx=15, pady=(10, 2), anchor="w")

        self.results_panel = PipelineResultsPanel(right)
        self.results_panel.pack(fill="both", expand=True, padx=15, pady=(0, 4))

        ctk.CTkLabel(right, text="Activity Log",
                     font=ctk.CTkFont(size=10), text_color="#555").pack(padx=15, anchor="w")

        self.log_widget = LogWidget(right)
        self.log_widget.textbox.configure(height=80)
        self.log_widget.pack(fill="x", padx=15, pady=(2, 15))

    def _browse(self, entry, key):
        if "dir" in key or "folder" in key.lower():
            path = filedialog.askdirectory()
        else:
            path = filedialog.askopenfilename()
        if path:
            entry.delete(0, "end")
            entry.insert(0, path)

    def _save_config(self):
        for key, entry in self.pipe_fields.items():
            self.config.set("pipeline", key, entry.get().strip())
        try:
            self.config.set("pipeline", "max_ats_iterations", int(self.max_iter.get()))
            self.config.set("pipeline", "ats_pass_threshold", int(self.threshold.get()))
        except ValueError:
            pass
        self.config.save()
        self.log_widget.append("INFO", "✓ Pipeline config saved.")

        self.config.set("pipeline", "use_jd_location", self.use_jd_loc_var.get())
    def _run(self):
        if self.running:
            return
        self._save_config()
        self.stop_event.clear()
        self.running = True
        self.run_btn.configure(state="disabled", text="⏳  Running...")
        self.stop_btn.configure(state="normal")
        self.status.set_status("running", "Processing...")
        self.log_widget.clear()
        self.results_panel.clear()

        q = self.log_queue
        self._worker = threading.Thread(target=self._pipeline_thread, args=(q,), daemon=True)
        self._worker.start()
        self.after(200, self._poll_logs)

    def _stop_pipeline(self):
        self.log_widget.append("WARNING", "Stopping pipeline... (finishing current job)")
        self.stop_event.set()
        self.stop_btn.configure(state="disabled")

    def _pipeline_thread(self, q):
        try:
            sys.path.insert(0, os.getcwd())

            q.put(("log", "INFO", "Starting pipeline..."))

            # Inject custom prompts into environment before importing agents
            self._inject_prompts()

            from pipeline import main as pipeline_main
            pipeline_main(stop_event=self.stop_event, event_queue=q)

            q.put(("log", "INFO", "✅ Pipeline complete!"))
            q.put(("status", "done", "Done"))

        except Exception as e:
            q.put(("log", "ERROR", f"Pipeline failed: {e}"))
            import traceback
            q.put(("log", "ERROR", traceback.format_exc()[:500]))
            q.put(("status", "error", "Failed"))

        q.put(("done", None, None))

    def _inject_prompts(self):
        try:
            import agents.resume_builder as rb
            import re as _re
            prompts = self.config.get("prompts", default={})
            roles   = self.config.get("experience_roles", default=None)

            # 1. Experience Prompt (Inject Roles)
            exp_prompt = prompts.get("experience_section", rb.PROMPT_EXPERIENCE)
            if roles:
                roles_text = "\n".join([
                    f"Role {i+1}: {r['title']} @ {r['company']} | {r['dates']}\n"
                    f"→ {r['total_bullets']} bullets: {r['fabricated_bullets']} fabricated + "
                    f"{r['real_bullets']} real ({r['domain']} domain)"
                    for i, r in enumerate(roles)
                ])
                exp_prompt = _re.sub(
                    r'Role 1:.*?(?=\n━━━|\nBULLET RULES|\nRAW LaTeX)',
                    roles_text,
                    exp_prompt,
                    flags=_re.DOTALL
                )
            rb.PROMPT_EXPERIENCE = exp_prompt

            # 2. Projects Prompt (Inject Links)
            profile = self.config.get("profile", default={})
            include_proj_links = profile.get("include_project_links", True)
            proj_prompt = prompts.get("projects_section", rb.PROMPT_PROJECTS)

            if include_proj_links:
                # Replace standard heading with href version and force empty second arg
                proj_prompt = proj_prompt.replace(
                    "\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}",
                    "\\href{url}{{\\textbf{Project Name}}} $|$ \\emph{{Tech1, Tech2, Tech3}}"
                )
                proj_prompt += "\n\nCRITICAL: For project headings, you MUST use the format: \\resumeProjectHeading{\\href{ACTUAL_URL}{{\\textbf{Project Name}}} $|$ \\emph{Techs}}{{}}. Replace 'ACTUAL_URL' with the real link found in the project data. The second set of braces must be EMPTY."
            else:
                # Remove href if present and force empty second arg
                proj_prompt = proj_prompt.replace(
                    "\\href{url}{{\\textbf{Project Name}}} $|$ \\emph{{Tech1, Tech2, Tech3}}",
                    "\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}"
                )
                proj_prompt += "\n\nCRITICAL: For project headings, you MUST use the format: \\resumeProjectHeading{\\textbf{Project Name} $|$ \\emph{Techs}}{{}}. The second set of braces must be EMPTY."
            rb.PROMPT_PROJECTS = proj_prompt

            # 3. Other prompts (Simple override)
            rb.PROMPT_SKILLS = prompts.get("skills_section", rb.PROMPT_SKILLS)
            rb.PROMPT_COVER_LETTER = prompts.get("cover_letter", rb.PROMPT_COVER_LETTER)

        except Exception as e:
            self.log_queue.put(("log", "WARNING", f"Could not inject prompts: {e}"))
            
    def _poll_logs(self):
        while not self.log_queue.empty():
            try:
                kind, level, msg = self.log_queue.get_nowait()
                if kind == "log":
                    self.log_widget.append(level, msg)
                elif kind == "pipeline_event":
                    self._handle_pipeline_event(level, msg)
                elif kind == "status":
                    self.status.set_status(level, msg)
                elif kind == "done":
                    self._finish_run()
                    return
            except queue.Empty:
                break
        if self.running:
            if hasattr(self, '_worker') and not self._worker.is_alive():
                self.log_widget.append("WARNING", "Worker thread ended without 'done' signal — resetting UI.")
                self._finish_run()
                return
            self.after(200, self._poll_logs)

    def _handle_pipeline_event(self, event_type: str, data):
        if not isinstance(data, dict):
            return
        company = data.get("company", "")
        title   = data.get("title",   "")

        if event_type == "job_list":
            self.results_panel.set_job_list(data.get("jobs", []))

        elif event_type == "job_start":
            idx   = data.get("index", 1)
            total = data.get("total", 1)
            self.results_panel.mark_processing(
                company, title, detail=f"Building resume... ({idx}/{total})"
            )

        elif event_type == "ats_score":
            self.results_panel.update_ats(
                company, title,
                score=data.get("score", 0),
                iteration=data.get("iteration", 1),
                max_iter=data.get("max_iter", 3),
                passed=data.get("pass", False),
                sections=data.get("sections_to_rewrite", []),
            )

        elif event_type == "job_done":
            self.results_panel.mark_done(
                company, title,
                score=data.get("score", 0),
                status=data.get("status", "done"),
                sections_rebuilt=data.get("sections_rebuilt", []),
            )

    def _finish_run(self):
        self.running = False
        self.run_btn.configure(state="normal", text="▶  Generate Resumes & Cover Letters")
        self.stop_btn.configure(state="disabled")


# ══════════════════════════════════════════════════════════════
#  TAB: Prompts
# ══════════════════════════════════════════════════════════════
class PromptsTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Prompt Editor",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self,
                     text="Customize AI prompts. Output schemas are always enforced — prompts cannot break the app.",
                     text_color="gray", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=30, pady=(0, 15))

        prompt_info = {
            "job_screener":      ("🔍", "Job Screener", "Evaluates whether to apply to a job"),
            "skills_section":    ("🛠", "Skills Section", "Generates the Technical Skills LaTeX section"),
            "experience_section":("💼", "Experience Section", "Generates the Experience LaTeX section"),
            "projects_section":  ("📦", "Projects Section", "Generates the Relevant Projects LaTeX section"),
            "cover_letter":      ("✉", "Cover Letter", "Generates the plain-text cover letter"),
        }

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=30)

        for key, (icon, title, desc) in prompt_info.items():
            card = ctk.CTkFrame(scroll, corner_radius=10)
            card.pack(fill="x", pady=6)

            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=15, pady=12)

            ctk.CTkLabel(row, text=f"{icon}  {title}",
                         font=ctk.CTkFont(size=13, weight="bold")).pack(side="left")
            ctk.CTkLabel(row, text=desc, text_color="gray").pack(side="left", padx=10)

            btn_frame = ctk.CTkFrame(row, fg_color="transparent")
            btn_frame.pack(side="right")

            ctk.CTkButton(
                btn_frame, text="Reset", width=70, height=30,
                fg_color="#333", hover_color="#444",
                command=lambda k=key: self._reset_prompt(k)
            ).pack(side="left", padx=4)

            ctk.CTkButton(
                btn_frame, text="Edit Prompt", width=110, height=30,
                command=lambda k=key: self._open_editor(k)
            ).pack(side="left")

            # Preview
            current = self.config.get("prompts", key, default="")
            preview = current[:120].replace("\n", " ").strip() + "..."
            ctk.CTkLabel(card, text=preview, text_color="#777",
                         font=ctk.CTkFont(family="Consolas", size=10),
                         anchor="w", wraplength=700).pack(padx=15, pady=(0, 10), anchor="w")

    def _open_editor(self, key):
        current = self.config.get("prompts", key,
                                  default=DEFAULT_CONFIG["prompts"].get(key, ""))
        PromptEditorDialog(self, key, current, self._save_prompt)

    def _save_prompt(self, key, text):
        self.config.set("prompts", key, text)
        self.config.save()
        messagebox.showinfo("Saved", f"Prompt '{key}' saved successfully.")
        # Rebuild to refresh previews
        for widget in self.winfo_children():
            widget.destroy()
        self._build()

    def _reset_prompt(self, key):
        if messagebox.askyesno("Reset", f"Reset '{key}' to default?"):
            default = DEFAULT_CONFIG["prompts"].get(key, "")
            self.config.set("prompts", key, default)
            self.config.save()
            for widget in self.winfo_children():
                widget.destroy()
            self._build()


class CustomSectionDialog(ctk.CTkToplevel):
    def __init__(self, master, on_save, existing_section=None, config=None):
        super().__init__(master)
        self.title("Add/Edit Custom Section")
        self.geometry("820x680")
        self.grab_set()
        self.on_save = on_save
        self.existing_section = existing_section
        self.config = config
        self._ai_thread = None
        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=20)

        # ── Section name ────────────────────────────────────────
        ctk.CTkLabel(scroll, text="Section Name",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(0, 8))

        name_row = ctk.CTkFrame(scroll, fg_color="transparent")
        name_row.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(name_row, text="Display Name", width=120, anchor="w").pack(side="left")
        self.name_entry = ctk.CTkEntry(name_row, placeholder_text='e.g. "Certifications"')
        self.name_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        if self.existing_section:
            self.name_entry.insert(0, self.existing_section.get("name", ""))

        # Auto-generate on focus-out of name field
        self.name_entry.bind("<FocusOut>", self._on_name_focus_out)

        id_row = ctk.CTkFrame(scroll, fg_color="transparent")
        id_row.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(id_row, text="Internal ID", width=120, anchor="w").pack(side="left")
        self.id_entry = ctk.CTkEntry(id_row, placeholder_text='e.g. "certifications"')
        self.id_entry.pack(side="left", fill="x", expand=True)
        if self.existing_section:
            self.id_entry.insert(0, self.existing_section.get("id", ""))
            self.id_entry.configure(state="disabled")

        ctk.CTkFrame(scroll, height=1, fg_color="#30363d").pack(fill="x", pady=(0, 12))

        # ── Status / AI button row ───────────────────────────────
        ai_row = ctk.CTkFrame(scroll, fg_color="transparent")
        ai_row.pack(fill="x", pady=(0, 10))

        self._status_label = ctk.CTkLabel(ai_row,
            text="💡 Type a section name above — prompts will auto-fill when you tab away.",
            text_color="#8b949e", font=ctk.CTkFont(size=12), anchor="w")
        self._status_label.pack(side="left", fill="x", expand=True)

        self._ai_btn = ctk.CTkButton(ai_row, text="✨ Enhance with AI", width=160, height=32,
                                     fg_color="#1f4e79", hover_color="#2563a8",
                                     command=self._enhance_with_ai)
        self._ai_btn.pack(side="right")

        # ── System Prompt ────────────────────────────────────────
        sp_hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        sp_hdr.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(sp_hdr, text="System Prompt",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left")
        ctk.CTkLabel(sp_hdr, text="(auto-generated — controls LaTeX output format)",
                     text_color="#555", font=ctk.CTkFont(size=11)).pack(side="left", padx=8)

        self.sys_prompt_box = ctk.CTkTextbox(scroll, height=170,
                                             font=ctk.CTkFont(family="Consolas", size=11))
        self.sys_prompt_box.pack(fill="x", pady=(0, 14))
        if self.existing_section:
            self.sys_prompt_box.insert("1.0", self.existing_section.get("system_prompt", ""))

        # ── User Prompt ──────────────────────────────────────────
        up_hdr = ctk.CTkFrame(scroll, fg_color="transparent")
        up_hdr.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(up_hdr, text="User Prompt",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left")
        ctk.CTkLabel(up_hdr, text="(customize freely — this is what the AI sees per job)",
                     text_color="#555", font=ctk.CTkFont(size=11)).pack(side="left", padx=8)

        self.user_prompt_box = ctk.CTkTextbox(scroll, height=170,
                                              font=ctk.CTkFont(family="Consolas", size=11))
        self.user_prompt_box.pack(fill="x", pady=(0, 14))
        if self.existing_section:
            self.user_prompt_box.insert("1.0", self.existing_section.get("user_prompt", ""))

        # ── Placeholder reference ────────────────────────────────
        ref = ctk.CTkFrame(scroll, fg_color="#161b22", corner_radius=6)
        ref.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(ref, text="Available placeholders:  {full_name}  {title}  {company}"
                               "  {description}  {existing_resume}  {ats_feedback}",
                     text_color="#58a6ff", font=ctk.CTkFont(family="Consolas", size=11)).pack(padx=12, pady=8)

        # ── Buttons ──────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(4, 0))
        ctk.CTkButton(btn_frame, text="Cancel", width=100,
                      fg_color="#333", hover_color="#444",
                      command=self.destroy).pack(side="right", padx=5)
        ctk.CTkButton(btn_frame, text="💾 Save Section", width=150,
                      command=self._save).pack(side="right")

        # If adding new section with no name yet, nothing to auto-fill
        if not self.existing_section:
            self.name_entry.focus_set()

    # ── Prompt generation ─────────────────────────────────────────────

    def _on_name_focus_out(self, _event=None):
        name = self.name_entry.get().strip()
        if not name:
            return
        # Auto-set ID from name if not editing
        if not self.existing_section:
            auto_id = name.lower().replace(" ", "_")
            self.id_entry.delete(0, "end")
            self.id_entry.insert(0, auto_id)
        # Only auto-fill prompts when boxes are empty (don't overwrite existing)
        current_sys = self.sys_prompt_box.get("1.0", "end-1c").strip()
        if not current_sys:
            sys_p, user_p = self._build_template_prompts(name)
            self._fill_boxes(sys_p, user_p)
            self._status_label.configure(
                text=f'✅ Prompts generated for "{name}". Click ✨ Enhance with AI for a better version.',
                text_color="#3fb950")

    @staticmethod
    def _build_template_prompts(section_name: str):
        """Instant template-based generation — no AI needed."""
        name = section_name.strip()
        sys_p = (
            f"You are an expert resume writer. Output ONLY raw LaTeX — no backticks, no explanation. "
            f"Generate ONLY the {name} section body.\n\n"
            f"OUTPUT FORMAT CONTRACT:\n"
            f"\\\\section{{{{{name}}}}}\n"
            f"\\\\resumeSubHeadingListStart\n"
            f"  \\\\resumeItem{{{{Specific, quantified point 1}}}}\n"
            f"  \\\\resumeItem{{{{Specific, quantified point 2}}}}\n"
            f"\\\\resumeSubHeadingListEnd\n\n"
            f"ABSOLUTE RULES:\n"
            f"- NO \\\\documentclass, NO \\\\usepackage, NO \\\\begin{{{{document}}}}, NO \\\\end{{{{document}}}}\n"
            f"- Use \\\\% for percent signs, \\\\& for ampersands\n"
            f"- Maximum 4 bullet points\n"
            f"- Every point must be specific and relevant to the target job"
        )
        user_p = (
            f"Write a {name} section for {{full_name}} applying for {{title}} at {{company}}.\n\n"
            f"JOB DESCRIPTION:\n{{description}}\n\n"
            f"CANDIDATE RESUME:\n{{existing_resume}}\n\n"
            f"ATS FEEDBACK:\n{{ats_feedback}}\n\n"
            f"Select only the most relevant items for this specific role. Be concise and impactful."
        )
        return sys_p, user_p

    def _fill_boxes(self, sys_p: str, user_p: str):
        self.sys_prompt_box.delete("1.0", "end")
        self.sys_prompt_box.insert("1.0", sys_p)
        self.user_prompt_box.delete("1.0", "end")
        self.user_prompt_box.insert("1.0", user_p)

    def _enhance_with_ai(self):
        name = self.name_entry.get().strip()
        if not name:
            messagebox.showwarning("Missing Name", "Enter the section display name first.")
            return
        if self._ai_thread and self._ai_thread.is_alive():
            return

        self._ai_btn.configure(text="⏳ Generating...", state="disabled")
        self._status_label.configure(text="🤖 AI is writing tailored prompts...", text_color="#d29922")

        import threading
        self._ai_thread = threading.Thread(target=self._run_ai_generation, args=(name,), daemon=True)
        self._ai_thread.start()

    def _run_ai_generation(self, section_name: str):
        try:
            model = "gemma4:31b-cloud"
            if self.config:
                model = self.config.get("model", "pipeline", default=model)

            assistant = OllamaAssistant(model=model)

            meta_prompt = f"""You are a resume AI prompt engineer. Generate professional prompts for a resume section called "{section_name}".

Output ONLY a JSON object (no markdown, no backticks, no explanation):
{{
  "system_prompt": "...",
  "user_prompt": "..."
}}

Rules for system_prompt:
- Start exactly with: You are an expert resume writer. Output ONLY raw LaTeX — no backticks, no explanation. Generate ONLY the {section_name} section body.
- Include an OUTPUT FORMAT CONTRACT showing the exact LaTeX structure (use \\\\section, \\\\resumeSubHeadingListStart, \\\\resumeItem, etc.)
- End with ABSOLUTE RULES listing: no \\\\documentclass, no \\\\usepackage, no \\\\begin{{document}}, no \\\\end{{document}}
- Tailor the LaTeX format to what makes sense for a "{section_name}" section specifically

Rules for user_prompt:
- Ask the AI to write the {section_name} section for {{full_name}} applying for the role of {{title}} at {{company}}
- Include these exact placeholder blocks: JOB DESCRIPTION:\\n{{description}}\\nCANDIDATE RESUME:\\n{{existing_resume}}\\nATS FEEDBACK:\\n{{ats_feedback}}
- Add specific guidance about what to include in a {section_name} section
- Keep it focused and practical"""

            response = send_with_retries(assistant, meta_prompt, tries=2, backoff_sec=1.0)

            # Parse JSON from response
            import re as _re, json as _json
            txt = response.strip()
            if "```" in txt:
                m = _re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", txt)
                if m:
                    txt = m.group(1).strip()
            brace_s = txt.find("{")
            brace_e = txt.rfind("}")
            if brace_s != -1 and brace_e > brace_s:
                data = _json.loads(txt[brace_s:brace_e + 1])
                sys_p = data.get("system_prompt", "")
                user_p = data.get("user_prompt", "")
                if sys_p and user_p:
                    self.after(0, lambda: self._on_ai_done(sys_p, user_p))
                    return

            raise ValueError("Could not parse AI response")

        except Exception as e:
            self.after(0, lambda: self._on_ai_error(str(e)))

    def _on_ai_done(self, sys_p: str, user_p: str):
        self._fill_boxes(sys_p, user_p)
        self._ai_btn.configure(text="✨ Enhance with AI", state="normal")
        self._status_label.configure(
            text="✅ AI-enhanced prompts ready. Edit the User Prompt freely.",
            text_color="#3fb950")

    def _on_ai_error(self, err: str):
        self._ai_btn.configure(text="✨ Enhance with AI", state="normal")
        self._status_label.configure(
            text=f"⚠️ AI failed ({err[:60]}). Template prompts are still usable.",
            text_color="#d29922")

    # ── Save ─────────────────────────────────────────────────────────

    def _save(self):
        sec_id = self.id_entry.get().strip().lower().replace(" ", "_")
        name = self.name_entry.get().strip()
        sys_p = self.sys_prompt_box.get("1.0", "end-1c").strip()
        user_p = self.user_prompt_box.get("1.0", "end-1c").strip()

        if not sec_id or not name:
            messagebox.showwarning("Input Error", "ID and Name are required.")
            return
        if not sys_p or not user_p:
            messagebox.showwarning("Missing Prompts",
                "Prompts are empty. Tab out of the name field to auto-generate, or type them manually.")
            return

        self.on_save({"id": sec_id, "name": name,
                      "system_prompt": sys_p, "user_prompt": user_p})
        self.destroy()

class CustomSectionsTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Custom Resume Sections",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self,
                     text="Add additional AI-generated sections to your resume (e.g. Summary, Achievements, Certifications).",
                     text_color="gray", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=30, pady=(0, 15))

        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=30, pady=(0, 10))
        ctk.CTkButton(top_bar, text="➕ Add New Section", height=32,
                      command=self._add_section).pack(side="left")

        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=30, pady=(0, 20))
        self._refresh_list()

    def _refresh_list(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        sections = self.config.get("custom_sections", default=[])
        if not sections:
            ctk.CTkLabel(self.scroll, text="No custom sections added yet.", text_color="gray").pack(pady=20)
            return

        for idx, sec in enumerate(sections):
            card = ctk.CTkFrame(self.scroll, corner_radius=10)
            card.pack(fill="x", pady=6)

            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=15, pady=12)

            ctk.CTkLabel(row, text=f"{sec.get('name', 'Unnamed')} ({sec.get('id', 'n/a')})",
                         font=ctk.CTkFont(weight="bold")).pack(side="left")

            btn_frame = ctk.CTkFrame(row, fg_color="transparent")
            btn_frame.pack(side="right")

            ctk.CTkButton(btn_frame, text="Edit", width=60, height=28, fg_color="#333",
                          command=lambda s=sec: self._edit_section(s)).pack(side="left", padx=4)
            ctk.CTkButton(btn_frame, text="Delete", width=60, height=28, fg_color="#442222",
                          command=lambda i=idx: self._delete_section(i)).pack(side="left", padx=4)

    def _add_section(self):
        CustomSectionDialog(self, self._save_section, config=self.config)

    def _edit_section(self, section):
        CustomSectionDialog(self, self._save_section, existing_section=section, config=self.config)

    def _save_section(self, sec_data):
        sections = self.config.get("custom_sections", default=[])
        # Update if exists, else append
        updated = False
        for i, s in enumerate(sections):
            if s.get("id") == sec_data["id"]:
                sections[i] = sec_data
                updated = True
                break
        if not updated:
            sections.append(sec_data)

        self.config.set("custom_sections", sections)
        self.config.save()
        self._refresh_list()
        messagebox.showinfo("Saved", "Custom section updated successfully.")

    def _delete_section(self, index):
        if messagebox.askyesno("Delete", "Are you sure you want to remove this section?"):
            sections = self.config.get("custom_sections", default=[])
            if 0 <= index < len(sections):
                sections.pop(index)
                self.config.set("custom_sections", sections)
                self.config.save()
                self._refresh_list()


class SettingsTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Settings",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=30)

        # ── AI Model ──────────────────────────────────────────
        self._section(scroll, "🤖 AI Model")

        model_card = ctk.CTkFrame(scroll, corner_radius=10)
        model_card.pack(fill="x", pady=(0, 12))

        # Scraping Model
        row = ctk.CTkFrame(model_card, fg_color="transparent")
        row.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(row, text="Scraping Model", width=130, anchor="w").pack(side="left")
        self.model_scraping = ctk.CTkEntry(row)
        self.model_scraping.insert(0, self.config.get("model", "scraping", default="gemma4:31b-cloud"))
        self.model_scraping.pack(side="left", fill="x", expand=True)

        # Pipeline Model
        row2 = ctk.CTkFrame(model_card, fg_color="transparent")
        row2.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(row2, text="Pipeline Model", width=130, anchor="w").pack(side="left")
        self.model_pipeline = ctk.CTkEntry(row2)
        self.model_pipeline.insert(0, self.config.get("model", "pipeline", default="gemma4:31b-cloud"))
        self.model_pipeline.pack(side="left", fill="x", expand=True)

        row3 = ctk.CTkFrame(model_card, fg_color="transparent")
        row3.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(row3, text="API Keys (one per line)", width=130, anchor="w").pack(side="left", anchor="n")
        self.api_keys_box = ctk.CTkTextbox(row3, height=80)
        keys = self.config.get("model", "api_keys", default=[])
        self.api_keys_box.insert("1.0", "\n".join(keys))
        self.api_keys_box.pack(side="left", fill="x", expand=True)


        for label, key, default in [
            ("Context Window", "num_ctx", "32768"),
            ("Max Tokens", "num_predict", "64384"),
            ("Temperature", "temperature", "0.3"),
        ]:
            r = ctk.CTkFrame(model_card, fg_color="transparent")
            r.pack(fill="x", padx=15, pady=4)
            ctk.CTkLabel(r, text=label, width=130, anchor="w").pack(side="left")
            e = ctk.CTkEntry(r, width=100)
            e.insert(0, str(self.config.get("model", key, default=default)))
            e.pack(side="left")
            setattr(self, f"model_{key}", e)

        ctk.CTkButton(model_card, text="Save Model Settings", height=36,
                      command=self._save_model).pack(padx=15, pady=(8, 15), fill="x")

        # ── Google Integration ────────────────────────────────
        self._section(scroll, "☁ Google Drive & Sheets")

        google_card = ctk.CTkFrame(scroll, corner_radius=10)
        google_card.pack(fill="x", pady=(0, 12))

        g_row = ctk.CTkFrame(google_card, fg_color="transparent")
        g_row.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(g_row, text="Enable Integration", width=140, anchor="w").pack(side="left")
        self.google_enabled = ctk.BooleanVar(
            value=self.config.get("google", "enabled", default=False))
        ctk.CTkSwitch(g_row, variable=self.google_enabled, text="").pack(side="left")

        self.google_fields = {}
        google_cfg = [
            ("spreadsheet_id", "Spreadsheet ID", "1LedHC091RPLwoDcByN8xo..."),
            ("drive_folder",   "Drive Folder",   "Job Applications"),
            ("sheet_tab",      "Sheet Tab",      "Sheet1"),
        ]
        for key, label, ph in google_cfg:
            r = ctk.CTkFrame(google_card, fg_color="transparent")
            r.pack(fill="x", padx=15, pady=4)
            ctk.CTkLabel(r, text=label, width=140, anchor="w").pack(side="left")
            e = ctk.CTkEntry(r, placeholder_text=ph)
            e.insert(0, self.config.get("google", key, default=""))
            e.pack(side="left", fill="x", expand=True)
            self.google_fields[key] = e

        btn_row = ctk.CTkFrame(google_card, fg_color="transparent")
        btn_row.pack(fill="x", padx=15, pady=(8, 15))
        ctk.CTkButton(btn_row, text="Save Google Settings", height=36,
                      command=self._save_google).pack(side="left", fill="x", expand=True, padx=(0, 5))
        ctk.CTkButton(btn_row, text="Test Auth", height=36, width=100,
              fg_color="#333", hover_color="#444",
              command=self._test_google_auth).pack(side="left")

        # ── GitHub Integration ──────────────────────────────────
        self._section(scroll, "🐙 GitHub API")
        gh_card = ctk.CTkFrame(scroll, corner_radius=10)
        gh_card.pack(fill="x", pady=(0, 12))

        gh_row = ctk.CTkFrame(gh_card, fg_color="transparent")
        gh_row.pack(fill="x", padx=15, pady=15)
        ctk.CTkLabel(gh_row, text="Personal Access Token", width=140, anchor="w").pack(side="left")
        self.github_token_entry = ctk.CTkEntry(gh_row, placeholder_text="ghp_...", show="*")
        self.github_token_entry.insert(0, self.config.get("github", "token", default=""))
        self.github_token_entry.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(gh_card, text="Save GitHub Token", height=36,
                      command=self._save_github).pack(padx=15, pady=(0, 15), fill="x")

        # ── Appearance ────────────────────────────────────────
        self._section(scroll, "🎨 Appearance")
        app_card = ctk.CTkFrame(scroll, corner_radius=10)
        app_card.pack(fill="x", pady=(0, 12))
        r = ctk.CTkFrame(app_card, fg_color="transparent")
        r.pack(fill="x", padx=15, pady=12)
        ctk.CTkLabel(r, text="Theme", width=100, anchor="w").pack(side="left")
        self.theme = ctk.CTkSegmentedButton(r, values=["dark", "light", "system"], height=28,
                                            command=lambda v: ctk.set_appearance_mode(v))
        self.theme.set(ctk.get_appearance_mode().lower())
        self.theme.pack(side="left")
        ctk.CTkFrame(app_card, height=8, fg_color="transparent").pack()

    def _section(self, parent, title):
        ctk.CTkLabel(parent, text=title,
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(12, 4))

    def _save_model(self):
        self.config.set("model", "scraping", self.model_scraping.get().strip())
        self.config.set("model", "pipeline", self.model_pipeline.get().strip())
        keys_raw = self.api_keys_box.get("1.0", "end").strip()
        keys = [k.strip() for k in keys_raw.splitlines() if k.strip()]
        self.config.set("model", "api_keys", keys)
        for attr in ("num_ctx", "num_predict", "temperature"):
            try:
                val = getattr(self, f"model_{attr}").get()
                self.config.set("model", attr, float(val) if "." in val else int(val))
            except ValueError:
                pass
        self.config.save()
        messagebox.showinfo("Saved", "Model settings saved.")

    def _save_google(self):
        self.config.set("google", "enabled", self.google_enabled.get())
        for key, entry in self.google_fields.items():
            self.config.set("google", key, entry.get().strip())
        self.config.save()
        messagebox.showinfo("Saved", "Google settings saved.")

    def _save_github(self):
        token = self.github_token_entry.get().strip()
        self.config.set("github", "token", token)
        self.config.save()
        messagebox.showinfo("Saved", "GitHub token saved successfully.")

    def _test_google_auth(self):
        try:
            sys.path.insert(0, os.getcwd())
            from google_integration import _get_service
            drive = _get_service("drive", "v3")
            about = drive.about().get(fields="user").execute()
            email = about["user"]["emailAddress"]
            messagebox.showinfo("✅ Connected", f"Google auth successful!\nConnected as: {email}")
        except Exception as e:
            messagebox.showerror("Auth Failed", f"Google auth failed:\n{e}\n\nMake sure credentials.json is in project folder.")

# ══════════════════════════════════════════════════════════════
#  TAB: Screener Config
# ══════════════════════════════════════════════════════════════
class ScreenerConfigTab(ctk.CTkFrame):
    """Visual configurator for the AI job screener — no manual prompt editing needed."""

    _SEC_FG    = "#161b22"
    _ENTRY_FG  = "#1c2128"
    _BORDER    = "#30363d"

    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._company_vars: list[str] = []
        self._build()
        self._load_values()

    # ── Build ────────────────────────────────────────────────────

    def _build(self):
        ctk.CTkLabel(self, text="Screener Configuration",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self,
                     text="Set AI screening rules visually — the prompt is generated automatically.",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(anchor="w", padx=30, pady=(0, 12))

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=30, pady=(0, 8))

        def _section(title):
            f = ctk.CTkFrame(scroll, fg_color=self._SEC_FG, corner_radius=10)
            f.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(f, text=title, font=ctk.CTkFont(size=13, weight="bold"),
                         text_color="#58a6ff").pack(anchor="w", padx=16, pady=(10, 6))
            return f

        def _row(parent):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", padx=16, pady=3)
            return r

        def _lbl(parent, text, width=240):
            ctk.CTkLabel(parent, text=text, width=width, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left")

        # ─── Thresholds ───────────────────────────────────────────
        thr = _section("Screening Thresholds")

        r = _row(thr)
        _lbl(r, "Max years experience (auto-reject above):")
        self._max_years_var = ctk.IntVar(value=5)
        self._max_years_lbl = ctk.CTkLabel(r, text="5 yrs", width=55, anchor="w",
                                            font=ctk.CTkFont(size=12, weight="bold"),
                                            text_color="#79c0ff")
        ctk.CTkSlider(r, from_=0, to=15, number_of_steps=15,
                      variable=self._max_years_var, width=200,
                      command=lambda v: self._max_years_lbl.configure(
                          text=f"{int(v)} yr{'s' if int(v) != 1 else ''}")
                      ).pack(side="left", padx=10)
        self._max_years_lbl.pack(side="left")

        r = _row(thr)
        _lbl(r, "YES if skills match ≥:")
        self._yes_pct_var = ctk.IntVar(value=70)
        self._yes_pct_lbl = ctk.CTkLabel(r, text="70%", width=55, anchor="w",
                                          font=ctk.CTkFont(size=12, weight="bold"),
                                          text_color="#3fb950")
        ctk.CTkSlider(r, from_=30, to=100, number_of_steps=70,
                      variable=self._yes_pct_var, width=200,
                      command=lambda v: self._yes_pct_lbl.configure(text=f"{int(v)}%")
                      ).pack(side="left", padx=10)
        self._yes_pct_lbl.pack(side="left")

        r = _row(thr)
        _lbl(r, "MAYBE if skills match ≥:")
        self._maybe_pct_var = ctk.IntVar(value=45)
        self._maybe_pct_lbl = ctk.CTkLabel(r, text="45%", width=55, anchor="w",
                                            font=ctk.CTkFont(size=12, weight="bold"),
                                            text_color="#d29922")
        ctk.CTkSlider(r, from_=10, to=80, number_of_steps=70,
                      variable=self._maybe_pct_var, width=200,
                      command=lambda v: self._maybe_pct_lbl.configure(text=f"{int(v)}%")
                      ).pack(side="left", padx=10)
        self._maybe_pct_lbl.pack(side="left")
        ctk.CTkLabel(thr, text="  Jobs below the MAYBE threshold become NO.",
                     font=ctk.CTkFont(size=10), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 10))

        # ─── Role Levels ──────────────────────────────────────────
        lvl = _section("Accepted Role Levels")
        lvl_r = _row(lvl)
        self._lvl_vars: dict[str, ctk.BooleanVar] = {}
        for name, default in [("Junior", True), ("Mid", True), ("Senior", False)]:
            v = ctk.BooleanVar(value=default)
            self._lvl_vars[name.lower()] = v
            ctk.CTkCheckBox(lvl_r, text=name, variable=v,
                            font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 24))
        ctk.CTkLabel(lvl, text="  Unchecked levels are auto-rejected before AI screening.",
                     font=ctk.CTkFont(size=10), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 10))

        # ─── Skills ───────────────────────────────────────────────
        sk = _section("Skills")
        ctk.CTkLabel(sk, text="Required Skills  (one per line — AI will prioritise these):",
                     font=ctk.CTkFont(size=11), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 4))
        self._req_box = ctk.CTkTextbox(sk, height=80, fg_color=self._ENTRY_FG,
                                        border_color=self._BORDER, font=ctk.CTkFont(size=11))
        self._req_box.pack(fill="x", padx=16, pady=(0, 10))

        ctk.CTkLabel(sk, text="Preferred / Nice-to-have Skills  (one per line):",
                     font=ctk.CTkFont(size=11), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 4))
        self._pref_box = ctk.CTkTextbox(sk, height=70, fg_color=self._ENTRY_FG,
                                         border_color=self._BORDER, font=ctk.CTkFont(size=11))
        self._pref_box.pack(fill="x", padx=16, pady=(0, 10))

        # ─── Keyword Filters ──────────────────────────────────────
        kw = _section("Keyword Filters")
        ctk.CTkLabel(kw, text="Auto-REJECT if job title or description contains  (one per line):",
                     font=ctk.CTkFont(size=11), text_color="#f85149"
                     ).pack(anchor="w", padx=16, pady=(0, 4))
        self._reject_box = ctk.CTkTextbox(kw, height=70, fg_color=self._ENTRY_FG,
                                           border_color=self._BORDER, font=ctk.CTkFont(size=11))
        self._reject_box.pack(fill="x", padx=16, pady=(0, 10))

        ctk.CTkLabel(kw, text="Boost to YES if job title contains  (one per line):",
                     font=ctk.CTkFont(size=11), text_color="#3fb950"
                     ).pack(anchor="w", padx=16, pady=(0, 4))
        self._accept_box = ctk.CTkTextbox(kw, height=60, fg_color=self._ENTRY_FG,
                                           border_color=self._BORDER, font=ctk.CTkFont(size=11))
        self._accept_box.pack(fill="x", padx=16, pady=(0, 10))

        # ─── Blacklisted Companies ────────────────────────────────
        bl = _section("Blacklisted Companies")
        ctk.CTkLabel(bl,
                     text="Jobs from these companies are silently skipped before AI screening.",
                     font=ctk.CTkFont(size=11), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 8))

        add_r = _row(bl)
        self._bl_entry = ctk.CTkEntry(add_r, placeholder_text="Company name...",
                                       fg_color=self._ENTRY_FG, border_color=self._BORDER,
                                       width=290, font=ctk.CTkFont(size=12))
        self._bl_entry.pack(side="left", padx=(0, 8))
        self._bl_entry.bind("<Return>", lambda e: self._add_company())
        ctk.CTkButton(add_r, text="Add", width=70, height=30,
                      fg_color="#1f4e79", hover_color="#2563a8",
                      command=self._add_company).pack(side="left")

        self._bl_frame = ctk.CTkScrollableFrame(bl, fg_color="#0d1117",
                                                  height=130, corner_radius=6)
        self._bl_frame.pack(fill="x", padx=16, pady=(8, 12))

        # ─── Deduplication ────────────────────────────────────────
        dd = _section("Duplicate Detection")
        r = _row(dd)
        self._skip_applied_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(r, variable=self._skip_applied_var, font=ctk.CTkFont(size=12),
                        text="Skip jobs already in applied.db (already applied)"
                        ).pack(side="left")
        r = _row(dd)
        self._fuzzy_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(r, variable=self._fuzzy_var, font=ctk.CTkFont(size=12),
                        text="Skip near-duplicate jobs across sites (90% title + company match)"
                        ).pack(side="left")
        ctk.CTkLabel(dd, text="  Deduplication runs before AI screening to save API calls.",
                     font=ctk.CTkFont(size=10), text_color="#8b949e"
                     ).pack(anchor="w", padx=16, pady=(0, 10))

        # ─── Save / Reset ─────────────────────────────────────────
        btn_r = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_r.pack(fill="x", pady=(4, 20))
        ctk.CTkButton(btn_r, text="💾 Save Configuration", width=180, height=36,
                      fg_color="#238636", hover_color="#2ea043",
                      command=self._save).pack(side="left", padx=(0, 12))
        ctk.CTkButton(btn_r, text="↩ Reset to Defaults", width=160, height=36,
                      fg_color="#333", hover_color="#444",
                      command=self._reset).pack(side="left")
        self._status_lbl = ctk.CTkLabel(btn_r, text="", font=ctk.CTkFont(size=12))
        self._status_lbl.pack(side="left", padx=16)

    # ── Load / Save ──────────────────────────────────────────────

    def _load_values(self):
        cfg = self.config.data.get("screener", {})

        self._max_years_var.set(int(cfg.get("max_years_exp", 5)))
        self._max_years_lbl.configure(text=f"{int(cfg.get('max_years_exp', 5))} yrs")
        self._yes_pct_var.set(int(cfg.get("yes_match_pct", 70)))
        self._yes_pct_lbl.configure(text=f"{int(cfg.get('yes_match_pct', 70))}%")
        self._maybe_pct_var.set(int(cfg.get("maybe_match_pct", 45)))
        self._maybe_pct_lbl.configure(text=f"{int(cfg.get('maybe_match_pct', 45))}%")

        accepted = cfg.get("accept_role_levels", ["junior", "mid"])
        for lvl, var in self._lvl_vars.items():
            var.set(lvl in accepted)

        self._req_box.delete("1.0", "end")
        self._req_box.insert("1.0", cfg.get("required_skills", ""))
        self._pref_box.delete("1.0", "end")
        self._pref_box.insert("1.0", cfg.get("preferred_skills", ""))
        self._reject_box.delete("1.0", "end")
        self._reject_box.insert("1.0", cfg.get("reject_keywords", ""))
        self._accept_box.delete("1.0", "end")
        self._accept_box.insert("1.0", cfg.get("accept_keywords", ""))

        self._company_vars = list(cfg.get("blacklisted_companies", []))
        self._refresh_bl()

        self._skip_applied_var.set(bool(cfg.get("skip_applied", True)))
        self._fuzzy_var.set(bool(cfg.get("fuzzy_dedup", True)))

    def _save(self):
        screener = {
            "max_years_exp":         self._max_years_var.get(),
            "yes_match_pct":         self._yes_pct_var.get(),
            "maybe_match_pct":       self._maybe_pct_var.get(),
            "accept_role_levels":    [l for l, v in self._lvl_vars.items() if v.get()],
            "required_skills":       self._req_box.get("1.0", "end").strip(),
            "preferred_skills":      self._pref_box.get("1.0", "end").strip(),
            "reject_keywords":       self._reject_box.get("1.0", "end").strip(),
            "accept_keywords":       self._accept_box.get("1.0", "end").strip(),
            "blacklisted_companies": list(self._company_vars),
            "skip_applied":          self._skip_applied_var.get(),
            "fuzzy_dedup":           self._fuzzy_var.get(),
        }
        self.config.data["screener"] = screener
        self.config.save()
        self._status_lbl.configure(text="✓ Saved", text_color="#3fb950")
        self.after(2500, lambda: self._status_lbl.configure(text=""))

    def _reset(self):
        self.config.data.pop("screener", None)
        self.config.save()
        self._load_values()
        self._status_lbl.configure(text="✓ Reset to defaults", text_color="#d29922")
        self.after(2500, lambda: self._status_lbl.configure(text=""))

    # ── Blacklist CRUD ───────────────────────────────────────────

    def _add_company(self):
        name = self._bl_entry.get().strip()
        if not name:
            return
        if name.lower() not in [c.lower() for c in self._company_vars]:
            self._company_vars.append(name)
            self._refresh_bl()
        self._bl_entry.delete(0, "end")

    def _remove_company(self, company: str):
        self._company_vars = [c for c in self._company_vars if c != company]
        self._refresh_bl()

    def _refresh_bl(self):
        for w in self._bl_frame.winfo_children():
            w.destroy()
        if not self._company_vars:
            ctk.CTkLabel(self._bl_frame, text="No companies blacklisted.",
                         font=ctk.CTkFont(size=11), text_color="#8b949e").pack(pady=10)
            return
        for co in self._company_vars:
            r = ctk.CTkFrame(self._bl_frame, fg_color="#1c2128", corner_radius=6)
            r.pack(fill="x", pady=2)
            ctk.CTkLabel(r, text=co, font=ctk.CTkFont(size=12),
                         anchor="w").pack(side="left", padx=10, pady=5)
            ctk.CTkButton(r, text="✕", width=28, height=24,
                          fg_color="#3a1010", hover_color="#6a1010",
                          command=lambda c=co: self._remove_company(c)
                          ).pack(side="right", padx=6, pady=5)


# ══════════════════════════════════════════════════════════════
#  TAB: Job Tracker
# ══════════════════════════════════════════════════════════════

class JobDetailWindow(ctk.CTkToplevel):
    def __init__(self, master, config, job_data, on_save):
        super().__init__(master)
        self.title(f"Job Details — {job_data['company']}")
        self.geometry("700x800")
        self.grab_set()
        self.config = config
        self.job_data = job_data
        self.on_save = on_save

        self._build()

    def _build(self):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        # ── HERO SECTION ───────────────────────────────────────────────
        hero = ctk.CTkFrame(scroll, fg_color="#161b22", corner_radius=0)
        hero.pack(fill="x", pady=(0, 20))

        hero_content = ctk.CTkFrame(hero, fg_color="transparent")
        hero_content.pack(fill="x", padx=30, pady=30)

        title_lbl = ctk.CTkLabel(hero_content, text=self.job_data.get("title", "N/A"),
                                 font=ctk.CTkFont(size=26, weight="bold"), text_color="#ffffff")
        title_lbl.pack(anchor="w")

        company_lbl = ctk.CTkLabel(hero_content, text=self.job_data.get("company", "Unknown Company"),
                                   font=ctk.CTkFont(size=18), text_color="#8b949e")
        company_lbl.pack(anchor="w", pady=(0, 4))

        loc_raw = self.job_data.get("location", "")
        loc_text = "" if (not loc_raw or str(loc_raw).lower() in ("nan", "none")) else str(loc_raw)
        if loc_text:
            ctk.CTkLabel(hero_content, text=f"📍 {loc_text}",
                         font=ctk.CTkFont(size=13), text_color="#58a6ff").pack(anchor="w", pady=(0, 11))

        action_row = ctk.CTkFrame(hero_content, fg_color="transparent")
        action_row.pack(fill="x", pady=10)

        link = self.job_data.get("link", "")
        if link:
            ctk.CTkButton(action_row, text="View Original Posting ↗", height=36,
                          fg_color="#238636", hover_color="#2ea043", font=ctk.CTkFont(weight="bold"),
                          command=lambda: os.system(f'start {link}')).pack(side="left", padx=(0, 15))

        ctk.CTkButton(action_row, text="Generate Resume & Cover Letter", height=36,
                      fg_color="#1f6feb", hover_color="#388bfd", font=ctk.CTkFont(weight="bold"),
                      command=self._generate).pack(side="left", padx=(0, 15))

        # Status Badge
        verdict = self.job_data.get("AI_recommendation", "maybe").lower()
        v_color = "#238636" if verdict == "yes" else "#d29922" if verdict == "maybe" else "#da3633"
        status_badge = ctk.CTkLabel(action_row, text=f"AI Verdict: {verdict.upper()}",
                                    fg_color=v_color, text_color="white",
                                    corner_radius=8, font=ctk.CTkFont(size=12, weight="bold"),
                                    padx=12, pady=6)
        status_badge.pack(side="left")

        # ── QUICK STATS BAR ────────────────────────────────────────────
        stats_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        stats_frame.pack(fill="x", padx=30, pady=10)

        stat_cols = [
            ("⏳ Experience", self.job_data.get("years_required", "N/A")),
            ("🎯 Match Score", f"{self.job_data.get('skills_match_pct', 'N/A')}%"),
            ("👔 Role Level", self.job_data.get("role_level", "N/A")),
        ]

        for i, (label, value) in enumerate(stat_cols):
            card = ctk.CTkFrame(stats_frame, fg_color="#161b22", corner_radius=12, border_width=1, border_color="#30363d")
            card.pack(side="left", expand=True, fill="both", padx=5)

            ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=12), text_color="#8b949e").pack(pady=(12, 0))
            ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=16, weight="bold"), text_color="#ffffff").pack(pady=(0, 12))

        # ── MAIN CONTENT GRID ─────────────────────────────────────────
        main_content = ctk.CTkFrame(scroll, fg_color="transparent")
        main_content.pack(fill="both", expand=True, padx=30, pady=20)

        # Left Column: Job Description
        left_col = ctk.CTkFrame(main_content, fg_color="transparent")
        left_col.pack(side="left", fill="both", expand=True, padx=(0, 20))

        ctk.CTkLabel(left_col, text="Job Description", font=ctk.CTkFont(size=18, weight="bold"), text_color="#ffffff").pack(anchor="w", pady=(0, 10))
        jd_box = ctk.CTkTextbox(left_col, height=500, font=ctk.CTkFont(size=13), wrap="word",
                                fg_color="#161b22", border_color="#30363d", border_width=1, corner_radius=12)
        jd_box.pack(fill="both", expand=True)
        jd_box.insert("1.0", self.job_data.get("description", "No description available."))
        jd_box.configure(state="disabled")

        # Right Column: Analysis & Controls
        right_col = ctk.CTkFrame(main_content, fg_color="transparent")
        right_col.pack(side="right", fill="y")

        # Control Card
        ctrl_card = ctk.CTkFrame(right_col, fg_color="#161b22", corner_radius=12, border_width=1, border_color="#30363d")
        ctrl_card.pack(fill="x", pady=(0, 20))

        ctk.CTkLabel(ctrl_card, text="Management", font=ctk.CTkFont(size=14, weight="bold")).pack(padx=20, pady=(15, 10), anchor="w")

        # Verdict Edit
        v_row = ctk.CTkFrame(ctrl_card, fg_color="transparent")
        v_row.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(v_row, text="Verdict", font=ctk.CTkFont(size=12)).pack(side="left")
        self.verdict_var = ctk.StringVar(value=self.job_data.get("AI_recommendation", "maybe"))
        self.verdict_menu = ctk.CTkSegmentedButton(v_row, values=["yes", "maybe", "no"],
                                                   variable=self.verdict_var, height=28)
        self.verdict_menu.pack(side="right")

        # Applied Edit
        a_row = ctk.CTkFrame(ctrl_card, fg_color="transparent")
        a_row.pack(fill="x", padx=20, pady=5)
        ctk.CTkLabel(a_row, text="Status", font=ctk.CTkFont(size=12)).pack(side="left")
        self.applied_var = ctk.BooleanVar(value=self.job_data.get("application_status") == "Applied")
        ctk.CTkSwitch(a_row, text="Applied", variable=self.applied_var).pack(side="right")

        ctk.CTkButton(ctrl_card, text="Save Changes", height=32, font=ctk.CTkFont(weight="bold"),
                      command=self._save).pack(fill="x", padx=20, pady=(15, 20))

        # Matched Skills
        self._add_tag_section(right_col, "Matched Skills", self.job_data.get("matched_skills", ""), "green")

        # Missing Skills
        self._add_tag_section(right_col, "Missing Skills", self.job_data.get("missing_skills", ""), "red")

        # AI Reasoning
        ctk.CTkLabel(right_col, text="AI Insights", font=ctk.CTkFont(size=16, weight="bold"), text_color="#ffffff").pack(anchor="w", pady=(20, 10))
        reasoning_box = ctk.CTkTextbox(right_col, height=150, font=ctk.CTkFont(size=13), wrap="word",
                                      fg_color="#161b22", border_color="#30363d", border_width=1, corner_radius=12)
        reasoning_box.pack(fill="x")
        reasoning_box.insert("1.0", self.job_data.get("reasoning", "No reasoning provided."))
        reasoning_box.configure(state="disabled")

        # Additional Info
        add_info_card = ctk.CTkFrame(right_col, fg_color="#161b22", corner_radius=12,
                                     border_width=1, border_color="#30363d")
        add_info_card.pack(fill="x", pady=(20, 0))

        ctk.CTkLabel(add_info_card, text="Additional Info",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(padx=20, pady=(15, 10), anchor="w")

        ctk.CTkButton(add_info_card, text="📝  LaTeX Editor & PDF Preview", height=36,
                      font=ctk.CTkFont(weight="bold"),
                      fg_color="#6e40c9", hover_color="#8250df",
                      command=self._open_latex_editor).pack(fill="x", padx=20, pady=(0, 15))

    def _add_tag_section(self, parent, title, skills_text, color):
        ctk.CTkLabel(parent, text=title, font=ctk.CTkFont(size=14, weight="bold"), text_color="#8b949e").pack(anchor="w", pady=(20, 5))

        tag_frame = ctk.CTkFrame(parent, fg_color="transparent")
        tag_frame.pack(fill="x", pady=(0, 10))

        skills_str = "" if skills_text is None or (isinstance(skills_text, float)) else str(skills_text)
        skills = [s.strip() for s in skills_str.split(",") if s.strip()]
        if not skills:
            ctk.CTkLabel(tag_frame, text="None", font=ctk.CTkFont(size=12), text_color="#484f58").pack(anchor="w")
            return

        for skill in skills:
            tag = ctk.CTkLabel(tag_frame, text=skill, fg_color=color, text_color="white",
                                corner_radius=6, font=ctk.CTkFont(size=11),
                                padx=8, pady=2)
            tag.pack(anchor="w", pady=2)

    def _generate(self):
        SingleJobPipelineDialog(self, self.config, self.job_data)

    def _open_latex_editor(self):
        LatexEditorDialog(self, self.config, self.job_data)

    def _save(self):
        new_verdict = self.verdict_menu.get()
        status = "Applied" if self.applied_var.get() else "Not Applied"
        self.on_save(self.job_data.get("link"), new_verdict, status)
        self.destroy()


class SingleJobPipelineDialog(ctk.CTkToplevel):
    """Runs the resume pipeline for a single job and shows live progress."""

    def __init__(self, master, config, job_data):
        super().__init__(master)
        self.title(f"Generate — {job_data.get('title', 'Job')} @ {job_data.get('company', '')}")
        self.geometry("700x480")
        self.grab_set()
        self.config   = config
        self.job_data = job_data
        self.running  = False

        self.log_queue = queue.Queue()
        self._log_handler = QueueLogHandler(self.log_queue)
        self._log_handler.setFormatter(logging.Formatter("%(name)s — %(message)s"))
        logging.getLogger().addHandler(self._log_handler)

        self._build()
        self._start()

    def _build(self):
        ctk.CTkLabel(self, text="Generate Resume & Cover Letter",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(padx=20, pady=(15, 2), anchor="w")
        ctk.CTkLabel(self,
                     text=f"{self.job_data.get('title', '')} at {self.job_data.get('company', '')}",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(padx=20, pady=(0, 10), anchor="w")

        self.log_widget = LogWidget(self)
        self.log_widget.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 15))

        self.status_lbl = StatusBadge(btn_frame, font=ctk.CTkFont(size=12))
        self.status_lbl.pack(side="left")

        self.close_btn = ctk.CTkButton(btn_frame, text="Close", width=100,
                                       fg_color="#333", hover_color="#444",
                                       command=self._on_close, state="disabled")
        self.close_btn.pack(side="right")

    def _on_close(self):
        logging.getLogger().removeHandler(self._log_handler)
        self.destroy()

    def _inject_prompts(self):
        try:
            import agents.resume_builder as rb
            import re as _re
            prompts = self.config.get("prompts", default={})
            roles   = self.config.get("experience_roles", default=None)

            exp_prompt = prompts.get("experience_section", rb.PROMPT_EXPERIENCE)
            if roles:
                roles_text = "\n".join([
                    f"Role {i+1}: {r['title']} @ {r['company']} | {r['dates']}\n"
                    f"→ {r['total_bullets']} bullets: {r['fabricated_bullets']} fabricated + "
                    f"{r['real_bullets']} real ({r['domain']} domain)"
                    for i, r in enumerate(roles)
                ])
                exp_prompt = _re.sub(
                    r'Role 1:.*?(?=\n━━━|\nBULLET RULES|\nRAW LaTeX)',
                    roles_text, exp_prompt, flags=_re.DOTALL
                )
            rb.PROMPT_EXPERIENCE = exp_prompt

            profile = self.config.get("profile", default={})
            include_proj_links = profile.get("include_project_links", True)
            proj_prompt = prompts.get("projects_section", rb.PROMPT_PROJECTS)
            if include_proj_links:
                proj_prompt = proj_prompt.replace(
                    "\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}",
                    "\\href{url}{{\\textbf{Project Name}}} $|$ \\emph{{Tech1, Tech2, Tech3}}"
                )
                proj_prompt += "\n\nCRITICAL: For project headings, you MUST use the format: \\resumeProjectHeading{\\href{ACTUAL_URL}{{\\textbf{Project Name}}} $|$ \\emph{Techs}}{{}}. Replace 'ACTUAL_URL' with the real link found in the project data. The second set of braces must be EMPTY."
            else:
                proj_prompt = proj_prompt.replace(
                    "\\href{url}{{\\textbf{Project Name}}} $|$ \\emph{{Tech1, Tech2, Tech3}}",
                    "\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}"
                )
                proj_prompt += "\n\nCRITICAL: For project headings, you MUST use the format: \\resumeProjectHeading{\\textbf{Project Name} $|$ \\emph{Techs}}{{}}. The second set of braces must be EMPTY."
            rb.PROMPT_PROJECTS = proj_prompt

            rb.PROMPT_SKILLS       = prompts.get("skills_section",  rb.PROMPT_SKILLS)
            rb.PROMPT_COVER_LETTER = prompts.get("cover_letter",     rb.PROMPT_COVER_LETTER)

        except Exception as e:
            self.log_queue.put(("log", "WARNING", f"Could not inject prompts: {e}"))

    def _pipeline_thread(self):
        try:
            sys.path.insert(0, os.getcwd())
            self._inject_prompts()

            from pipeline import load_env, collect_ollama_keys, load_text_file, process_job
            from agents.api_client import RotatingOllamaClient
            from agents.resume_builder import ResumeBuilderAgent
            from agents.hiring_agent_ats import HiringAgentATSChecker
            from utils.latex_compiler import is_pdflatex_available

            load_env(".env")

            app_cfg  = self.config.data
            pipe_cfg = app_cfg.get("pipeline", {})
            model    = app_cfg.get("model", {}).get("pipeline", "gemma4:31b-cloud")

            output_dir     = pipe_cfg.get("output_dir",             "outputs")
            resume_path    = pipe_cfg.get("resume_path",            "resume.txt")
            projects_path  = pipe_cfg.get("projects_path",          "Projects.txt")
            max_iterations = int(pipe_cfg.get("max_ats_iterations", 2))
            pass_threshold = int(pipe_cfg.get("ats_pass_threshold", 85))

            os.environ["RESUME_FILENAME"]       = pipe_cfg.get("resume_filename",       "Resume")
            os.environ["COVER_LETTER_FILENAME"] = pipe_cfg.get("cover_letter_filename", "Cover_Letter")

            config_keys = app_cfg.get("model", {}).get("api_keys", [])
            for i, key in enumerate(config_keys, 1):
                os.environ[f"OLLAMA_API_KEY_{i}"] = key

            api_keys = collect_ollama_keys()
            client   = RotatingOllamaClient(api_keys=api_keys, model=model)

            existing_resume = load_text_file(resume_path,   "resume")
            projects_text   = load_text_file(projects_path, "projects")

            builder      = ResumeBuilderAgent(client, projects_text, existing_resume)
            checker      = HiringAgentATSChecker(client)
            use_pdflatex = is_pdflatex_available()

            self.log_queue.put(("log", "INFO",
                f"Starting pipeline for: {self.job_data.get('title')} @ {self.job_data.get('company')}"))

            result = process_job(
                self.job_data, builder, checker, output_dir, use_pdflatex,
                use_jd_location=pipe_cfg.get("use_jd_location",   True),
                default_location=pipe_cfg.get("default_location", "Canada"),
                max_iterations=max_iterations,
                pass_threshold=pass_threshold,
            )

            self.log_queue.put(("log", "INFO",
                f"✅ Done! ATS Score: {result['score']} | Status: {result['status']}"))
            if result.get("resume_path"):
                self.log_queue.put(("log", "INFO", f"Resume → {result['resume_path']}"))
            if result.get("cover_path"):
                self.log_queue.put(("log", "INFO", f"Cover Letter → {result['cover_path']}"))
            self.log_queue.put(("status", "done", f"Done — Score: {result['score']}"))

        except Exception as e:
            import traceback
            self.log_queue.put(("log", "ERROR", f"Pipeline failed: {e}"))
            self.log_queue.put(("log", "ERROR", traceback.format_exc()[:500]))
            self.log_queue.put(("status", "error", "Failed"))

        self.log_queue.put(("done", None, None))

    def _start(self):
        self.running = True
        self.status_lbl.set_status("running", "Generating...")
        threading.Thread(target=self._pipeline_thread, daemon=True).start()
        self.after(200, self._poll_logs)

    def _poll_logs(self):
        while not self.log_queue.empty():
            try:
                kind, level, msg = self.log_queue.get_nowait()
                if kind == "log":
                    self.log_widget.append(level, msg)
                elif kind == "status":
                    self.status_lbl.set_status(level, msg)
                elif kind == "done":
                    self.running = False
                    self.close_btn.configure(state="normal")
                    return
            except queue.Empty:
                break
        if self.running:
            self.after(200, self._poll_logs)


class LatexEditorDialog(ctk.CTkToplevel):
    """Side-by-side LaTeX source editor and live PDF preview for a generated resume."""

    def __init__(self, master, config, job_data):
        super().__init__(master)
        title   = job_data.get("title",   "Job")
        company = job_data.get("company", "")
        self.title(f"LaTeX Editor — {title} @ {company}")
        self.geometry("1300x780")
        self.minsize(900, 600)

        self.config          = config
        self.job_data        = job_data
        self._pdf_images          = []   # PhotoImage refs — must be kept alive
        self._compile_thread      = None
        self._bullet_improvements = {}   # {line_num: (original, improved)}
        self._overlay_btns        = []   # [(tk.Button, line_num), ...]
        self._ai_running          = False

        self._build()
        self._load_tex()

    # ── Path helpers ───────────────────────────────────────────
    def _slug(self, name: str) -> str:
        import re
        name = re.sub(r"[^\w\s-]", "", str(name))
        name = re.sub(r"\s+", "_", name.strip())
        return name[:60] or "Unknown"

    def _tex_path(self) -> str:
        out_dir     = self.config.get("pipeline", "output_dir",      default="outputs")
        resume_file = self.config.get("pipeline", "resume_filename", default="Resume")
        return os.path.join(
            out_dir,
            self._slug(self.job_data.get("company", "Unknown")),
            self._slug(self.job_data.get("title",   "Unknown")),
            f"{resume_file}.tex",
        )

    def _pdf_path(self) -> str:
        return self._tex_path().replace(".tex", ".pdf")

    # ── UI ─────────────────────────────────────────────────────
    def _build(self):
        # Header bar
        header = ctk.CTkFrame(self, fg_color="#161b22", corner_radius=0, height=46)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header,
            text=f"{self.job_data.get('title','')} @ {self.job_data.get('company','')}",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#58a6ff",
        ).pack(side="left", padx=16)
        self._path_lbl = ctk.CTkLabel(header, text="", font=ctk.CTkFont(size=10),
                                      text_color="#484f58")
        self._path_lbl.pack(side="left", padx=8)

        # PanedWindow for resizable split
        paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashwidth=5,
                               bg="#21262d", sashrelief="flat", bd=0)
        paned.pack(fill="both", expand=True)

        # ── Left: editor ──────────────────────────────────────
        left = tk.Frame(paned, bg="#0d1117")

        ed_bar = tk.Frame(left, bg="#161b22", height=34)
        ed_bar.pack(fill="x")
        ed_bar.pack_propagate(False)
        tk.Label(ed_bar, text="  LaTeX Source", bg="#161b22", fg="#8b949e",
                 font=("Segoe UI", 10, "bold")).pack(side="left", pady=6)

        # Native tk Text widget gives us line numbers + better performance
        editor_outer = tk.Frame(left, bg="#0d1117")
        editor_outer.pack(fill="both", expand=True)

        self._line_canvas = tk.Canvas(editor_outer, width=40, bg="#161b22",
                                      highlightthickness=0)
        self._line_canvas.pack(side="left", fill="y")

        self._editor = tk.Text(
            editor_outer,
            font=("Consolas", 11),
            bg="#0d1117", fg="#e6edf3",
            insertbackground="#e6edf3",
            selectbackground="#264f78",
            wrap="none",
            undo=True,
            relief="flat",
            padx=8, pady=4,
        )
        ed_scroll_y = tk.Scrollbar(editor_outer, orient="vertical",
                                   command=self._editor.yview)
        ed_scroll_x = tk.Scrollbar(left, orient="horizontal",
                                   command=self._editor.xview)
        self._ed_scroll_y = ed_scroll_y
        self._editor.configure(yscrollcommand=self._on_editor_yscroll,
                               xscrollcommand=ed_scroll_x.set)
        ed_scroll_y.pack(side="right", fill="y")
        ed_scroll_x.pack(side="bottom", fill="x")
        self._editor.pack(side="left", fill="both", expand=True)

        # Tag for highlighting bullet lines that have AI suggestions
        self._editor.tag_configure(
            "bullet_hl",
            background="#142236",
            foreground="#79c0ff",
        )

        self._editor.bind("<KeyRelease>", self._update_line_numbers)
        self._editor.bind("<MouseWheel>", self._update_line_numbers)
        self._editor.bind("<Configure>", lambda e: self._reposition_overlays())

        # ── Right: preview ────────────────────────────────────
        right = tk.Frame(paned, bg="#1c2128")

        prev_bar = tk.Frame(right, bg="#161b22", height=34)
        prev_bar.pack(fill="x")
        prev_bar.pack_propagate(False)
        tk.Label(prev_bar, text="  PDF Preview", bg="#161b22", fg="#8b949e",
                 font=("Segoe UI", 10, "bold")).pack(side="left", pady=6)
        self._page_lbl = tk.Label(prev_bar, text="", bg="#161b22", fg="#484f58",
                                  font=("Segoe UI", 9))
        self._page_lbl.pack(side="right", padx=10)

        preview_outer = tk.Frame(right, bg="#1c2128")
        preview_outer.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(preview_outer, bg="#1c2128",
                                 highlightthickness=0, cursor="hand2")
        prev_scroll = tk.Scrollbar(preview_outer, orient="vertical",
                                   command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=prev_scroll.set)
        prev_scroll.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._canvas.bind("<MouseWheel>",
                          lambda e: self._canvas.yview_scroll(
                              -1 if e.delta > 0 else 1, "units"))
        self._canvas.bind("<Configure>", self._on_canvas_resize)

        # Add panes
        paned.add(left,  minsize=380, stretch="always")
        paned.add(right, minsize=300, stretch="always")
        paned.paneconfigure(left,  width=680)
        paned.paneconfigure(right, width=520)

        # ── Bottom toolbar ────────────────────────────────────
        toolbar = ctk.CTkFrame(self, fg_color="#161b22", height=52, corner_radius=0)
        toolbar.pack(fill="x", side="bottom")
        toolbar.pack_propagate(False)

        self._status_lbl = ctk.CTkLabel(
            toolbar, text="Load or generate a resume, then click Compile.",
            font=ctk.CTkFont(size=11), text_color="#8b949e",
        )
        self._status_lbl.pack(side="left", padx=16)

        ctk.CTkButton(toolbar, text="📂  Open Folder", width=120, height=34,
                      fg_color="#333", hover_color="#444",
                      command=self._open_folder).pack(side="right", padx=6, pady=9)

        ctk.CTkButton(toolbar, text="💾  Save .tex", width=110, height=34,
                      fg_color="#333", hover_color="#444",
                      command=self._save_tex).pack(side="right", padx=(0, 6), pady=9)

        self._ai_btn = ctk.CTkButton(
            toolbar, text="🤖  Build with AI", width=150, height=34,
            fg_color="#6e40c9", hover_color="#8250df",
            font=ctk.CTkFont(weight="bold"),
            command=self._build_with_ai,
        )
        self._ai_btn.pack(side="right", padx=(0, 6), pady=9)

        self._compile_btn = ctk.CTkButton(
            toolbar, text="▶  Compile & Preview", width=165, height=34,
            fg_color="#1f6feb", hover_color="#388bfd",
            font=ctk.CTkFont(weight="bold"),
            command=self._start_compile,
        )
        self._compile_btn.pack(side="right", padx=(0, 6), pady=9)

    # ── Load / save ────────────────────────────────────────────
    def _load_tex(self):
        path = self._tex_path()
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
            self._editor.delete("1.0", "end")
            self._editor.insert("1.0", content)
            self._path_lbl.configure(text=path)
            self._set_status(f"Loaded: {path}", "ok")
            self._update_line_numbers()
            # Auto-render existing PDF if present
            pdf = self._pdf_path()
            if os.path.exists(pdf):
                self._render_pdf(pdf)
        else:
            self._path_lbl.configure(text="No .tex file found")
            self._set_status(
                "No resume found for this job. Generate one first via 'Generate Resume & Cover Letter'.",
                "error",
            )

    def _save_tex(self):
        path = self._tex_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        content = self._editor.get("1.0", "end-1c")
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        self._set_status(f"Saved → {path}", "ok")

    # ── Compile ────────────────────────────────────────────────
    def _start_compile(self):
        if self._compile_thread and self._compile_thread.is_alive():
            return
        self._compile_btn.configure(state="disabled", text="⏳  Compiling...")
        self._set_status("Compiling LaTeX…", "info")
        self._compile_thread = threading.Thread(
            target=self._compile_worker, daemon=True
        )
        self._compile_thread.start()

    def _compile_worker(self):
        try:
            from utils.latex_compiler import compile_latex_to_pdf

            content = self._editor.get("1.0", "end-1c")
            tex     = self._tex_path()
            pdf     = self._pdf_path()
            os.makedirs(os.path.dirname(tex), exist_ok=True)

            with open(tex, "w", encoding="utf-8") as f:
                f.write(content)

            ok = compile_latex_to_pdf(content, pdf)

            if ok and os.path.exists(pdf):
                self._safe_after(0, lambda: self._render_pdf(pdf))
                self._safe_after(0, lambda: self._set_status(f"✅ Compiled → {pdf}", "ok"))
            else:
                self._safe_after(0, lambda: self._set_status(
                    "Compilation failed — check LaTeX syntax", "error"))
        except Exception as exc:
            msg = str(exc)
            self._safe_after(0, lambda: self._set_status(f"Error: {msg}", "error"))
        finally:
            self._safe_after(0, lambda: self._alive() and self._compile_btn.configure(
                state="normal", text="▶  Compile & Preview"))

    # ── PDF rendering ──────────────────────────────────────────
    def _render_pdf(self, pdf_path: str):
        if not self._alive():
            return
        try:
            import fitz
            from PIL import Image, ImageTk

            doc   = fitz.open(pdf_path)
            pages = len(doc)
            self._canvas.delete("all")
            self._pdf_images.clear()

            cw     = max(self._canvas.winfo_width(), 400)
            y_off  = 12
            margin = 12

            for i in range(pages):
                page  = doc[i]
                scale = (cw - margin * 2) / page.rect.width
                mat   = fitz.Matrix(scale, scale)
                pix   = page.get_pixmap(matrix=mat, alpha=False)

                img   = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                photo = ImageTk.PhotoImage(img)
                self._pdf_images.append(photo)

                self._canvas.create_image(cw // 2, y_off, anchor="n", image=photo)
                y_off += pix.height + margin

            doc.close()
            self._canvas.configure(scrollregion=(0, 0, cw, y_off + 10))
            self._canvas.yview_moveto(0)
            self._page_lbl.configure(text=f"{pages} page{'s' if pages != 1 else ''}")

        except Exception as exc:
            self._set_status(f"Preview error: {exc}", "error")

    def _on_canvas_resize(self, event):
        """Re-render PDF at new width when the preview panel is resized."""
        pdf = self._pdf_path()
        if self._pdf_images and os.path.exists(pdf):
            self._render_pdf(pdf)

    # ── Line numbers ───────────────────────────────────────────
    def _update_line_numbers(self, event=None):
        if not self._alive():
            return
        try:
            self._line_canvas.delete("all")
        except Exception:
            return
        i    = self._editor.index("@0,0")
        lc   = self._line_canvas
        font = ("Consolas", 11)
        while True:
            dline = self._editor.dlineinfo(i)
            if dline is None:
                break
            y        = dline[1]
            h        = dline[3]
            num      = str(i).split(".")[0]
            line_int = int(num)

            if line_int in self._bullet_improvements:
                tag = f"bi_{line_int}"
                lc.create_text(20, y + h // 2, anchor="center",
                               text="✨", fill="#8250df",
                               font=("Segoe UI", 10), tags=(tag,))
                lc.tag_bind(tag, "<Button-1>",
                            lambda e, ln=line_int: self._on_bullet_icon_click(ln))
                lc.tag_bind(tag, "<Enter>",  lambda e: lc.configure(cursor="hand2"))
                lc.tag_bind(tag, "<Leave>",  lambda e: lc.configure(cursor=""))
            else:
                lc.create_text(35, y + 4, anchor="ne", text=num,
                               fill="#484f58", font=font)

            i = self._editor.index(f"{i}+1line")
            if self._editor.compare(i, ">=", "end"):
                break

        self._reposition_overlays()

    # ── Overlay helpers ────────────────────────────────────────
    def _on_editor_yscroll(self, *args):
        """Proxy for yscrollcommand — also repositions overlay buttons."""
        self._ed_scroll_y.set(*args)
        self._reposition_overlays()

    def _draw_bullet_overlays(self):
        """Apply text highlights + floating buttons for every bullet with a suggestion."""
        if not self._alive():
            return
        # Destroy stale buttons
        for btn, _ in self._overlay_btns:
            try:
                btn.destroy()
            except Exception:
                pass
        self._overlay_btns.clear()

        # Remove old highlights
        try:
            self._editor.tag_remove("bullet_hl", "1.0", "end")
        except Exception:
            pass

        for line_num, (original, improved) in self._bullet_improvements.items():
            try:
                self._editor.tag_add("bullet_hl", f"{line_num}.0", f"{line_num}.end+1c")
            except Exception:
                pass

            btn = tk.Button(
                self._editor,
                text="✨ Improve",
                bg="#6e40c9", fg="white",
                activebackground="#8250df", activeforeground="white",
                relief="flat", bd=0, cursor="hand2",
                font=("Segoe UI", 8, "bold"),
                padx=6, pady=1,
                command=lambda ln=line_num: self._on_bullet_icon_click(ln),
            )
            self._overlay_btns.append((btn, line_num))

        self._reposition_overlays()
        self._update_line_numbers()

    def _reposition_overlays(self):
        """Move each overlay button to the current pixel Y of its bullet line."""
        if not self._alive():
            return
        for btn, line_num in self._overlay_btns:
            try:
                dline = self._editor.dlineinfo(f"{line_num}.0")
                if dline:
                    _, y, _, h, _ = dline
                    btn_h = 18
                    btn.place(relx=1.0, x=-90, y=y + max(0, (h - btn_h) // 2),
                              width=82, height=btn_h)
                else:
                    btn.place_forget()
            except Exception:
                pass

    # ── Helpers ────────────────────────────────────────────────
    def _alive(self) -> bool:
        """Return True if the dialog window still exists in tk."""
        try:
            return bool(self.winfo_exists())
        except Exception:
            return False

    def _safe_after(self, ms: int, func) -> None:
        """Schedule func on the main thread; silently drop if window is gone."""
        try:
            self.after(ms, func)
        except Exception:
            pass

    def _set_status(self, msg: str, kind: str = "info"):
        if not self._alive():
            return
        colors = {"ok": "#3fb950", "error": "#f85149", "info": "#FFC107"}
        try:
            self._status_lbl.configure(text=msg, text_color=colors.get(kind, "#8b949e"))
        except Exception:
            pass

    def _open_folder(self):
        folder = os.path.dirname(self._tex_path())
        if os.path.exists(folder):
            os.startfile(folder)
        else:
            self._set_status("Output folder not found — generate a resume first.", "error")

    # ── Build with AI ──────────────────────────────────────────
    def _build_with_ai(self):
        tex = self._tex_path()
        if not os.path.exists(tex):
            self._set_status("No resume found — generating first…", "info")
            self._ai_btn.configure(state="disabled", text="⏳  Generating…")
            threading.Thread(target=self._generate_then_analyze, daemon=True).start()
        else:
            if not self._editor.get("1.0", "end-1c").strip():
                self._load_tex()
            self._start_ai_analysis()

    def _inject_prompts(self):
        try:
            import agents.resume_builder as rb
            import re as _re
            prompts = self.config.get("prompts", default={})
            roles   = self.config.get("experience_roles", default=None)
            exp_prompt = prompts.get("experience_section", rb.PROMPT_EXPERIENCE)
            if roles:
                roles_text = "\n".join([
                    f"Role {i+1}: {r['title']} @ {r['company']} | {r['dates']}\n"
                    f"→ {r['total_bullets']} bullets: {r['fabricated_bullets']} fabricated + "
                    f"{r['real_bullets']} real ({r['domain']} domain)"
                    for i, r in enumerate(roles)
                ])
                exp_prompt = _re.sub(
                    r'Role 1:.*?(?=\n━━━|\nBULLET RULES|\nRAW LaTeX)',
                    roles_text, exp_prompt, flags=_re.DOTALL
                )
            rb.PROMPT_EXPERIENCE = exp_prompt
            profile = self.config.get("profile", default={})
            proj_prompt = prompts.get("projects_section", rb.PROMPT_PROJECTS)
            if profile.get("include_project_links", True):
                proj_prompt = proj_prompt.replace(
                    "\\textbf{{Project Name}} $|$ \\emph{{Tech1, Tech2, Tech3}}",
                    "\\href{url}{{\\textbf{Project Name}}} $|$ \\emph{{Tech1, Tech2, Tech3}}"
                )
                proj_prompt += "\n\nCRITICAL: Use \\resumeProjectHeading{\\href{ACTUAL_URL}{{\\textbf{Name}}} $|$ \\emph{Techs}}{{}}."
            rb.PROMPT_PROJECTS    = proj_prompt
            rb.PROMPT_SKILLS       = prompts.get("skills_section",  rb.PROMPT_SKILLS)
            rb.PROMPT_COVER_LETTER = prompts.get("cover_letter",     rb.PROMPT_COVER_LETTER)
        except Exception:
            pass

    def _generate_then_analyze(self):
        """Run the pipeline for this job (no .tex found), then start AI analysis."""
        try:
            self._inject_prompts()
            from pipeline import load_env, collect_ollama_keys, load_text_file, process_job
            from agents.api_client import RotatingOllamaClient
            from agents.resume_builder import ResumeBuilderAgent
            from agents.hiring_agent_ats import HiringAgentATSChecker
            from utils.latex_compiler import is_pdflatex_available

            load_env(".env")
            app_cfg  = self.config.data
            pipe_cfg = app_cfg.get("pipeline", {})
            model    = app_cfg.get("model", {}).get("pipeline", "gemma4:31b-cloud")
            os.environ["RESUME_FILENAME"]       = pipe_cfg.get("resume_filename",       "Resume")
            os.environ["COVER_LETTER_FILENAME"] = pipe_cfg.get("cover_letter_filename", "Cover_Letter")
            for i, key in enumerate(app_cfg.get("model", {}).get("api_keys", []), 1):
                os.environ[f"OLLAMA_API_KEY_{i}"] = key

            client   = RotatingOllamaClient(api_keys=collect_ollama_keys(), model=model)
            builder  = ResumeBuilderAgent(
                client,
                load_text_file(pipe_cfg.get("projects_path", "Projects.txt"), "projects"),
                load_text_file(pipe_cfg.get("resume_path",   "resume.txt"),   "resume"),
            )
            checker  = HiringAgentATSChecker(client)

            self._safe_after(0, lambda: self._set_status("Generating resume…", "info"))
            result = process_job(
                self.job_data, builder, checker,
                pipe_cfg.get("output_dir", "outputs"),
                is_pdflatex_available(),
                use_jd_location=pipe_cfg.get("use_jd_location",   True),
                default_location=pipe_cfg.get("default_location", "Canada"),
                max_iterations=int(pipe_cfg.get("max_ats_iterations", 2)),
                pass_threshold=int(pipe_cfg.get("ats_pass_threshold", 85)),
            )
            if result.get("status") != "failed":
                self._safe_after(0, self._load_tex)
                self._safe_after(300, self._start_ai_analysis)
            else:
                self._safe_after(0, lambda: self._set_status("Resume generation failed.", "error"))
        except Exception as exc:
            msg = str(exc)
            self._safe_after(0, lambda: self._set_status(f"Generation failed: {msg}", "error"))
        finally:
            self._safe_after(0, lambda: self._alive() and self._ai_btn.configure(
                state="normal", text="🤖  Build with AI"))

    def _start_ai_analysis(self):
        if self._ai_running:
            return
        self._clear_overlays()
        self._ai_running = True
        self._ai_btn.configure(state="disabled", text="🤖  Analyzing…")
        self._set_status("AI is reading bullet points…", "info")
        threading.Thread(target=self._analyze_bullets_thread, daemon=True).start()

    def _analyze_bullets_thread(self):
        try:
            import re
            from ai import OllamaAssistant, send_with_retries

            bullet_map = self._parse_bullet_positions()
            if not bullet_map:
                self._safe_after(0, lambda: self._set_status(
                    "No \\resumeItem bullets found in this resume.", "info"))
                return

            total   = len(bullet_map)
            app_cfg = self.config.data
            model   = app_cfg.get("model", {}).get("pipeline", "gemma4:31b-cloud")
            keys    = app_cfg.get("model", {}).get("api_keys", [""])
            if not keys:
                keys = [""]

            self._safe_after(0, lambda: self._set_status(
                f"Analyzing {total} bullet points — this may take a minute…", "info"))

            for idx, (line_num, original) in enumerate(bullet_map.items(), 1):
                if not self._alive():
                    break           # window closed — stop processing
                os.environ["OLLAMA_API_KEY"] = keys[(idx - 1) % len(keys)]
                try:
                    improved = send_with_retries(
                        OllamaAssistant(model=model),
                        self._build_improvement_prompt(original),
                    ).strip()
                    # Strip any \resumeItem{} wrapper the AI may have added
                    improved = re.sub(r'^\\resumeItem\{', '', improved)
                    if improved.endswith("}"):
                        improved = improved[:-1]
                    improved = improved.strip()
                    if improved and improved != original:
                        self._bullet_improvements[line_num] = (original, improved)
                        self._safe_after(0, self._draw_bullet_overlays)
                except Exception:
                    pass
                n = idx
                self._safe_after(0, lambda n=n, t=total: self._set_status(
                    f"Analyzed {n}/{t} bullets…", "info"))

            count = len(self._bullet_improvements)
            self._safe_after(0, lambda: self._set_status(
                f"✅ {count} suggestions ready — click ✨ in the gutter to review each bullet.",
                "ok"))
        except Exception as exc:
            msg = str(exc)
            self._safe_after(0, lambda: self._set_status(f"AI analysis failed: {msg}", "error"))
        finally:
            self._ai_running = False
            self._safe_after(0, lambda: self._alive() and self._ai_btn.configure(
                state="normal", text="🤖  Build with AI"))

    def _build_improvement_prompt(self, bullet_text: str) -> str:
        title       = self.job_data.get("title",       "Software Developer")
        company     = self.job_data.get("company",     "")
        description = str(self.job_data.get("description", ""))[:2000]
        return (
            "You are an expert resume writer. Improve the following single resume bullet point.\n\n"
            f"JOB TITLE: {title}\nCOMPANY: {company}\n"
            f"JOB DESCRIPTION (excerpt):\n{description}\n\n"
            f"CURRENT BULLET CONTENT:\n{bullet_text}\n\n"
            "RULES:\n"
            "- Stronger action verbs (Architected, Engineered, Delivered, Optimized)\n"
            "- Add or sharpen specific metrics (%, count, time saved, scale)\n"
            "- Weave in relevant JD keywords naturally\n"
            "- Preserve LaTeX formatting like \\textbf{} for 2-3 key terms\n"
            "- All % → \\%, all & → \\&\n"
            "- Output ONLY the improved bullet content (NOT the \\resumeItem{} wrapper)\n"
            "- Keep to 2 compiled lines maximum\n\n"
            "IMPROVED BULLET CONTENT:"
        )

    def _parse_bullet_positions(self) -> dict:
        """Return {line_num: bullet_content} for every \\resumeItem in the editor."""
        content = self._editor.get("1.0", "end-1c")
        results = {}
        i       = 0
        marker  = r"\resumeItem{"
        while True:
            idx = content.find(marker, i)
            if idx == -1:
                break
            line_num = content[:idx].count("\n") + 1
            start    = idx + len(marker)
            depth    = 1
            j        = start
            while j < len(content) and depth > 0:
                if   content[j] == "{": depth += 1
                elif content[j] == "}": depth -= 1
                j += 1
            results[line_num] = content[start:j - 1].strip()
            i = j
        return results

    def _on_bullet_icon_click(self, line_num: int):
        if line_num not in self._bullet_improvements:
            return
        original, improved = self._bullet_improvements[line_num]
        BulletSuggestionDialog(
            self, original, improved,
            on_apply=lambda new: self._apply_bullet(line_num, new),
        )

    def _apply_bullet(self, line_num: int, new_content: str):
        """Replace the content inside \\resumeItem{} at line_num with new_content."""
        content = self._editor.get("1.0", "end-1c")
        marker  = r"\resumeItem{"
        lines   = content.split("\n")
        # Character offset to the start of target line (0-indexed)
        char_offset = sum(len(l) + 1 for l in lines[:line_num - 1])
        idx = content.find(marker, char_offset)
        if idx == -1:
            return
        start = idx + len(marker)
        depth = 1
        j     = start
        while j < len(content) and depth > 0:
            if   content[j] == "{": depth += 1
            elif content[j] == "}": depth -= 1
            j += 1
        new_full = content[:start] + new_content + content[j - 1:]
        self._editor.delete("1.0", "end")
        self._editor.insert("1.0", new_full)
        self._bullet_improvements.pop(line_num, None)
        self._draw_bullet_overlays()
        self._save_tex()
        self._start_compile()

    def _clear_overlays(self):
        self._bullet_improvements.clear()
        for btn, _ in self._overlay_btns:
            try:
                btn.destroy()
            except Exception:
                pass
        self._overlay_btns.clear()
        try:
            self._editor.tag_remove("bullet_hl", "1.0", "end")
        except Exception:
            pass
        self._update_line_numbers()


class BulletSuggestionDialog(ctk.CTkToplevel):
    """Side-by-side comparison of current vs AI-improved bullet point."""

    def __init__(self, master, original: str, improved: str, on_apply):
        super().__init__(master)
        self.title("Improve Bullet Point")
        self.geometry("940x430")
        self.resizable(True, False)
        self.grab_set()
        self._improved = improved
        self._on_apply = on_apply
        self._build(original, improved)

    def _build(self, original: str, improved: str):
        ctk.CTkLabel(self, text="Choose which version to use in your resume",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(
                     padx=24, pady=(16, 12), anchor="w")

        grid = ctk.CTkFrame(self, fg_color="transparent")
        grid.pack(fill="both", expand=True, padx=24, pady=(0, 20))
        grid.grid_columnconfigure(0, weight=1)
        grid.grid_columnconfigure(1, weight=1)

        # ── Current ─────────────────────────────────────────────
        orig = ctk.CTkFrame(grid, fg_color="#161b22", corner_radius=10,
                             border_width=1, border_color="#30363d")
        orig.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        ctk.CTkLabel(orig, text="  Current Version",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#8b949e", anchor="w").pack(fill="x", padx=4, pady=(12, 4))

        ob = ctk.CTkTextbox(orig, height=240, wrap="word",
                            font=ctk.CTkFont(family="Consolas", size=11),
                            fg_color="#0d1117", border_width=0, corner_radius=0)
        ob.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        ob.insert("1.0", original)
        ob.configure(state="disabled")

        ctk.CTkButton(orig, text="Keep Original", height=36,
                      fg_color="#30363d", hover_color="#444",
                      command=self.destroy).pack(fill="x", padx=12, pady=(0, 14))

        # ── AI Improved ─────────────────────────────────────────
        impr = ctk.CTkFrame(grid, fg_color="#0d1f3c", corner_radius=10,
                             border_width=1, border_color="#1f6feb")
        impr.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        hdr = ctk.CTkFrame(impr, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(12, 4))
        ctk.CTkLabel(hdr, text="✨  AI Improved",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="#58a6ff").pack(side="left")

        ib = ctk.CTkTextbox(impr, height=240, wrap="word",
                            font=ctk.CTkFont(family="Consolas", size=11),
                            fg_color="#0d1117", border_width=0, corner_radius=0)
        ib.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        ib.insert("1.0", improved)
        ib.configure(state="disabled")

        ctk.CTkButton(impr, text="✅  Use AI Version", height=36,
                      fg_color="#1f6feb", hover_color="#388bfd",
                      font=ctk.CTkFont(weight="bold"),
                      command=self._use_improved).pack(fill="x", padx=12, pady=(0, 14))

    def _use_improved(self):
        self._on_apply(self._improved)
        self.destroy()


class AddJobDialog(ctk.CTkToplevel):
    """Dialog to manually add a job posting. Supports URL fetch, LinkedIn guest API, and manual paste."""
    """Dialog to manually add a job posting. Supports URL fetch, LinkedIn guest API, and manual paste."""

    # Keywords that indicate a login/auth wall was returned instead of the job
    _LOGIN_SIGNALS = ("sign in", "log in", "create an account", "join now", "sign up to", "please log in")

    def __init__(self, master, config, on_saved):
        super().__init__(master)
        self.config = config
        self.on_saved = on_saved
        self._worker = None

        self.title("Add Job Posting")
        self.geometry("620x560")
        self.resizable(False, False)
        self.grab_set()

        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Add Job Posting",
                     font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(18, 2))
        ctk.CTkLabel(self, text="Paste the URL and/or the job description. AI will screen it.",
                     font=ctk.CTkFont(size=12), text_color="gray").pack(pady=(0, 12))

        # ── URL row ────────────────────────────────────────────
        url_frame = ctk.CTkFrame(self, fg_color="transparent")
        url_frame.pack(fill="x", padx=24, pady=(0, 4))
        ctk.CTkLabel(url_frame, text="Job URL:", width=80, anchor="w",
                     font=ctk.CTkFont(size=12)).pack(side="left")
        self.url_entry = ctk.CTkEntry(url_frame, placeholder_text="https://linkedin.com/jobs/view/…", height=32)
        self.url_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.fetch_btn = ctk.CTkButton(url_frame, text="⬇ Fetch", width=72, height=32,
                                       fg_color="#333", hover_color="#444",
                                       command=self._start_fetch)
        self.fetch_btn.pack(side="left")

        # ── Title / Company row ────────────────────────────────
        meta_frame = ctk.CTkFrame(self, fg_color="transparent")
        meta_frame.pack(fill="x", padx=24, pady=(6, 0))

        title_col = ctk.CTkFrame(meta_frame, fg_color="transparent")
        title_col.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(title_col, text="Job Title:", anchor="w",
                     font=ctk.CTkFont(size=12)).pack(anchor="w")
        self.title_entry = ctk.CTkEntry(title_col, placeholder_text="e.g. Software Engineer", height=32)
        self.title_entry.pack(fill="x")

        company_col = ctk.CTkFrame(meta_frame, fg_color="transparent")
        company_col.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(company_col, text="Company:", anchor="w",
                     font=ctk.CTkFont(size=12)).pack(anchor="w")
        self.company_entry = ctk.CTkEntry(company_col, placeholder_text="e.g. Google", height=32)
        self.company_entry.pack(fill="x")

        # ── Description box ────────────────────────────────────
        ctk.CTkLabel(self, text="Job Description (paste here if URL requires login):",
                     anchor="w", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=24, pady=(10, 2))
        self.desc_box = ctk.CTkTextbox(self, height=160, font=ctk.CTkFont(size=11))
        self.desc_box.pack(fill="x", padx=24)

        # ── Status / progress ──────────────────────────────────
        self.status_label = ctk.CTkLabel(self, text="", font=ctk.CTkFont(size=11), text_color="gray")
        self.status_label.pack(pady=(6, 0))
        self.progress = ctk.CTkProgressBar(self, mode="indeterminate", width=460)

        # ── Analyze button ─────────────────────────────────────
        self.analyze_btn = ctk.CTkButton(self, text="🤖 Analyze & Save", height=38,
                                         font=ctk.CTkFont(size=13, weight="bold"),
                                         fg_color="#238636", hover_color="#2ea043",
                                         command=self._start_analyze)
        self.analyze_btn.pack(pady=(10, 16))

    # ── Fetch helpers ──────────────────────────────────────────

    def _start_fetch(self):
        url = self.url_entry.get().strip()
        if not url.startswith("http"):
            self._set_status("Enter a valid URL first.", "#da3633")
            return
        self.fetch_btn.configure(state="disabled", text="…")
        self._set_status("Fetching…", "gray")
        threading.Thread(target=self._run_fetch, args=(url,), daemon=True).start()

    def _run_fetch(self, url):
        import requests
        from bs4 import BeautifulSoup
        try:
            title, company, description = "", "", ""

            # LinkedIn: try the public guest API first
            if "linkedin.com" in url.lower():
                job_id = self._extract_linkedin_id(url)
                if job_id:
                    self._set_status("Trying LinkedIn guest API…", "gray")
                    guest_url = f"https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}"
                    r = requests.get(guest_url, timeout=15,
                                     headers={"User-Agent": "Mozilla/5.0"})
                    if r.status_code == 200:
                        soup = BeautifulSoup(r.text, "html.parser")
                        title   = (soup.find("h2") or soup.find("h1") or object())
                        title   = title.get_text(strip=True) if hasattr(title, "get_text") else ""
                        company = (soup.find("a", {"class": lambda c: c and "company" in c}) or object())
                        company = company.get_text(strip=True) if hasattr(company, "get_text") else ""
                        for tag in soup(["script", "style"]): tag.decompose()
                        description = soup.get_text(separator=" ", strip=True)[:6000]

            # Generic fetch (also fallback for LinkedIn if guest API gave nothing)
            if not description:
                self._set_status("Fetching page…", "gray")
                headers = {
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"
                    ),
                    "Accept-Language": "en-US,en;q=0.9",
                }
                r = requests.get(url, headers=headers, timeout=20)
                soup = BeautifulSoup(r.text, "html.parser")
                for tag in soup(["script", "style", "nav", "footer", "header"]):
                    tag.decompose()
                page_text = soup.get_text(separator=" ", strip=True)

                # Detect login wall
                page_lower = page_text.lower()
                if any(sig in page_lower[:500] for sig in self._LOGIN_SIGNALS):
                    self._set_status(
                        "⚠ This site requires login. Please paste the job description below manually.",
                        "#d29922"
                    )
                    self.after(0, lambda: self.fetch_btn.configure(state="normal", text="⬇ Fetch"))
                    return

                if not title:
                    h = soup.find("h1") or soup.find("h2")
                    title = h.get_text(strip=True) if h else (soup.title.get_text(strip=True) if soup.title else "")
                if not company:
                    og = soup.find("meta", property="og:site_name")
                    company = og.get("content", "") if og else ""
                description = page_text[:6000]

            # Populate the fields on the main thread
            self.after(0, lambda t=title, c=company, d=description: self._populate(t, c, d))
            self._set_status("Fetched! Review the fields then click Analyze & Save.", "#238636")

        except Exception as e:
            self._set_status(f"Fetch error: {e}", "#da3633")

        self.after(0, lambda: self.fetch_btn.configure(state="normal", text="⬇ Fetch"))

    def _populate(self, title, company, description):
        if title:
            self.title_entry.delete(0, "end")
            self.title_entry.insert(0, title)
        if company:
            self.company_entry.delete(0, "end")
            self.company_entry.insert(0, company)
        if description:
            self.desc_box.delete("1.0", "end")
            self.desc_box.insert("1.0", description)

    @staticmethod
    def _extract_linkedin_id(url):
        import re
        m = re.search(r"/jobs/view/(\d+)", url)
        if m:
            return m.group(1)
        m = re.search(r"currentJobId=(\d+)", url)
        if m:
            return m.group(1)
        return None

    # ── Analyze helpers ────────────────────────────────────────

    def _start_analyze(self):
        description = self.desc_box.get("1.0", "end").strip()
        title       = self.title_entry.get().strip()
        url         = self.url_entry.get().strip()

        if not description and not url.startswith("http"):
            self._set_status("Enter a URL or paste a job description.", "#da3633")
            return
        if not description:
            self._set_status("Please fetch the page first (⬇ Fetch) or paste the description.", "#da3633")
            return

        self.analyze_btn.configure(state="disabled", text="Analyzing…")
        self.fetch_btn.configure(state="disabled")
        self.progress.pack(pady=(4, 0))
        self.progress.start()
        self._set_status("Running AI screening…", "gray")

        self._worker = threading.Thread(target=self._run_analyze, daemon=True)
        self._worker.start()
        self.after(200, self._poll_worker)

    def _poll_worker(self):
        if self._worker and self._worker.is_alive():
            self.after(200, self._poll_worker)

    def _set_status(self, msg, color="gray"):
        self.after(0, lambda: self.status_label.configure(text=msg, text_color=color))

    def _run_analyze(self):
        try:
            url         = self.url_entry.get().strip()
            title       = self.title_entry.get().strip() or "Unknown Position"
            company     = self.company_entry.get().strip() or "Unknown"
            description = self.desc_box.get("1.0", "end").strip()
            site        = self._detect_site(url) if url.startswith("http") else "web"

            from jobs import DEFAULT_JOB_FINDER_PROMPT, format_prompt, load_df, write_excel_safely
            from setup_wizard import load_profile, build_profile_prompt_context

            profile     = load_profile()
            profile_ctx = build_profile_prompt_context(profile)

            try:
                cfg_path = os.path.join(BASE_DIR, "app_config.json")
                with open(cfg_path, "r", encoding="utf-8") as f:
                    cfg_data = json.load(f)
                prompt_template = cfg_data.get("prompts", {}).get("job_screener", DEFAULT_JOB_FINDER_PROMPT)
            except Exception:
                prompt_template = DEFAULT_JOB_FINDER_PROMPT

            prompt_template = prompt_template.replace("{candidate_profile}", profile_ctx)

            resume_path = self.config.get("pipeline", "resume_path", default="resume.txt")
            try:
                with open(resume_path, "r", encoding="utf-8") as f:
                    resume_text = f.read().strip()
            except Exception:
                resume_text = ""

            msg       = format_prompt(prompt_template, title=title, description=description, resume_text=resume_text)
            model     = self.config.get("ai", "model", default="gemma4:31b-cloud")
            assistant = OllamaAssistant(model=model)
            ai_resp   = send_with_retries(assistant, msg, tries=3, backoff_sec=2.0)
            result    = parse_ai_verdict(ai_resp)

            matched = result.get("matched_skills", [])
            missing = result.get("missing_skills", [])
            if isinstance(matched, list): matched = ", ".join(matched)
            if isinstance(missing, list): missing = ", ".join(missing)

            new_row = {
                "AI_recommendation": result["verdict"],
                "site":              site,
                "company":           company,
                "title":             title,
                "link":              url or "N/A",
                "years_required":    str(result.get("years_required", "unspecified")),
                "role_level":        str(result.get("role_level", "unspecified")).lower(),
                "skills_match_pct":  result.get("skills_match_pct", 0),
                "matched_skills":    matched,
                "missing_skills":    missing,
                "reasoning":         result.get("reasoning", ""),
                "description":       description[:2000],
                "posted_date":       "",
            }

            excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
            existing   = load_df()
            new_df     = pd.DataFrame([new_row])
            combined   = pd.concat([existing, new_df], ignore_index=True)
            write_excel_safely(combined, excel_path)

            self._set_status(
                f"✅ Saved!  Verdict: {result['verdict'].upper()}  |  Match: {result.get('skills_match_pct', 0)}%",
                "#238636"
            )
            self.after(0, self._finish_ok)

        except Exception as e:
            self._set_status(f"Error: {e}", "#da3633")
            self.after(0, self._finish_error)

    @staticmethod
    def _detect_site(url):
        url_l = url.lower()
        if "linkedin.com" in url_l:    return "linkedin"
        if "indeed.com" in url_l:      return "indeed"
        if "glassdoor.com" in url_l:   return "glassdoor"
        if "jobright" in url_l:        return "jobright"
        if "ziprecruiter" in url_l:    return "zip_recruiter"
        if "naukri.com" in url_l:      return "naukri"
        if "google.com/about/careers" in url_l or "careers.google" in url_l: return "google"
        if "bayt.com" in url_l:        return "bayt"
        if "bdjobs.com" in url_l:      return "bdjobs"
        return "web"

    def _finish_ok(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.analyze_btn.configure(state="normal", text="🤖 Analyze & Save")
        self.fetch_btn.configure(state="normal")
        self.on_saved()
        self.after(2000, self.destroy)

    def _finish_error(self):
        self.progress.stop()
        self.progress.pack_forget()
        self.analyze_btn.configure(state="normal", text="🤖 Analyze & Save")
        self.fetch_btn.configure(state="normal")


# ══════════════════════════════════════════════════════════════
#  Applied Job Detail Window
# ══════════════════════════════════════════════════════════════
class AppliedDetailWindow(ctk.CTkToplevel):
    def __init__(self, master, job_data: dict, config=None):
        super().__init__(master)
        company = str(job_data.get("company", "Applied Job"))
        self.title(f"Applied — {company}")
        self.geometry("780x720")
        self.resizable(True, True)
        self.grab_set()
        self._config = config
        self._build(job_data)

    # ── File finder ───────────────────────────────────────────
    def _find_job_files(self, data: dict):
        """Return (resume_path, cover_letter_path) — either or both may be None."""
        import re

        def _sanitize(name):
            name = re.sub(r"[^\w\s-]", "", name)
            return re.sub(r"\s+", "_", name.strip())[:60] or "Unknown"

        applied_folder = (self._config.get("pipeline", "applied_folder_path", default="fultime")
                          if self._config else "fultime")

        company_raw = str(data.get("company", ""))
        title_raw   = str(data.get("title",   ""))
        company_s   = _sanitize(company_raw)
        title_s     = _sanitize(title_raw) if title_raw.lower() not in ("nan", "none", "") else ""

        # Build a priority list of directories to scan
        candidate_dirs = []

        # 1. Stored folder_path (set by _migrate_to_applied for new-format jobs)
        fp = data.get("folder_path", "")
        if fp and str(fp).lower() not in ("nan", "none", "") and os.path.isdir(str(fp)):
            candidate_dirs.append(str(fp))

        # 2. Search sanitized and raw company folder for title sub-dirs
        for cname in [company_s, company_raw]:
            cdir = os.path.join(applied_folder, cname)
            if not os.path.isdir(cdir):
                continue
            # Title sub-directories (e.g. SoftwareEngineer_2026-05-27)
            if title_s:
                try:
                    for entry in os.scandir(cdir):
                        if entry.is_dir() and entry.name.startswith(title_s):
                            candidate_dirs.append(entry.path)
                except OSError:
                    pass
            # Also the company folder itself (old flat layout)
            candidate_dirs.append(cdir)

        resume_path      = None
        cover_letter_path = None

        for folder in candidate_dirs:
            if not os.path.isdir(folder):
                continue
            try:
                entries = list(os.scandir(folder))
            except OSError:
                continue
            for f in entries:
                if not f.is_file():
                    continue
                nl = f.name.lower()
                if not (nl.endswith(".pdf") or nl.endswith(".docx")):
                    continue
                if "resume" in nl and resume_path is None:
                    resume_path = f.path
                elif ("cover" in nl or "_cl" in nl or "cl_" in nl) and cover_letter_path is None:
                    cover_letter_path = f.path
            if resume_path or cover_letter_path:
                break

        return resume_path, cover_letter_path

    def _build(self, data: dict):
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        # Hero section
        hero = ctk.CTkFrame(scroll, fg_color="#161b22", corner_radius=0)
        hero.pack(fill="x", pady=(0, 16))
        hc = ctk.CTkFrame(hero, fg_color="transparent")
        hc.pack(fill="x", padx=28, pady=24)

        title_str = str(data.get("title", "N/A"))
        ctk.CTkLabel(hc, text=title_str,
                     font=ctk.CTkFont(size=22, weight="bold")).pack(anchor="w")

        company_str = str(data.get("company", "Unknown"))
        ctk.CTkLabel(hc, text=company_str,
                     font=ctk.CTkFont(size=15), text_color="#8b949e").pack(anchor="w", pady=(2, 4))

        loc_raw = data.get("location", "")
        loc_text = "" if (not loc_raw or str(loc_raw).lower() in ("nan", "none")) else str(loc_raw)
        if loc_text:
            ctk.CTkLabel(hc, text=f"📍 {loc_text}",
                         font=ctk.CTkFont(size=13), text_color="#58a6ff").pack(anchor="w", pady=(0, 4))

        # Applied date badge
        applied_raw = data.get("applied_date", "")
        applied_str = str(applied_raw) if (applied_raw and str(applied_raw).lower() not in ("nan", "none", "nat", "")) else ""
        if applied_str:
            ctk.CTkLabel(hc, text=f"✅ Applied: {applied_str[:10]}",
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color="#3fb950").pack(anchor="w", pady=(0, 6))

        action_row = ctk.CTkFrame(hc, fg_color="transparent")
        action_row.pack(fill="x", pady=(6, 0))

        link = data.get("link", "")
        if link and str(link).lower() not in ("nan", "none", ""):
            ctk.CTkButton(action_row, text="View Posting ↗", height=34,
                          fg_color="#238636", hover_color="#2ea043",
                          font=ctk.CTkFont(weight="bold"),
                          command=lambda: os.system(f'start {link}')).pack(side="left", padx=(0, 8))

        verdict = str(data.get("AI_recommendation", "")).lower()
        v_color = "#238636" if verdict == "yes" else "#d29922" if verdict == "maybe" else "#da3633"
        ctk.CTkLabel(action_row, text=f"AI: {verdict.upper() if verdict else 'N/A'}",
                     fg_color=v_color, text_color="white", corner_radius=8,
                     font=ctk.CTkFont(size=12, weight="bold"),
                     padx=10, pady=6).pack(side="left", padx=(0, 8))

        # Resume / Cover Letter buttons
        resume_path, cl_path = self._find_job_files(data)
        if resume_path:
            rp = resume_path
            ctk.CTkButton(action_row, text="📄 Resume", height=34, width=100,
                          fg_color="#1f4e79", hover_color="#2563a8",
                          command=lambda: os.startfile(rp)).pack(side="left", padx=(0, 8))
        if cl_path:
            cp = cl_path
            ctk.CTkButton(action_row, text="✉ Cover Letter", height=34, width=120,
                          fg_color="#4a1d5e", hover_color="#6b2b88",
                          command=lambda: os.startfile(cp)).pack(side="left")

        if not resume_path and not cl_path:
            ctk.CTkLabel(action_row, text="No documents found",
                         font=ctk.CTkFont(size=11), text_color="#8b949e").pack(side="left", padx=8)

        # Stats bar
        stats = ctk.CTkFrame(scroll, fg_color="transparent")
        stats.pack(fill="x", padx=28, pady=(0, 14))
        for label, val in [
            ("⏳ Experience",  str(data.get("years_required", "N/A"))),
            ("🎯 Match Score", f"{data.get('skills_match_pct', 'N/A')}%"),
            ("👔 Role Level",  str(data.get("role_level", "N/A"))),
        ]:
            card = ctk.CTkFrame(stats, fg_color="#161b22", corner_radius=12,
                                border_width=1, border_color="#30363d")
            card.pack(side="left", expand=True, fill="both", padx=5)
            ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=12),
                         text_color="#8b949e").pack(pady=(10, 0))
            ctk.CTkLabel(card, text=val, font=ctk.CTkFont(size=15, weight="bold")).pack(pady=(0, 10))

        def _section(title_text):
            ctk.CTkLabel(scroll, text=title_text, font=ctk.CTkFont(size=13, weight="bold"),
                         text_color="#8b949e").pack(anchor="w", padx=28, pady=(10, 4))

        def _row_field(label, value):
            if value is None or str(value).lower() in ("nan", "none", "nat", ""):
                return
            row = ctk.CTkFrame(scroll, fg_color="#1c2128", corner_radius=6)
            row.pack(fill="x", padx=28, pady=2)
            ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=11, weight="bold"),
                         text_color="#8b949e", width=140, anchor="w").pack(side="left", padx=(12, 0), pady=8)
            val_str = str(value)
            if len(val_str) > 100:
                tb = ctk.CTkTextbox(row, height=72, font=ctk.CTkFont(size=11),
                                    fg_color="transparent", border_width=0, wrap="word")
                tb.pack(side="left", fill="x", expand=True, padx=8, pady=4)
                tb.insert("1.0", val_str)
                tb.configure(state="disabled")
            else:
                ctk.CTkLabel(row, text=val_str, font=ctk.CTkFont(size=11),
                             anchor="w", wraplength=480).pack(side="left", fill="x", expand=True, padx=8, pady=8)

        _section("Application Details")
        _row_field("Site",         data.get("site"))
        _row_field("Applied Date", applied_str or None)
        _row_field("Posted Date",  data.get("posted_date"))

        _section("AI Analysis")
        _row_field("Matched Skills", data.get("matched_skills"))
        _row_field("Missing Skills", data.get("missing_skills"))
        _row_field("Reasoning",      data.get("reasoning"))

        desc_raw = data.get("description", "")
        if desc_raw and str(desc_raw).lower() not in ("nan", "none", ""):
            _section("Job Description")
            tb_frame = ctk.CTkFrame(scroll, fg_color="#1c2128", corner_radius=6)
            tb_frame.pack(fill="x", padx=28, pady=2)
            tb = ctk.CTkTextbox(tb_frame, height=200, font=ctk.CTkFont(size=11),
                                fg_color="transparent", border_width=0, wrap="word")
            tb.pack(fill="x", expand=True, padx=8, pady=6)
            tb.insert("1.0", str(desc_raw))
            tb.configure(state="disabled")


class JobTrackerTab(ctk.CTkFrame):
    # Fonts created once, reused across all rows
    _font_site    = None
    _font_company = None
    _font_title   = None
    _font_small   = None
    _font_small_b = None
    _font_match   = None
    _font_verdict = None
    _font_link    = None

    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._df      = None
        self._page    = 0
        self._loading = False
        if JobTrackerTab._font_site is None:
            JobTrackerTab._font_site    = ctk.CTkFont(size=10)
            JobTrackerTab._font_company = ctk.CTkFont(size=11, weight="bold")
            JobTrackerTab._font_title   = ctk.CTkFont(size=11)
            JobTrackerTab._font_small   = ctk.CTkFont(size=10)
            JobTrackerTab._font_small_b = ctk.CTkFont(size=10, weight="bold")
            JobTrackerTab._font_match   = ctk.CTkFont(size=11, weight="bold")
            JobTrackerTab._font_verdict = ctk.CTkFont(size=10, weight="bold")
            JobTrackerTab._font_link    = ctk.CTkFont(size=14, weight="bold")
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Job Tracker",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self, text="Review and manage your AI-screened opportunities.",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(anchor="w", padx=30, pady=(0, 15))

        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=30, pady=(0, 10))

        ctk.CTkLabel(top_bar, text="Filter:").pack(side="left", padx=(0, 6))
        self.filter_var = ctk.StringVar(value="All")
        self.filter_menu = ctk.CTkSegmentedButton(top_bar, values=["All", "yes", "maybe", "no"],
                                                  variable=self.filter_var,
                                                  command=self._refresh_list, height=28)
        self.filter_menu.pack(side="left")

        ctk.CTkButton(top_bar, text="🗑 Remove All", width=100, height=28,
                      fg_color="#5a1a1a", hover_color="#7a1a1a",
                      command=self._remove_all).pack(side="right", padx=5)
        ctk.CTkButton(top_bar, text="🗑 Clear NOs", width=100, height=28,
                      fg_color="#442222", hover_color="#662222",
                      command=self._remove_nos).pack(side="right", padx=5)
        ctk.CTkButton(top_bar, text="🗑 Clear Not Applied", width=120, height=28,
                      fg_color="#442222", hover_color="#662222",
                      command=self._remove_not_applied).pack(side="right", padx=5)

        ctk.CTkButton(top_bar, text="🔄 Refresh Data", width=100, height=28,
                      fg_color="#333", hover_color="#444",
                      command=lambda: self._refresh_list(reload=True)).pack(side="right")

        ctk.CTkButton(top_bar, text="➕ Add Job by URL", width=130, height=28,
                      fg_color="#1f4e79", hover_color="#2563a8",
                      command=self._open_add_job_dialog).pack(side="right", padx=5)

        # Fixed column widths — header and rows must use identical values
        # (label_text, pixel_width, anchor)
        self._col_defs = [
            ("Site",     100, "w"),
            ("Company",  140, "w"),
            ("Position", 185, "w"),
            ("Location", 140, "w"),
            ("Years",     60, "w"),
            ("Level",     60, "w"),
            ("Match %",   65, "w"),
            ("Verdict",   80, "w"),
            ("Applied",   70, "w"),
        ]

        # Table Header
        header_frame = ctk.CTkFrame(self, fg_color="#161b22", height=45, corner_radius=8)
        header_frame.pack(fill="x", padx=30, pady=(0, 5))
        header_frame.pack_propagate(False)

        for text, width, anchor in self._col_defs:
            ctk.CTkLabel(header_frame, text=text, width=width, anchor=anchor,
                         font=ctk.CTkFont(weight="bold", size=11),
                         text_color="#8b949e").pack(side="left", padx=(6, 0), pady=8)

        # Table Content
        self.list_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.list_frame.pack(fill="both", expand=True, padx=30, pady=(0, 5))

        # Pagination bar
        pg_bar = ctk.CTkFrame(self, fg_color="#161b22", height=40, corner_radius=8)
        pg_bar.pack(fill="x", padx=30, pady=(0, 12))

        self._btn_prev = ctk.CTkButton(pg_bar, text="◀ Prev", width=80, height=28,
                                       fg_color="#333", hover_color="#444",
                                       command=self._prev_page)
        self._btn_prev.pack(side="left", padx=10, pady=6)

        self._page_label = ctk.CTkLabel(pg_bar, text="Page 1 of 1  (0 jobs)",
                                        font=ctk.CTkFont(size=12), text_color="#8b949e")
        self._page_label.pack(side="left", expand=True)

        self._btn_next = ctk.CTkButton(pg_bar, text="Next ▶", width=80, height=28,
                                       fg_color="#333", hover_color="#444",
                                       command=self._next_page)
        self._btn_next.pack(side="right", padx=10, pady=6)

        self._refresh_list()

    def _load_df(self):
        from db_manager import JobsDB
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        db = JobsDB()
        if db.count() == 0 and os.path.exists(excel_path):
            db.import_from_excel(excel_path)
        df = db.get_all()
        if df.empty:
            self._df = None
            return
        if "ai_recommendation" in df.columns:
            df = df.rename(columns={"ai_recommendation": "AI_recommendation"})
        self._df = df.drop(columns=["id", "created_at"], errors="ignore")

    def _refresh_list(self, *args, reload=False):
        if reload or self._df is None:
            self._loading = True
            self._draw_loading()
            threading.Thread(target=self._load_df_bg, daemon=True).start()
        else:
            self._page = 0
            self._draw_page()

    def _draw_loading(self):
        for w in self.list_frame.winfo_children():
            w.destroy()
        ctk.CTkLabel(self.list_frame, text="Loading jobs...",
                     font=ctk.CTkFont(size=13), text_color="#8b949e").pack(pady=30)

    def _load_df_bg(self):
        self._load_df()
        self.after(0, self._on_df_ready)

    def _on_df_ready(self):
        self._loading = False
        self._page = 0
        self._draw_page()

    def _draw_page(self):
        for w in self.list_frame.winfo_children():
            w.destroy()

        if self._df is None:
            ctk.CTkLabel(self.list_frame, text="No jobs found. Run the scraper first!").pack(pady=20)
            self._page_label.configure(text="Page 1 of 1  (0 jobs)")
            self._btn_prev.configure(state="disabled")
            self._btn_next.configure(state="disabled")
            return

        try:
            df = self._df
            filter_val = self.filter_var.get()
            if filter_val != "All":
                df = df[df["AI_recommendation"].str.lower() == filter_val.lower()]

            total = len(df)
            page_size = 25
            total_pages = max(1, (total + page_size - 1) // page_size)
            self._page = max(0, min(self._page, total_pages - 1))

            start = self._page * page_size
            page_df = df.iloc[start:start + page_size]

            for idx, row in page_df.iterrows():
                self._add_row(row, idx)

            self._page_label.configure(
                text=f"Page {self._page + 1} of {total_pages}  ({total} jobs)"
            )
            self._btn_prev.configure(state="normal" if self._page > 0 else "disabled")
            self._btn_next.configure(state="normal" if self._page < total_pages - 1 else "disabled")
        except Exception as e:
            ctk.CTkLabel(self.list_frame, text=f"Error loading jobs: {e}", text_color="red").pack(pady=20)

    def _prev_page(self):
        self._page -= 1
        self._draw_page()

    def _next_page(self):
        self._page += 1
        self._draw_page()

    def _open_add_job_dialog(self):
        AddJobDialog(self, self.config, on_saved=lambda: self._refresh_list(reload=True))

    @staticmethod
    def _site_info(site_val, link_val):
        """Return (emoji, display_name) from the site column or URL fallback."""
        _MAP = {
            "linkedin":      ("🔗", "LinkedIn"),
            "indeed":        ("💼", "Indeed"),
            "glassdoor":     ("🏢", "Glassdoor"),
            "jobright":      ("📌", "JobRight"),
            "naukri":        ("🎯", "Naukri"),
            "zip_recruiter": ("📮", "ZipRecruiter"),
            "ziprecruiter":  ("📮", "ZipRecruiter"),
            "google":        ("🔍", "Google"),
            "bayt":          ("🌍", "Bayt"),
            "bdjobs":        ("📋", "BDJobs"),
        }
        key = str(site_val).strip().lower()
        if key and key not in ("", "nan", "unknown", "web"):
            if key in _MAP:
                return _MAP[key]
            return ("🌐", site_val.title())
        # URL fallback
        url = str(link_val).lower()
        for k, v in _MAP.items():
            if k.replace("_", "") in url:
                return v
        return ("🌐", "Web")

    def _add_row(self, row, idx):
        row_frame = ctk.CTkFrame(self.list_frame, fg_color="#1c2128", height=46, corner_radius=6)
        row_frame.pack(fill="x", pady=2)
        row_frame.pack_propagate(False)

        verdict = str(row.get("AI_recommendation", "")).lower()
        verdict_color = "#238636" if verdict == "yes" else "#d29922" if verdict == "maybe" else "#da3633"

        emoji, site_name = self._site_info(row.get("site", ""), row.get("link", ""))

        # All labels use fixed width + pack(side="left") so columns align perfectly
        PAD = (6, 0)

        # 0. Site
        site_lbl = ctk.CTkLabel(row_frame, text=f"{emoji} {site_name}", width=100, anchor="w",
                                 font=self._font_site, text_color="#58a6ff")
        site_lbl.pack(side="left", padx=PAD, pady=8)

        # 1. Company
        company = str(row.get("company", "Unknown"))[:22]
        c_lbl = ctk.CTkLabel(row_frame, text=company, width=140, anchor="w", font=self._font_company)
        c_lbl.pack(side="left", padx=PAD, pady=8)

        # 2. Position
        title = str(row.get("title", "Unknown"))[:28]
        p_lbl = ctk.CTkLabel(row_frame, text=title, width=185, anchor="w", font=self._font_title)
        p_lbl.pack(side="left", padx=PAD, pady=8)

        # 3. Location
        loc_raw = row.get("location", "")
        loc_str = "" if (loc_raw is None or str(loc_raw).lower() in ("nan", "none", "")) else str(loc_raw)
        loc_lbl = ctk.CTkLabel(row_frame, text=loc_str[:22], width=140, anchor="w",
                                font=self._font_small, text_color="#8b949e")
        loc_lbl.pack(side="left", padx=PAD, pady=8)

        # 4. Years Required
        years_raw = str(row.get("years_required", "N/A"))
        years = years_raw if years_raw not in ("", "nan", "None") else "N/A"
        y_lbl = ctk.CTkLabel(row_frame, text=years[:8], width=60, anchor="w",
                              font=self._font_small, text_color="#79c0ff")
        y_lbl.pack(side="left", padx=PAD, pady=8)

        # 5. Level
        level_raw = str(row.get("role_level", "")).lower()
        level = {"junior": "JR", "mid": "MID", "senior": "SR"}.get(level_raw, level_raw[:3].upper() or "N/A")
        lv_lbl = ctk.CTkLabel(row_frame, text=level, width=60, anchor="w",
                               font=self._font_small_b, text_color="#a5d6ff")
        lv_lbl.pack(side="left", padx=PAD, pady=8)

        # 6. Match %
        try:
            match_pct = int(row.get("skills_match_pct", 0))
            match_color = "#238636" if match_pct >= 70 else "#d29922" if match_pct >= 45 else "#da3633"
            match_text = f"{match_pct}%"
        except (ValueError, TypeError):
            match_color = "#8b949e"
            match_text = "N/A"
        m_lbl = ctk.CTkLabel(row_frame, text=match_text, width=65, anchor="w",
                              font=self._font_match, text_color=match_color)
        m_lbl.pack(side="left", padx=PAD, pady=8)

        # 7. Verdict
        v_lbl = ctk.CTkLabel(row_frame, text=f"● {verdict.upper()}", width=80, anchor="w",
                              font=self._font_verdict, text_color=verdict_color)
        v_lbl.pack(side="left", padx=PAD, pady=8)

        # 8. Applied toggle (in a fixed-width container so switch doesn't shift columns)
        sw_frame = ctk.CTkFrame(row_frame, fg_color="transparent", width=70)
        sw_frame.pack(side="left", padx=PAD, pady=8)
        sw_frame.pack_propagate(False)
        status = row.get("application_status", "Not Applied")
        app_switch = ctk.CTkSwitch(sw_frame, text="", width=40, height=20,
                                   variable=ctk.BooleanVar(value=(status == "Applied")),
                                   command=lambda l=row.get("link"): self._update_application_status(l))
        app_switch.pack(side="left", pady=10)

        # Entire row opens detail window on click
        for widget in [site_lbl, c_lbl, p_lbl, loc_lbl, y_lbl, lv_lbl, m_lbl, v_lbl, row_frame]:
            widget.bind("<Button-1>", lambda e, r=row: self._open_details(r))

    def _open_details(self, row):
        JobDetailWindow(self, self.config, row.to_dict(), self._update_verdict)

    def _migrate_to_applied(self, job_link):
        """
        Moves a job record from jobs.xlsx to Job-Tracker.xlsx and
        moves generated files to the structured applied folder.
        """
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        applied_excel = self.config.get("pipeline", "applied_excel_path", default="Job-Tracker.xlsx")
        applied_folder_root = self.config.get("pipeline", "applied_folder_path", default="fultime")
        output_root = self.config.get("pipeline", "output_dir", default="outputs")

        try:
            df = pd.read_excel(excel_path)
            job_row = df[df["link"] == job_link]

            if job_row.empty:
                logging.warning(f"Migration failed: No job found with link {job_link}")
                return False

            # Compute paths up-front so we can store them in the Excel record
            company = str(job_row.iloc[0].get("company", "Unknown_Company"))
            title   = str(job_row.iloc[0].get("title",   "Unknown_Position"))

            def sanitize(name):
                import re
                name = re.sub(r"[^\w\s-]", "", name)
                return re.sub(r"\s+", "_", name.strip())[:60] or "Unknown"

            company_s = sanitize(company)
            title_s   = sanitize(title)
            date_str  = datetime.now().strftime("%Y-%m-%d")

            src_dir  = os.path.join(output_root,         company_s, title_s)
            dest_dir = os.path.join(applied_folder_root, company_s, f"{title_s}_{date_str}")

            # 1. Data Migration — stamp applied_date + folder_path
            record = job_row.iloc[0].to_frame().T
            record["applied_date"] = date_str
            record["folder_path"]  = dest_dir

            # Load or create Job-Tracker.xlsx
            if os.path.exists(applied_excel):
                app_df = pd.read_excel(applied_excel)
                app_df = pd.concat([app_df, record], ignore_index=True)
            else:
                app_df = record

            app_df.to_excel(applied_excel, index=False)

            # Remove from main jobs.xlsx
            df = df[df["link"] != job_link]
            df.to_excel(excel_path, index=False)

            # 2. File Migration

            logging.info(f"Attempting to move files: {src_dir} -> {dest_dir}")

            if os.path.exists(src_dir):
                os.makedirs(os.path.dirname(dest_dir), exist_ok=True)
                shutil.move(src_dir, dest_dir)
                logging.info(f"Successfully moved files to {dest_dir}")
            else:
                logging.warning(f"Source directory not found: {src_dir}. Files might not have been generated.")

            return True
        except Exception as e:
            logging.error(f"Migration failed with error: {e}")
            return False

    def _update_application_status(self, job_link):
        if not job_link: return
        from db_manager import JobsDB
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        try:
            df = pd.read_excel(excel_path)

            if "application_status" not in df.columns:
                df["application_status"] = "Not Applied"
            else:
                df["application_status"] = df["application_status"].astype(str)

            current_status = df.loc[df["link"] == job_link, "application_status"].values
            new_status = "Applied" if len(current_status) == 0 or current_status[0] != "Applied" else "Not Applied"

            df.loc[df["link"] == job_link, "application_status"] = new_status
            df.to_excel(excel_path, index=False)

            # Sync to SQLite
            try:
                JobsDB().set_application_status(str(job_link), new_status)
            except Exception as _dbe:
                logging.warning("SQLite status sync failed: %s", _dbe)

            if new_status == "Applied":
                if self._migrate_to_applied(job_link):
                    messagebox.showinfo("Applied", f"Job moved to {self.config.get('pipeline', 'applied_excel_path')} and files archived.")
                else:
                    messagebox.showwarning("Partial Success", "Updated status, but could not migrate files/record.")

            self._df = None
            self._refresh_list(reload=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not update application status: {e}")

    def _update_verdict(self, job_link, new_verdict, status="Not Applied"):
        from db_manager import JobsDB
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        try:
            df = pd.read_excel(excel_path)
            if "AI_recommendation" in df.columns:
                df["AI_recommendation"] = df["AI_recommendation"].astype(str)
            if "application_status" in df.columns:
                df["application_status"] = df["application_status"].astype(str)
            else:
                df["application_status"] = "Not Applied"

            df.loc[df["link"] == job_link, "AI_recommendation"] = new_verdict
            df.loc[df["link"] == job_link, "application_status"] = status
            df.to_excel(excel_path, index=False)

            # Sync to SQLite
            try:
                db = JobsDB()
                db.set_verdict(str(job_link), new_verdict)
                db.set_application_status(str(job_link), status)
            except Exception as _dbe:
                logging.warning("SQLite verdict sync failed: %s", _dbe)

            if status == "Applied":
                self._migrate_to_applied(job_link)

            self._df = None
            self._refresh_list(reload=True)
        except Exception as e:
            messagebox.showerror("Save Error", f"Could not update record: {e}")

    def _remove_nos(self):
        from db_manager import JobsDB
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        db = JobsDB()
        if db.count() == 0 and os.path.exists(excel_path):
            db.import_from_excel(excel_path)
        try:
            before = db.count()
            db.delete_by_verdict("no")
            removed = before - db.count()
            db.sync_to_excel(excel_path)
            messagebox.showinfo("Cleaned", f"Removed {removed} 'No' verdict records.")
            self._df = None
            self._refresh_list(reload=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not remove NOs: {e}")

    def _remove_not_applied(self):
        from db_manager import JobsDB
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        db = JobsDB()
        if db.count() == 0 and os.path.exists(excel_path):
            db.import_from_excel(excel_path)
        try:
            before = db.count()
            db.delete_not_applied()
            removed = before - db.count()
            db.sync_to_excel(excel_path)
            messagebox.showinfo("Cleaned", f"Removed {removed} not-applied records.")
            self._df = None
            self._refresh_list(reload=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not remove Not Applied: {e}")

    def _remove_all(self):
        from db_manager import JobsDB
        if not messagebox.askyesno("Remove All",
                                   "Remove ALL jobs from the tracker?\nThis cannot be undone."):
            return
        excel_path = self.config.get("pipeline", "excel_path", default="jobs.xlsx")
        db = JobsDB()
        try:
            db.delete_all()
            db.sync_to_excel(excel_path)
            self._df = None
            self._refresh_list(reload=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not remove all: {e}")

# ══════════════════════════════════════════════════════════════
#  TAB: Applied
# ══════════════════════════════════════════════════════════════
class AppliedTab(ctk.CTkFrame):
    _font_company = None
    _font_title   = None
    _font_small   = None
    _font_match   = None
    _font_verdict = None

    _PAGE_SIZE = 25

    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config
        self._df          = None
        self._filtered_df = None
        self._page        = 0
        if AppliedTab._font_company is None:
            AppliedTab._font_company = ctk.CTkFont(size=11, weight="bold")
            AppliedTab._font_title   = ctk.CTkFont(size=11)
            AppliedTab._font_small   = ctk.CTkFont(size=10)
            AppliedTab._font_match   = ctk.CTkFont(size=11, weight="bold")
            AppliedTab._font_verdict = ctk.CTkFont(size=10, weight="bold")
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Applied Jobs",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)

        applied_path = self.config.get("pipeline", "applied_db_path", default="applied.db")
        self._file_label = ctk.CTkLabel(
            self,
            text=f"Reading from: {applied_path}",
            font=ctk.CTkFont(size=11), text_color="#58a6ff"
        )
        self._file_label.pack(anchor="w", padx=30, pady=(0, 12))

        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=30, pady=(0, 10))

        ctk.CTkLabel(top_bar, text="Search:", font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 6))
        self._search_var = ctk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._filter())
        ctk.CTkEntry(top_bar, textvariable=self._search_var, width=280,
                     placeholder_text="Company, title, location, skills...",
                     fg_color="#1c2128", border_color="#30363d",
                     font=ctk.CTkFont(size=12)).pack(side="left", padx=(0, 10))

        ctk.CTkButton(top_bar, text="🔄 Refresh", width=90, height=28,
                      fg_color="#333", hover_color="#444",
                      command=lambda: self._load(reload=True)).pack(side="right")

        self._count_label = ctk.CTkLabel(top_bar, text="", font=ctk.CTkFont(size=11),
                                          text_color="#8b949e")
        self._count_label.pack(side="right", padx=10)

        # Column definitions — now driven by SQLite applied_jobs table
        self._col_defs = [
            ("Company",      140, "w"),
            ("Position",     175, "w"),
            ("Location",     110, "w"),
            ("Applied",       95, "w"),
            ("Score",         65, "w"),
            ("Status",        75, "w"),
            ("Actions",      225, "w"),
        ]

        header_frame = ctk.CTkFrame(self, fg_color="#161b22", height=45, corner_radius=8)
        header_frame.pack(fill="x", padx=30, pady=(0, 5))
        header_frame.pack_propagate(False)
        for text, width, anchor in self._col_defs:
            ctk.CTkLabel(header_frame, text=text, width=width, anchor=anchor,
                         font=ctk.CTkFont(weight="bold", size=11),
                         text_color="#8b949e").pack(side="left", padx=(6, 0), pady=8)

        self.list_frame = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.list_frame.pack(fill="both", expand=True, padx=30, pady=(0, 5))

        # Pagination bar
        pg_bar = ctk.CTkFrame(self, fg_color="#161b22", height=40, corner_radius=8)
        pg_bar.pack(fill="x", padx=30, pady=(0, 12))

        self._btn_prev = ctk.CTkButton(pg_bar, text="◀ Prev", width=80, height=28,
                                       fg_color="#333", hover_color="#444",
                                       command=self._prev_page)
        self._btn_prev.pack(side="left", padx=10, pady=6)

        self._page_label = ctk.CTkLabel(pg_bar, text="Page 1 of 1  (0 jobs)",
                                        font=ctk.CTkFont(size=12), text_color="#8b949e")
        self._page_label.pack(side="left", expand=True)

        self._btn_next = ctk.CTkButton(pg_bar, text="Next ▶", width=80, height=28,
                                       fg_color="#333", hover_color="#444",
                                       command=self._next_page)
        self._btn_next.pack(side="right", padx=10, pady=6)

        self._load()

    def _load(self, reload=False):
        from db_manager import AppliedDB
        applied_db_path = self.config.get("pipeline", "applied_db_path", default="applied.db")
        self._file_label.configure(text=f"Reading from: {applied_db_path}")
        try:
            adb = AppliedDB(applied_db_path)
            df = adb.get_all()
            self._df = df if not df.empty else None
        except Exception as e:
            self._df = None
            logging.warning("Could not load applied.db: %s", e)
        self._search_var.set("")
        self._filter()

    def _filter(self):
        query = self._search_var.get().strip().lower()
        if self._df is None:
            self._filtered_df = None
        elif not query:
            self._filtered_df = self._df
        else:
            search_cols = ["company", "title", "location", "status", "job_url"]
            mask = pd.Series([False] * len(self._df), index=self._df.index)
            for col in search_cols:
                if col in self._df.columns:
                    mask = mask | self._df[col].astype(str).str.lower().str.contains(query, na=False)
            self._filtered_df = self._df[mask]
        self._page = 0
        self._render_rows()

    def _prev_page(self):
        self._page = max(0, self._page - 1)
        self._render_rows()

    def _next_page(self):
        self._page += 1
        self._render_rows()

    def _render_rows(self):
        for w in self.list_frame.winfo_children():
            w.destroy()

        if self._filtered_df is None:
            self._count_label.configure(text="0 jobs")
            self._page_label.configure(text="Page 1 of 1  (0 jobs)")
            self._btn_prev.configure(state="disabled")
            self._btn_next.configure(state="disabled")
            ctk.CTkLabel(self.list_frame,
                         text="No applied jobs in database yet.\nResumes generated by the pipeline will appear here.",
                         justify="center", text_color="#8b949e").pack(pady=40)
            return

        total = len(self._filtered_df)
        if total == 0:
            self._count_label.configure(text="0 jobs")
            self._page_label.configure(text="Page 1 of 1  (0 jobs)")
            self._btn_prev.configure(state="disabled")
            self._btn_next.configure(state="disabled")
            ctk.CTkLabel(self.list_frame,
                         text="No jobs match your search." if self._search_var.get().strip()
                              else "No applied jobs found.",
                         justify="center", text_color="#8b949e").pack(pady=40)
            return

        PAGE = self._PAGE_SIZE
        total_pages = max(1, (total + PAGE - 1) // PAGE)
        self._page = max(0, min(self._page, total_pages - 1))
        start = self._page * PAGE
        page_df = self._filtered_df.iloc[start:start + PAGE]

        self._count_label.configure(text=f"{total} job{'s' if total != 1 else ''}")
        self._page_label.configure(text=f"Page {self._page + 1} of {total_pages}  ({total} jobs)")
        self._btn_prev.configure(state="normal" if self._page > 0 else "disabled")
        self._btn_next.configure(state="normal" if self._page < total_pages - 1 else "disabled")

        for _, row in page_df.iterrows():
            self._add_row(row)

    def _add_row(self, row):
        PAD = (6, 0)
        job_id  = int(row.get("id", 0))
        company = str(row.get("company", "Unknown"))
        title   = str(row.get("title",   "Unknown"))

        status_raw = str(row.get("status", "")).lower()
        status_color = (
            "#3fb950" if status_raw == "done"
            else "#58a6ff" if status_raw == "tex_only"
            else "#d29922"
        )

        row_frame = ctk.CTkFrame(self.list_frame, fg_color="#1c2128", height=46, corner_radius=6)
        row_frame.pack(fill="x", pady=2)
        row_frame.pack_propagate(False)

        # Company
        c_lbl = ctk.CTkLabel(row_frame, text=company[:20], width=140, anchor="w",
                              font=self._font_company)
        c_lbl.pack(side="left", padx=PAD, pady=8)

        # Title
        t_lbl = ctk.CTkLabel(row_frame, text=title[:26], width=175, anchor="w",
                              font=self._font_title)
        t_lbl.pack(side="left", padx=PAD, pady=8)

        # Location
        loc_raw = row.get("location", "")
        loc_str = "" if (not loc_raw or str(loc_raw).lower() in ("nan", "none", "")) else str(loc_raw)[:16]
        l_lbl = ctk.CTkLabel(row_frame, text=loc_str, width=110, anchor="w",
                              font=self._font_small, text_color="#8b949e")
        l_lbl.pack(side="left", padx=PAD, pady=8)

        # Applied Date
        applied_raw = row.get("applied_date", "")
        date_str = "" if (not applied_raw or str(applied_raw).lower() in ("nan", "none", "nat")) else str(applied_raw)[:10]
        d_lbl = ctk.CTkLabel(row_frame, text=date_str, width=95, anchor="w",
                              font=self._font_small, text_color="#3fb950")
        d_lbl.pack(side="left", padx=PAD, pady=8)

        # ATS Score
        try:
            score_val  = int(row.get("ats_score", 0))
            score_clr  = "#3fb950" if score_val >= 85 else "#d29922" if score_val >= 70 else "#da3633"
            score_text = f"{score_val}%"
        except (ValueError, TypeError):
            score_clr  = "#8b949e"
            score_text = "N/A"
        s_lbl = ctk.CTkLabel(row_frame, text=score_text, width=65, anchor="w",
                              font=self._font_match, text_color=score_clr)
        s_lbl.pack(side="left", padx=PAD, pady=8)

        # Status
        st_lbl = ctk.CTkLabel(row_frame, text=status_raw.upper()[:8], width=75, anchor="w",
                               font=self._font_verdict, text_color=status_color)
        st_lbl.pack(side="left", padx=PAD, pady=8)

        # Action buttons (Compile | View .tex | Delete)
        act_frame = ctk.CTkFrame(row_frame, fg_color="transparent", width=225)
        act_frame.pack(side="left", padx=(6, 4), pady=5)
        act_frame.pack_propagate(False)

        applied_excel = self.config.get("pipeline", "applied_excel_path", default="Job-Tracker.xlsx")
        applied_db    = self.config.get("pipeline", "applied_db_path",    default="applied.db")

        ctk.CTkButton(act_frame, text="Compile PDF", width=82, height=26,
                      font=ctk.CTkFont(size=10),
                      fg_color="#0d3042", hover_color="#1455a4",
                      command=lambda jid=job_id, c=company, t=title, adb=applied_db:
                          self._compile_pdf(jid, c, t, adb)
                      ).pack(side="left", padx=(0, 3))

        ctk.CTkButton(act_frame, text="View .tex", width=70, height=26,
                      font=ctk.CTkFont(size=10),
                      fg_color="#1a2a1a", hover_color="#1e4d1e",
                      command=lambda jid=job_id, c=company, t=title, adb=applied_db:
                          self._view_tex(jid, c, t, adb)
                      ).pack(side="left", padx=(0, 3))

        ctk.CTkButton(act_frame, text="Delete", width=60, height=26,
                      font=ctk.CTkFont(size=10),
                      fg_color="#3a1010", hover_color="#6a1010",
                      command=lambda jid=job_id, c=company, t=title, adb=applied_db, ae=applied_excel:
                          self._delete_applied(jid, c, t, adb, ae)
                      ).pack(side="left")

        for widget in [c_lbl, t_lbl, l_lbl, d_lbl, s_lbl, st_lbl, row_frame]:
            widget.bind("<Button-1>", lambda e, r=row: self._open_detail(r))

    def _open_detail(self, row):
        AppliedDetailWindow(self, row.to_dict(), config=self.config)

    # ── Action handlers ───────────────────────────────────────────────────────

    def _compile_pdf(self, job_id: int, company: str, title: str, applied_db_path: str):
        from db_manager import AppliedDB
        import threading
        db  = AppliedDB(applied_db_path)
        tex = db.get_tex(job_id)
        if not tex:
            messagebox.showwarning("No LaTeX",
                f"No .tex content stored for:\n{title} @ {company}\n\n"
                "Only resumes generated by the pipeline have stored LaTeX.")
            return

        safe_co = re.sub(r"[^\w\s-]", "", company).strip().replace(" ", "_")[:40]
        safe_ti = re.sub(r"[^\w\s-]", "", title  ).strip().replace(" ", "_")[:40]
        out_dir = os.path.join("compiled", safe_co, safe_ti)
        os.makedirs(out_dir, exist_ok=True)
        pdf_path = os.path.join(out_dir, "Resume.pdf")

        def _do():
            try:
                from utils.latex_compiler import compile_latex_to_pdf
                ok = compile_latex_to_pdf(tex, pdf_path)
                if ok:
                    messagebox.showinfo("Compiled!", f"PDF saved:\n{os.path.abspath(pdf_path)}")
                else:
                    messagebox.showerror("Compile Failed",
                        "pdflatex returned an error. Make sure pdflatex is installed.")
            except Exception as e:
                messagebox.showerror("Compile Error", str(e))

        threading.Thread(target=_do, daemon=True).start()

    def _view_tex(self, job_id: int, company: str, title: str, applied_db_path: str):
        from db_manager import AppliedDB
        db  = AppliedDB(applied_db_path)
        tex = db.get_tex(job_id)
        if not tex:
            messagebox.showwarning("No LaTeX",
                f"No .tex content stored for:\n{title} @ {company}")
            return

        win = ctk.CTkToplevel(self)
        win.title(f".tex — {title} @ {company}")
        win.geometry("920x620")
        win.grab_set()

        top = ctk.CTkFrame(win, fg_color="transparent")
        top.pack(fill="x", padx=10, pady=(8, 0))
        ctk.CTkLabel(top, text=f"{title} @ {company}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left")
        ctk.CTkButton(top, text="Copy All", width=80, height=26,
                      command=lambda: [win.clipboard_clear(),
                                       win.clipboard_append(tex),
                                       win.update()]
                      ).pack(side="right")

        tb = ctk.CTkTextbox(win, wrap="none",
                            font=ctk.CTkFont(family="Courier New", size=11))
        tb.pack(fill="both", expand=True, padx=10, pady=8)
        tb.insert("end", tex)
        tb.configure(state="disabled")

    def _delete_applied(self, job_id: int, company: str, title: str,
                        applied_db_path: str, applied_excel: str):
        from db_manager import AppliedDB
        if not messagebox.askyesno("Delete",
                f"Remove from applied tracker:\n{title} @ {company}?"):
            return
        db = AppliedDB(applied_db_path)
        db.delete_by_id(job_id)
        db.sync_to_excel(applied_excel)
        self._load(reload=True)


# ══════════════════════════════════════════════════════════════
#  TAB: Resume Order
# ══════════════════════════════════════════════════════════════
class ResumeOrderTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config = config

        self.base_section_names = {
            "education":  "Education",
            "skills":     "Technical Skills",
            "experience": "Experience",
            "projects":   "Relevant Projects",
        }

        # Build combined names map including custom sections
        custom_sections = self.config.get("custom_sections", default=[])
        self.section_names = dict(self.base_section_names)
        for cs in custom_sections:
            cs_id = cs.get("id")
            if cs_id:
                self.section_names[cs_id] = cs.get("name", cs_id)

        base_keys   = list(self.base_section_names.keys())
        custom_keys = [cs.get("id") for cs in custom_sections if cs.get("id")]
        default_order = base_keys + custom_keys

        saved = self.config.get("section_order", default=None)
        if saved and isinstance(saved, list):
            valid = [s for s in saved if s in self.section_names]
            for key in default_order:
                if key not in valid:
                    valid.append(key)
            self.sections = valid
        else:
            self.sections = default_order

        self._build()

    def _reload_sections(self):
        """Re-read custom sections from config and sync the order list."""
        custom_sections = self.config.get("custom_sections", default=[])
        self.section_names = dict(self.base_section_names)
        for cs in custom_sections:
            cs_id = cs.get("id")
            if cs_id:
                self.section_names[cs_id] = cs.get("name", cs_id)

        all_known = set(self.section_names.keys())
        valid = [s for s in self.sections if s in all_known]
        for key in list(self.base_section_names.keys()) + [cs.get("id") for cs in custom_sections if cs.get("id")]:
            if key not in valid:
                valid.append(key)
        self.sections = valid
        self._refresh_list()

    def _build(self):
        ctk.CTkLabel(self, text="Resume Section Order",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self,
                     text="Customize the order of sections in your LaTeX resume. Use the buttons to move sections up and down.",
                     text_color="gray", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=30, pady=(0, 15))

        top_bar = ctk.CTkFrame(self, fg_color="transparent")
        top_bar.pack(fill="x", padx=30, pady=(0, 5))
        ctk.CTkButton(top_bar, text="🔄 Refresh Sections", width=150, height=28,
                      fg_color="#333", hover_color="#444",
                      command=self._reload_sections).pack(side="left")

        self.list_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.list_frame.pack(fill="both", expand=True, padx=30)

        self._refresh_list()

        ctk.CTkButton(self, text="💾 Save Section Order", height=42,
                      command=self._save_order).pack(padx=30, pady=20, fill="x")

    def _refresh_list(self):
        for w in self.list_frame.winfo_children():
            w.destroy()

        for i, section in enumerate(self.sections):
            row = ctk.CTkFrame(self.list_frame, corner_radius=8)
            row.pack(fill="x", pady=5)

            name = self.section_names.get(section, section)
            ctk.CTkLabel(row, text=f"{i+1}. {name}",
                         font=ctk.CTkFont(size=13, weight="bold")).pack(side="left", padx=15, pady=10)

            btn_frame = ctk.CTkFrame(row, fg_color="transparent")
            btn_frame.pack(side="right", padx=10)

            ctk.CTkButton(btn_frame, text="▲", width=30, height=30,
                          fg_color="#333", hover_color="#444",
                          command=lambda idx=i: self._move_up(idx)).pack(side="left", padx=2)
            ctk.CTkButton(btn_frame, text="▼", width=30, height=30,
                          fg_color="#333", hover_color="#444",
                          command=lambda idx=i: self._move_down(idx)).pack(side="left", padx=2)

    def _move_up(self, index):
        if index > 0:
            self.sections[index], self.sections[index-1] = self.sections[index-1], self.sections[index]
            self._refresh_list()

    def _move_down(self, index):
        if index < len(self.sections) - 1:
            self.sections[index], self.sections[index+1] = self.sections[index+1], self.sections[index]
            self._refresh_list()

    def _save_order(self):
        self.config.set("section_order", self.sections)
        self.config.save()
        messagebox.showinfo("Saved", "Resume section order saved successfully!")

# ══════════════════════════════════════════════════════════════
#  TAB: Logs
# ══════════════════════════════════════════════════════════════
class LogsTab(ctk.CTkFrame):

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._build()

    def _build(self):
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=30, pady=(20, 8))

        ctk.CTkLabel(header, text="Application Logs",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(side="left")

        ctk.CTkButton(header, text="Clear", width=80, height=32,
                      fg_color="#333", hover_color="#444",
                      command=self.clear).pack(side="right")

        self.log_widget = LogWidget(self)
        self.log_widget.pack(fill="both", expand=True, padx=30, pady=(0, 20))

    def append(self, level, msg):
        self.log_widget.append(level, msg)

    def clear(self):
        self.log_widget.clear()

class ResumeTab(ctk.CTkFrame):
    def __init__(self, master, config, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.config   = config
        self.exp_rows = []
        self._build()

    def _build(self):
        ctk.CTkLabel(self, text="Resume & Projects",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(
                     pady=(20, 4), anchor="w", padx=30)
        ctk.CTkLabel(self,
                     text="Edit your resume, projects, and work experience used by the pipeline.",
                     text_color="gray", font=ctk.CTkFont(size=12)).pack(
                     anchor="w", padx=30, pady=(0, 12))

        # ── Tabview for sub-sections ──────────────────────────
        self.tabview = ctk.CTkTabview(self)
        self.tabview.pack(fill="both", expand=True, padx=30, pady=(0, 20))

        self.tabview.add("📄 Resume Text")
        self.tabview.add("📦 Projects")
        self.tabview.add("💼 Work Experience")

        self._build_resume_tab()
        self._build_projects_tab()
        self._build_experience_tab()

    # ── Resume text ───────────────────────────────────────────
    def _build_resume_tab(self):
        tab = self.tabview.tab("📄 Resume Text")

        header = ctk.CTkFrame(tab, fg_color="transparent")
        header.pack(fill="x", pady=(8, 4))
        ctk.CTkLabel(header, text="Paste your full resume text here",
                     text_color="gray", font=ctk.CTkFont(size=11)).pack(side="left")
        ctk.CTkButton(header, text="🗑 Clear", width=80, height=28,
                      fg_color="#442222", hover_color="#662222",
                      command=self._clear_resume).pack(side="right", padx=4)
        ctk.CTkButton(header, text="Load File", width=90, height=28,
                      fg_color="#333", hover_color="#444",
                      command=self._load_resume).pack(side="right", padx=4)
        ctk.CTkButton(header, text="💾 Save", width=80, height=28,
                      command=self._save_resume).pack(side="right")

        self.resume_box = ctk.CTkTextbox(
            tab, font=ctk.CTkFont(family="Consolas", size=11), wrap="word")
        self.resume_box.pack(fill="both", expand=True, pady=(4, 0))

        rp = self.config.get("pipeline", "resume_path", default="resume.txt")
        if os.path.exists(rp):
            with open(rp, "r", encoding="utf-8") as f:
                self.resume_box.insert("1.0", f.read())

    # ── Projects ──────────────────────────────────────────────
    def _build_projects_tab(self):
        tab = self.tabview.tab("📦 Projects")

        header = ctk.CTkFrame(tab, fg_color="transparent")
        header.pack(fill="x", pady=(8, 4))
        ctk.CTkLabel(header, text="List your projects — one per section",
                     text_color="gray", font=ctk.CTkFont(size=11)).pack(side="left")
        ctk.CTkButton(header, text="🌐 Import GitHub", width=110, height=28,
                      fg_color="#333", hover_color="#444",
                      command=self._import_github_project).pack(side="right", padx=4)
        ctk.CTkButton(header, text="➕ Add Link", width=100, height=28,
                      fg_color="#333", hover_color="#444",
                      command=self._add_project_link).pack(side="right", padx=4)
        ctk.CTkButton(header, text="Load File", width=90, height=28,
                      fg_color="#333", hover_color="#444",
                      command=self._load_projects).pack(side="right", padx=4)
        ctk.CTkButton(header, text="💾 Save", width=80, height=28,
                      command=self._save_projects).pack(side="right")

        self.projects_box = ctk.CTkTextbox(
            tab, font=ctk.CTkFont(family="Consolas", size=11), wrap="word")
        self.projects_box.pack(fill="both", expand=True, pady=(4, 0))

        pp = self.config.get("pipeline", "projects_path", default="Projects.txt")
        if os.path.exists(pp):
            with open(pp, "r", encoding="utf-8") as f:
                self.projects_box.insert("1.0", f.read())

    # ── Work Experience ───────────────────────────────────────
    def _build_experience_tab(self):
        tab = self.tabview.tab("💼 Work Experience")

        info = ctk.CTkLabel(tab,
            text="Define your work history. These are injected into the experience prompt automatically.",
            text_color="gray", font=ctk.CTkFont(size=11))
        info.pack(anchor="w", pady=(8, 4))

        # Scrollable container for experience rows
        self.exp_scroll = ctk.CTkScrollableFrame(tab, fg_color="transparent")
        self.exp_scroll.pack(fill="both", expand=True, pady=(4, 8))

        # Load saved experience or use defaults
        saved_exp = self.config.get("experience_roles", default=None)
        if saved_exp:
            roles = saved_exp
        else:
            roles = [
                {
                    "title":       "Job Title",
                    "company":     "Company Name",
                    "dates":       "Month Year – Month Year",
                    "domain":      "Your Domain",
                    "total_bullets":  4,
                    "real_bullets":   2,
                    "fabricated_bullets": 2,
                },
            ]

        self.exp_rows = []
        for role in roles:
            self._add_experience_row(role)

        # Bottom buttons
        btn_frame = ctk.CTkFrame(tab, fg_color="transparent")
        btn_frame.pack(fill="x", pady=4)
        ctk.CTkButton(btn_frame, text="+ Add Role", width=110, height=32,
                      fg_color="#333", hover_color="#444",
                      command=lambda: self._add_experience_row()).pack(side="left")
        ctk.CTkButton(btn_frame, text="💾 Save All", width=110, height=32,
                      command=self._save_experience).pack(side="right")

    def _add_experience_row(self, role: dict = None):
        role = role or {
            "title": "", "company": "", "dates": "",
            "domain": "", "total_bullets": 4,
            "real_bullets": 2, "fabricated_bullets": 2,
        }

        card = ctk.CTkFrame(self.exp_scroll, corner_radius=8)
        card.pack(fill="x", pady=5)

        # Row index label
        idx = len(self.exp_rows) + 1
        ctk.CTkLabel(card, text=f"Role {idx}",
                     font=ctk.CTkFont(weight="bold"),
                     text_color="#58a6ff").pack(anchor="w", padx=12, pady=(8, 4))

        fields = {}

        # Title + Company on same line
        row1 = ctk.CTkFrame(card, fg_color="transparent")
        row1.pack(fill="x", padx=12, pady=3)
        row1.grid_columnconfigure(1, weight=1)
        row1.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(row1, text="Job Title", width=80, anchor="w").grid(row=0, column=0, sticky="w")
        fields["title"] = ctk.CTkEntry(row1)
        fields["title"].insert(0, role.get("title", ""))
        fields["title"].grid(row=0, column=1, sticky="ew", padx=(4, 12))

        ctk.CTkLabel(row1, text="Company", width=70, anchor="w").grid(row=0, column=2, sticky="w")
        fields["company"] = ctk.CTkEntry(row1)
        fields["company"].insert(0, role.get("company", ""))
        fields["company"].grid(row=0, column=3, sticky="ew", padx=(4, 0))

        # Dates + Domain on same line
        row2 = ctk.CTkFrame(card, fg_color="transparent")
        row2.pack(fill="x", padx=12, pady=3)
        row2.grid_columnconfigure(1, weight=1)
        row2.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(row2, text="Dates", width=80, anchor="w").grid(row=0, column=0, sticky="w")
        fields["dates"] = ctk.CTkEntry(row2, placeholder_text="Jan 2024 – Dec 2024")
        fields["dates"].insert(0, role.get("dates", ""))
        fields["dates"].grid(row=0, column=1, sticky="ew", padx=(4, 12))

        ctk.CTkLabel(row2, text="Domain", width=70, anchor="w").grid(row=0, column=2, sticky="w")
        fields["domain"] = ctk.CTkEntry(row2, placeholder_text="e.g. AI/ML, full-stack, APIs")
        fields["domain"].insert(0, role.get("domain", ""))
        fields["domain"].grid(row=0, column=3, sticky="ew", padx=(4, 0))

        # Bullet counts
        row3 = ctk.CTkFrame(card, fg_color="transparent")
        row3.pack(fill="x", padx=12, pady=(3, 8))

        for col, (label, key, default) in enumerate([
            ("Total Bullets",      "total_bullets",      4),
            ("Real Bullets",       "real_bullets",       2),
            ("Fabricated Bullets", "fabricated_bullets", 2),
        ]):
            ctk.CTkLabel(row3, text=label, anchor="w").grid(row=0, column=col*2, padx=(0 if col==0 else 16, 4), sticky="w")
            e = ctk.CTkEntry(row3, width=50)
            e.insert(0, str(role.get(key, default)))
            e.grid(row=0, column=col*2+1, sticky="w")
            fields[key] = e

        # Delete button
        del_btn = ctk.CTkButton(row3, text="✕ Remove", width=90, height=26,
                                fg_color="#6e1a1a", hover_color="#8b2121",
                                command=lambda c=card, f=fields: self._remove_role(c, f))
        del_btn.grid(row=0, column=6, padx=(20, 0))
        row3.grid_columnconfigure(6, weight=1)

        self.exp_rows.append(fields)

    def _remove_role(self, card, fields):
        self.exp_rows = [r for r in self.exp_rows if r is not fields]
        card.destroy()

    def _save_experience(self):
        roles = []
        for fields in self.exp_rows:
            try:
                roles.append({
                    "title":             fields["title"].get().strip(),
                    "company":           fields["company"].get().strip(),
                    "dates":             fields["dates"].get().strip(),
                    "domain":            fields["domain"].get().strip(),
                    "total_bullets":     int(fields["total_bullets"].get() or 4),
                    "real_bullets":      int(fields["real_bullets"].get() or 2),
                    "fabricated_bullets":int(fields["fabricated_bullets"].get() or 2),
                })
            except Exception:
                pass
        self.config.set("experience_roles", roles)
        self.config.save()

        # Also rebuild the experience prompt with new roles
        self._inject_roles_into_prompt(roles)
        messagebox.showinfo("Saved", f"Saved {len(roles)} work experience roles.")

    def _inject_roles_into_prompt(self, roles: list):
        """Rebuild the experience prompt roles section from saved roles."""
        roles_text = "\n".join([
            f"Role {i+1}: {r['title']} @ {r['company']} | {r['dates']}\n"
            f"→ {r['total_bullets']} bullets: {r['fabricated_bullets']} fabricated + "
            f"{r['real_bullets']} real ({r['domain']} domain)"
            for i, r in enumerate(roles)
        ])

        current_prompt = self.config.get("prompts", "experience_section", default="")

        # Replace the roles section between the markers
        import re as _re
        new_prompt = _re.sub(
            r'(ROLES — KEEP EXACTLY AS SHOWN.*?\n━+\n)(.*?)(━━━\nBULLET RULES)',
            lambda m: m.group(1) + roles_text + "\n\n" + m.group(3),
            current_prompt,
            flags=_re.DOTALL
        )

        if new_prompt != current_prompt:
            self.config.set("prompts", "experience_section", new_prompt)
        else:
            # Fallback — append roles to prompt if markers not found
            self.config.set("experience_roles", roles)

        self.config.save()

    # ── File helpers ──────────────────────────────────────────
    def _load_resume(self):
        path = filedialog.askopenfilename(
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if path:
            with open(path, "r", encoding="utf-8") as f:
                self.resume_box.delete("1.0", "end")
                self.resume_box.insert("1.0", f.read())
            self.config.set("pipeline", "resume_path", path)
            self.config.save()

    def _save_resume(self):
        path = self.config.get("pipeline", "resume_path", default="resume.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.resume_box.get("1.0", "end-1c"))
        messagebox.showinfo("Saved", f"Resume saved → {path}")

    def _load_projects(self):
        path = filedialog.askopenfilename(
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if path:
            with open(path, "r", encoding="utf-8") as f:
                self.projects_box.delete("1.0", "end")
                self.projects_box.insert("1.0", f.read())
            self.config.set("pipeline", "projects_path", path)
            self.config.save()

    def _save_projects(self):
        path = self.config.get("pipeline", "projects_path", default="Projects.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(self.projects_box.get("1.0", "end-1c"))
        messagebox.showinfo("Saved", f"Projects saved → {path}")

    def _clear_resume(self):
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to permanently remove all resume text?"):
            self.resume_box.delete("1.0", "end")
            self._save_resume()
            messagebox.showinfo("Cleared", "Resume text has been cleared and saved.")

    def _add_project_link(self):
        def handle_save(name, link):
            text = f"{name} | {link}\n"
            self.projects_box.insert("end", text)

        ProjectLinkDialog(self, on_save=handle_save)

    def _import_github_project(self):
        from utils.github_importer import GitHubProjectImporter
        from agents.api_client import RotatingOllamaClient

        profile = self.config.get("profile", default={})
        github_url = profile.get("github", "").strip()

        if not github_url:
            messagebox.showwarning("Missing Profile Info", "Please add your GitHub URL in the Profile tab first.")
            return

        def handle_import():
            try:
                model_name = self.config.get("model", "name", default="gemma4:31b-cloud")
                api_keys = self.config.get("model", "api_keys", default=[])

                if not api_keys:
                    messagebox.showerror("Configuration Error", "No API keys found in settings.")
                    return

                import os
                original_key = os.environ.get("OLLAMA_API_KEY", "")
                if api_keys:
                    os.environ["OLLAMA_API_KEY"] = api_keys[0]

                client = RotatingOllamaClient(api_keys=api_keys, model=model_name)

                # Get GitHub token from settings
                github_token = self.config.get("github", "token", default="")
                importer = GitHubProjectImporter(client, github_token=github_token)

                # 1. Get total repo count for progress bar (briefly fetch list)
                username = importer._extract_username(github_url)
                if not username:
                    raise ValueError("Could not extract GitHub username.")

                import requests
                repos_url = f"https://api.github.com/users/{username}/repos"
                params = {"per_page": 100, "type": "owner"}

                # Use token if available for initial count fetch
                headers = {"Accept": "application/vnd.github.v3+json"}
                if github_token:
                    headers["Authorization"] = f"token {github_token}"

                resp = requests.get(repos_url, params=params, headers=headers)
                resp.raise_for_status()
                repos_data = resp.json()
                filtered_repos = [r for r in repos_data if not (r.get("forks_count", 0) > 0 and r.get("fork", False))]
                total = len(filtered_repos)

                if total == 0:
                    messagebox.showwarning("No Projects", "No public repositories found for this account.")
                    os.environ["OLLAMA_API_KEY"] = original_key
                    return

                # 2. Show progress dialog (in main thread)
                progress_dialog = GitHubImportProgressDialog(self, total)

                def worker():
                    try:
                        imported_entries = []
                        count = 0

                        # use the importer's generator
                        for name, entry in importer.import_user_projects(github_url):
                            count += 1
                            if entry:
                                imported_entries.append(entry)

                            # Update UI progress
                            self.after(0, lambda c=count: progress_dialog.update_progress(c))

                        if not imported_entries:
                            self.after(0, lambda: messagebox.showwarning("Import Finished", "No projects were successfully processed."))
                        else:
                            content = "\n\n".join(imported_entries)
                            self.after(0, lambda: self.projects_box.insert("end", f"\n\n{content}\n"))
                            self.after(0, lambda: messagebox.showinfo("Success", f"Imported {len(imported_entries)} projects from GitHub!"))

                    except Exception as e:
                        self.after(0, lambda: messagebox.showerror("Import Error", f"Error during import: {e}"))
                    finally:
                        os.environ["OLLAMA_API_KEY"] = original_key
                        self.after(0, progress_dialog.close)

                import threading
                threading.Thread(target=worker, daemon=True).start()

            except Exception as e:
                messagebox.showerror("Error", f"Initialization failed: {e}")

        handle_import()
# ══════════════════════════════════════════════════════════════
#  MAIN WINDOW
# ══════════════════════════════════════════════════════════════
class JobHunterApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("JobHunter")
        self.geometry("1280x800")
        self.minsize(1000, 650)

        self.config    = Config()
        self.log_queue = queue.Queue()

        # Setup logging
        handler = QueueLogHandler(self.log_queue)
        handler.setFormatter(logging.Formatter("%(name)s — %(message)s"))
        logging.getLogger().addHandler(handler)
        logging.getLogger().setLevel(logging.INFO)

        # Add file logging
        file_handler = logging.FileHandler("ai_logs.log", encoding="utf-8")
        file_handler.setFormatter(logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s"))
        logging.getLogger().addHandler(file_handler)

        self._start_log_monitor()

        # ← ADD THIS
        check_and_run_setup(self, self.config, on_complete=self._build_ui)

    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Sidebar ───────────────────────────────────────────
        sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color="#161b22")
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_rowconfigure(15, weight=1)

        # Logo
        logo = ctk.CTkFrame(sidebar, fg_color="transparent")
        logo.grid(row=0, column=0, padx=16, pady=(20, 8), sticky="ew")
        ctk.CTkLabel(logo, text="JobHunter",
                     font=ctk.CTkFont(size=18, weight="bold"),
                     text_color="#58a6ff").pack(side="left")
        ctk.CTkLabel(logo, text=" v1.0",
                     font=ctk.CTkFont(size=11),
                     text_color="#666").pack(side="left", pady=(4, 0))

        ctk.CTkFrame(sidebar, height=1, fg_color="#21262d").grid(
            row=1, column=0, sticky="ew", padx=12, pady=4)

        nav_items = [
            ("Dashboard",       "🏠"),
            ("Scraper",         "🔍"),
            ("Screener",        "🎯"),
            ("Pipeline",        "⚙"),
            ("Job Tracker",     "📊"),
            ("Applied",         "✅"),
            ("Profile",         "👤"),
            ("Resume",          "📄"),
            ("Resume Order",    "↕"),
            ("Custom Sections", "✨"),
            ("Prompts",         "✏"),
            ("Settings",        "⚙"),
            ("Logs",            "📋"),
        ]

        self.nav_buttons = {}
        for i, (name, icon) in enumerate(nav_items):
            btn = SidebarButton(sidebar, name, icon, lambda n=name: self._show_tab(n))
            btn.grid(row=i+2, column=0, padx=8, pady=2, sticky="ew")
            self.nav_buttons[name] = btn

        sidebar.grid_columnconfigure(0, weight=1)

        # Version info at bottom
        ctk.CTkLabel(sidebar, text="JobHunter AI",
                     font=ctk.CTkFont(size=10), text_color="#444").grid(
            row=16, column=0, padx=16, pady=(0, 8), sticky="sw")

        # ── Content area ──────────────────────────────────────
        self.content = ctk.CTkFrame(self, corner_radius=0, fg_color="#0d1117")
        self.content.grid(row=0, column=1, sticky="nsew")
        self.content.grid_columnconfigure(0, weight=1)
        self.content.grid_rowconfigure(0, weight=1)

        # Tab factories — built lazily on first visit
        self._tab_factories = {
            "Dashboard":       lambda: DashboardTab(self.content, self.config),
            "Scraper":         lambda: ScraperTab(self.content, self.config, self.log_queue),
            "Screener":        lambda: ScreenerConfigTab(self.content, self.config),
            "Pipeline":        lambda: PipelineTab(self.content, self.config, self.log_queue),
            "Job Tracker":     lambda: JobTrackerTab(self.content, self.config),
            "Applied":         lambda: AppliedTab(self.content, self.config),
            "Profile":         lambda: ProfileTab(self.content, self.config),
            "Resume":          lambda: ResumeTab(self.content, self.config),
            "Resume Order":    lambda: ResumeOrderTab(self.content, self.config),
            "Custom Sections": lambda: CustomSectionsTab(self.content, self.config),
            "Prompts":         lambda: PromptsTab(self.content, self.config),
            "Settings":        lambda: SettingsTab(self.content, self.config),
            "Logs":            lambda: LogsTab(self.content),
        }
        self.tabs = {}

        self._show_tab("Dashboard")

    def _show_tab(self, name: str):
        for n, btn in self.nav_buttons.items():
            btn.set_active(n == name)
        # Build tab on first visit
        if name not in self.tabs:
            tab = self._tab_factories[name]()
            tab.grid(row=0, column=0, sticky="nsew")
            self.tabs[name] = tab
        self.tabs[name].tkraise()

    def _start_log_monitor(self):
        """Forward logs from queue to the Logs tab.
        Skips queue entirely while a tab is running — _poll_logs owns the queue then."""
        tabs = getattr(self, 'tabs', {})
        scraper_running  = "Scraper"  in tabs and getattr(tabs["Scraper"],  "running", False)
        pipeline_running = "Pipeline" in tabs and getattr(tabs["Pipeline"], "running", False)

        if not scraper_running and not pipeline_running:
            while not self.log_queue.empty():
                try:
                    kind, level, msg = self.log_queue.get_nowait()
                    if kind == "log" and "Logs" in tabs:
                        tabs["Logs"].append(level, msg)
                except queue.Empty:
                    break

        self.after(500, self._start_log_monitor)


# ── Entry point ───────────────────────────────────────────────
if __name__ == "__main__":
    app = JobHunterApp()
    app.mainloop()