import logging
import traceback
import os
import json
import re
from datetime import date
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from ai import OllamaAssistant, parse_ai_verdict, send_with_retries
from jobs_scraper import *

# --------------------------- Config & Constants ---------------------------

required_columns = [
    "AI_recommendation",
    "site",
    "company",
    "title",
    "location",
    "link",
    "years_required",
    "role_level",
    "skills_match_pct",
    "matched_skills",
    "missing_skills",
    "reasoning",
    "description",
    "posted_date",
]

pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
logging.basicConfig(level=logging.INFO)

excel_file = "jobs.xlsx"

DEFAULT_JOB_FINDER_PROMPT = """
You are a ruthless but fair IT job screener evaluating whether a candidate should apply to a job posting.

-----------------------------------------------------
CANDIDATE PROFILE
-----------------------------------------------------
{candidate_profile}

-----------------------------------------------------
INPUTS
-----------------------------------------------------
JOB TITLE: {title}
JOB DESCRIPTION: {description}
RESUME: {resume_text}

-----------------------------------------------------
EVALUATION RULES (apply in order)
-----------------------------------------------------

STEP 1 — DOMAIN FILTER
Accept ONLY if the role clearly belongs to the candidate's specified domains.
Reject immediately if the role is non-technical or outside the candidate's expertise.
If rejected here — verdict = "no", stop evaluation.

STEP 2 — EXPERIENCE GAP CHECK
Extract the required years of experience from the JD.
Compare against candidate's experience:
- 0 to 4 years required — PASS (good fit range)
- 5 years required — BORDERLINE (check if skills compensate)
- 6+ years required — FAIL (too senior, do not apply)
- Not mentioned — NEUTRAL (continue to skill check)

STEP 3 — SKILLS MATCH SCORING
List every distinct technical skill, tool, framework, or platform mentioned in the JD.
For each one, check if it appears in the resume (directly or through a clearly equivalent technology).
Compute a match percentage: matched_skills / total_jd_skills * 100

Scoring bands:
- 70% or above — Strong match
- 45% to 69% — Partial match
- Below 45% — Weak match

STEP 4 — ROLE LEVEL CALIBRATION
Detect the seniority signal from the JD title and language:
- Junior / Entry-level / Associate / New Grad — BONUS (bump verdict up one level if on the fence)
- Intermediate / Mid-level / no seniority label — NEUTRAL
- Senior / Lead / Staff / Principal / Manager — PENALTY (bump verdict down one level)

STEP 5 — FINAL VERDICT MATRIX

Experience Check | Skills Match | Role Level   | Verdict
PASS             | Strong       | Any          | yes
PASS             | Partial      | Junior/Mid   | yes
PASS             | Partial      | Senior       | maybe
PASS             | Weak         | Junior       | maybe
PASS             | Weak         | Mid/Senior   | no
BORDERLINE       | Strong       | Junior/Mid   | maybe
BORDERLINE       | Partial/Weak | Any          | no
FAIL             | Any          | Any          | no

-----------------------------------------------------
OUTPUT FORMAT — return ONLY a valid JSON object, no extra text, no backticks, no explanation outside the JSON
-----------------------------------------------------

{
  "verdict": "<yes | maybe | no>",
  "years_required": "<exact number, range like 3-5, or unspecified>",
  "role_level": "<junior | mid | senior | unspecified>",
  "skills_match_pct": <integer 0-100>,
  "matched_skills": ["skill1", "skill2"],
  "missing_skills": ["skill1", "skill2"],
  "reasoning": "<2 sentences max — why this verdict, what the main gap or strength is>"
}
"""


# --------------------------- Helpers ---------------------------

def load_env_file(file_path):
    with open(file_path, encoding="utf-8") as f:
        for line in f:
            if line.strip() and not line.startswith("#") and not line.startswith(";"):
                key, value = line.strip().split("=", 1)
                os.environ[key] = value


def load_resume_text(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        logging.warning("Could not load resume text at %s; proceeding with empty resume.", path)
        return ""


def format_prompt(template: str, **kwargs) -> str:
    """Safely substitute placeholders while keeping other braces literal."""
    escaped = template.replace("{", "{{").replace("}", "}}")
    for k, v in kwargs.items():
        escaped = escaped.replace(f"{{{{{k}}}}}", str(v))
    return escaped


def load_df():
    if os.path.exists(excel_file):
        df = pd.read_excel(excel_file, engine="openpyxl")
        for col in required_columns:
            if col not in df.columns:
                df[col] = "" if col != "site" else "Unknown"
        return df
    else:
        return pd.DataFrame(columns=required_columns)


def parse_json_response(response_text: str):
    """
    Uses robust parsing from ai.py and formats for Excel.
    """
    result = parse_ai_verdict(response_text)

    matched_skills = result.get("matched_skills", [])
    if isinstance(matched_skills, list):
        matched_skills = ", ".join(matched_skills)

    missing_skills = result.get("missing_skills", [])
    if isinstance(missing_skills, list):
        missing_skills = ", ".join(missing_skills)

    return {
        "verdict": result["verdict"],
        "years_required": result["years_required"],
        "role_level": result["role_level"],
        "skills_match_pct": result["skills_match_pct"],
        "matched_skills": matched_skills,
        "missing_skills": missing_skills,
        "reasoning": result["reasoning"],
    }


def beautify_excel(path: str = None):
    target_path = path or excel_file
    wb = load_workbook(target_path)
    ws = wb.active

    # Header styling
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    ws.auto_filter.ref = ws.dimensions

    red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

    # Resolve column letters dynamically so column order changes don't break coloring
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).lower()] = cell.column_letter

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            col_letter = cell.column_letter

            # Color verdict column
            if col_letter == header_map.get("ai_recommendation"):
                val = str(cell.value).lower() if cell.value else ""
                if val == "no":
                    cell.fill = red_fill
                elif val == "yes":
                    cell.fill = green_fill
                elif val == "maybe":
                    cell.fill = yellow_fill

            # Color skills_match_pct column with a gradient feel
            if col_letter == header_map.get("skills_match_pct"):
                try:
                    pct = int(cell.value)
                    if pct >= 70:
                        cell.fill = green_fill
                    elif pct >= 45:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = orange_fill
                except (TypeError, ValueError):
                    pass

            # Hyperlink the link column
            if col_letter == header_map.get("link") and cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    # Auto-size columns
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(10, min(60, max_len + 2))

    wb.save(target_path)


def write_excel_safely(df: pd.DataFrame, path: str) -> str:
    try:
        df.to_excel(path, index=False, engine="openpyxl")
        beautify_excel(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        fallback = f"{base}_{date.today().isoformat()}{ext}"
        logging.warning(
            "Permission denied writing %s. Is it open in Excel? Saving to %s instead.",
            path,
            fallback,
        )
        df.to_excel(fallback, index=False, engine="openpyxl")
        beautify_excel(fallback)
        return fallback


# --------------------------- Core Flow ---------------------------

def scrape_and_filter_ai(unique_urls, assistant, instructions, resume_text, prompt_template):
    offset = 0
    new_data = [0]
    df = pd.DataFrame(columns=required_columns)

    while len(new_data) > 0:
        try:
            logging.info(f"Offset of scraping is: {offset}")
            new_data = scrape_all_jobs(
                os.getenv("sites"),
                os.getenv("search_term"),
                os.getenv("location"),
                os.getenv("hours_old"),
                os.getenv("results_wanted"),
                offset,
            )
            logging.info(f"Scraped {len(new_data)} jobs.")
        except Exception as e:
            logging.error("Scraping error: %s\n%s", e, traceback.format_exc())
            new_data = pd.DataFrame(columns=required_columns)

        for _, row in tqdm(new_data.iterrows(), total=len(new_data), desc="Analyzing Jobs"):
            try:
                job_url = row.get("job_url", "")
                if job_url in unique_urls:
                    continue

                msg = format_prompt(
                    prompt_template,
                    title=row.get("title", ""),
                    description=row.get("description", ""),
                    resume_text=resume_text,
                )

                ai_response = send_with_retries(assistant, msg, tries=3, backoff_sec=1.5)
                logging.info("Ollama response received.")

                result = parse_json_response(ai_response)

                new_row = {
                    "AI_recommendation": result["verdict"],
                    "site":              row.get("site", "Unknown"),
                    "company":           row.get("company", ""),
                    "title":             row.get("title", ""),
                    "location":          row.get("location", ""),
                    "link":              job_url,
                    "years_required":    result["years_required"],
                    "role_level":        result["role_level"],
                    "skills_match_pct":  result["skills_match_pct"],
                    "matched_skills":    result["matched_skills"],
                    "missing_skills":    result["missing_skills"],
                    "reasoning":         result["reasoning"],
                    "description":       row.get("description", ""),
                    "posted_date":       row.get("date_posted", ""),
                }
                df.loc[len(df)] = new_row
                unique_urls.add(job_url)

                logging.info(
                    "[%s] %s @ %s | match=%s%% | level=%s | missing=%s",
                    result["verdict"].upper(),
                    row.get("title", ""),
                    row.get("company", ""),
                    result["skills_match_pct"],
                    result["role_level"],
                    result["missing_skills"] or "none",
                )

            except Exception as e:
                logging.error("An error occurred while sending to AI: %s", e)
                logging.error("Stack trace: %s", traceback.format_exc())

        offset += len(new_data)
        break

    return df


def main():
    load_env_file(".env")
    data = load_df()
    unique_urls = set(data["link"])

    try:
        with open("instructions.txt", "r", encoding="utf-8") as file:
            instructions = file.read()
    except Exception:
        instructions = ""

    resume_text = load_resume_text(os.getenv("RESUME_PATH", "instructions.txt"))

    # Load user profile and prepare the AI prompt
    profile = load_profile()
    profile_ctx = build_profile_prompt_context(profile)

    # Load prompt template from config or use default
    try:
        with open("app_config.json", "r", encoding="utf-8") as f:
            cfg = json.load(f)
        prompt_template = cfg.get("prompts", {}).get("job_screener", DEFAULT_JOB_FINDER_PROMPT)
    except Exception:
        prompt_template = DEFAULT_JOB_FINDER_PROMPT

    # Inject candidate profile into the prompt
    prompt_template = prompt_template.replace("{candidate_profile}", profile_ctx)

    assistant = OllamaAssistant(model=os.getenv("model", "gemma4:31b-cloud"))
    logging.info(f"Ollama Assistant ready using model: {assistant.model}")

    new_df = scrape_and_filter_ai(unique_urls, assistant, instructions, resume_text, prompt_template)
    df = pd.concat([data, new_df], ignore_index=True)
    written_path = write_excel_safely(df, excel_file)
    logging.info(f"Excel written to: {written_path}")


if __name__ == "__main__":
    main()